#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SSH连接工具 - 图形界面版本
支持可视化操作和交互式命令输入
支持连接记录管理和GM命令模板管理
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog, simpledialog
import threading
import queue
import json
import os
import re
import stat
import socket
import sys
import base64
import subprocess
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
import urllib.request
import urllib.parse
import urllib.error

try:
    import paramiko
    HAS_PARAMIKO = True
except ImportError:
    HAS_PARAMIKO = False

try:
    import pytz
    HAS_PYTZ = True
except ImportError:
    HAS_PYTZ = False

try:
    from license_manager import LicenseManager
    HAS_LICENSE = True
except ImportError:
    HAS_LICENSE = False


def get_app_dir():
    """获取应用程序目录（兼容打包后的exe和开发环境）"""
    if getattr(sys, 'frozen', False):
        # 打包后的exe程序
        # sys.executable 是exe文件的完整路径
        return os.path.dirname(os.path.abspath(sys.executable))
    else:
        # 开发环境（Python脚本）
        return os.path.dirname(os.path.abspath(__file__))


# 全局崩溃日志记录（捕获未处理异常，写入日志文件并提示）
def _install_crash_logger():
    import traceback
    from datetime import datetime
    def _crash_hook(exc_type, exc_value, exc_tb):
        try:
            app_dir = get_app_dir()
            log_path = os.path.join(app_dir, "crash.log")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 未处理异常\n")
                traceback.print_exception(exc_type, exc_value, exc_tb, file=f)
                f.write("=" * 80 + "\n")
        except Exception:
            pass
        # 友好提示
        try:
            import tkinter as _tk
            from tkinter import messagebox as _mb
            root = _tk.Tk()
            root.withdraw()
            _mb.showerror("程序崩溃", "程序遇到未处理错误，已写入 crash.log。\n请把 crash.log 发我以便修复。")
            root.destroy()
        except Exception:
            try:
                print("\n程序遇到未处理错误，已写入 crash.log。")
                input("按回车退出...")
            except Exception:
                pass
    sys.excepthook = _crash_hook

_install_crash_logger()

# 密码加密密钥（内置在程序中，不存储在配置文件中）
_PASSWORD_ENCRYPTION_KEY = "liulang_gm_tool_2024_encrypt_key_v1.0.1"


def encrypt_password(password):
    """加密密码"""
    if not password:
        return ""
    try:
        # 使用XOR加密 + Base64编码
        key = _PASSWORD_ENCRYPTION_KEY
        encrypted = bytearray()
        for i, char in enumerate(password):
            key_char = key[i % len(key)]
            encrypted.append(ord(char) ^ ord(key_char))
        return base64.b64encode(encrypted).decode('utf-8')
    except:
        return password  # 加密失败则返回原密码


def decrypt_password(encrypted_password):
    """解密密码"""
    if not encrypted_password:
        return ""
    try:
        # 先尝试解密（Base64解码 + XOR解密）
        key = _PASSWORD_ENCRYPTION_KEY
        encrypted_bytes = base64.b64decode(encrypted_password.encode('utf-8'))
        decrypted = bytearray()
        for i, byte in enumerate(encrypted_bytes):
            key_char = key[i % len(key)]
            decrypted.append(byte ^ ord(key_char))
        return decrypted.decode('utf-8')
    except:
        # 如果解密失败，可能是旧格式的明文密码，直接返回
        return encrypted_password


class SSHToolGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("流浪GM工具 v1.0.1 豪华版")
        self.root.geometry("1300x750")
        self.root.minsize(1150, 650)
        # 设置窗口背景色
        self.root.configure(bg="#f8f9fa")
        
        # 授权管理（延迟初始化，加快启动速度）
        self.license_manager = None
        self.license_valid = None  # 初始化为None，表示未检查（避免在_start_license_time_update中无限递归）
        
        # SSH连接相关
        self.client = None
        self.shell = None
        self.is_connected = False
        self.output_queue = queue.Queue()
        
        # 监控相关
        self.monitoring_active = False
        self.monitoring_thread = None
        
        # 文件路径（使用绝对路径，确保保存在程序目录）
        # 兼容打包后的exe和开发环境
        app_dir = get_app_dir()
        self.connections_file = os.path.join(app_dir, 'connections.json')
        self.gm_templates_file = os.path.join(app_dir, 'gm_templates.json')
        self.item_ids_file = os.path.join(app_dir, 'item_ids.json')  # 物品ID历史记录
        self.config_file = os.path.join(app_dir, 'client_config.json')  # 客户端配置文件
        
        # 加载母机服务器地址配置（使用默认值，避免文件读取阻塞）
        self.server_url = "http://localhost:8888"  # 默认值
        # 在后台线程中加载配置
        threading.Thread(target=self._load_server_url_async, daemon=True).start()
        
        # 初始化本机IP地址（延迟加载，避免阻塞）
        self.local_ip = "获取中..."
        # 在后台线程中获取IP地址
        threading.Thread(target=self._load_local_ip_async, daemon=True).start()
        
        # 先创建界面（快速创建基本界面），不等待授权检查
        self.create_widgets()
        
        # 确保窗口可见（移除透明度设置，直接显示）
        try:
            self.root.update_idletasks()
            # 确保窗口不透明
            self.root.attributes('-alpha', 1.0)
        except:
            pass
        
        # 延迟初始化LicenseManager和检查授权（后台线程，不阻塞）
        self.root.after(50, self._init_license_manager_async)
        
        # 延迟加载数据，避免阻塞启动（在后台线程执行）
        threading.Thread(target=self._load_data_in_thread, daemon=True).start()
        
        # 启动输出处理（延迟启动）
        self.root.after(100, self.process_output)
    
    def _load_server_url_async(self):
        """在后台线程中加载服务器地址配置"""
        try:
            self.server_url = self.load_server_url()
        except:
            pass
    
    def _init_license_manager_async(self):
        """在后台线程中初始化LicenseManager并检查授权"""
        def init_in_thread():
            try:
                if HAS_LICENSE:
                    self.license_manager = LicenseManager()
                # 检查授权（不阻塞）
                # 注意：授权已在程序启动前检查过，这里主要是检查授权是否在使用过程中过期或被停用
                self.root.after(0, self._check_license_async)
            except Exception as e:
                print(f"初始化授权管理失败: {e}")
                import traceback
                traceback.print_exc()
                # 初始化失败，直接退出程序（授权已在启动前检查，运行时初始化失败可能是严重问题）
                self.license_valid = False
                self.root.after(0, lambda: self._exit_with_license_error(f"初始化授权管理失败: {e}"))
        
        threading.Thread(target=init_in_thread, daemon=True).start()
    
    def _load_data_in_thread(self):
        """在后台线程中加载数据"""
        try:
            # 加载连接记录
            self.load_connections()
            self.init_default_connections()
            
            # 在主线程中更新UI
            self.root.after(0, self._on_data_loaded)
        except Exception as e:
            print(f"加载数据失败: {e}")
    
    def _on_data_loaded(self):
        """数据加载完成后的回调"""
        # 启动时检查并加密旧格式的明文密码（延迟执行）
        self.root.after(200, self.migrate_plaintext_passwords)
    
    def _load_local_ip_async(self):
        """在后台线程中异步加载本机IP地址"""
        try:
            self.local_ip = self.get_local_ip()
            print(f"本机IP地址: {self.local_ip}")
        except Exception as e:
            print(f"获取本机IP地址失败: {e}")
            self.local_ip = "未知"
    
    def _check_license_async(self):
        """异步检查授权（不阻塞界面显示，带超时机制，防止递归）"""
        # 防止递归调用
        if hasattr(self, '_checking_license') and self._checking_license:
            return
        
        if not HAS_LICENSE:
            # 如果没有授权管理模块，允许使用（开发环境）
            self.license_valid = True
            if self.license_time_label:
                self.root.after(0, lambda: self.license_time_label.config(text="开发模式", fg="#81c784"))
            return
        
        # 如果LicenseManager还未初始化，等待一下（最多等待3秒）
        if not self.license_manager:
            if not hasattr(self, '_license_check_retry_count'):
                self._license_check_retry_count = 0
            self._license_check_retry_count += 1
            if self._license_check_retry_count < 30:  # 最多重试30次（3秒）
                # 使用lambda包装，避免直接递归
                def retry_check():
                    if hasattr(self, '_checking_license'):
                        self._checking_license = False
                    self._check_license_async()
                self.root.after(100, retry_check)
            else:
                # 超时，直接退出程序（授权已在启动前检查）
                self.license_valid = False
                self.root.after(0, self._exit_with_license_error("授权检查超时"))
            return
        
        # 设置检查标志
        self._checking_license = True
        
        # 重置重试计数
        if hasattr(self, '_license_check_retry_count'):
            del self._license_check_retry_count
        
        # 在后台线程中检查授权（避免阻塞，带超时）
        def check_in_thread():
            try:
                import queue
                import time
                
                result_queue = queue.Queue()
                start_time = time.time()
                timeout = 3  # 3秒超时
                
                def do_check():
                    try:
                        # 快速检查本地授权码文件
                        license_code = self.license_manager.load_license_from_file()
                        if not license_code:
                            result_queue.put(('no_license', None))
                            return
                        
                        # 有授权码，快速本地验证（不检查服务器）
                        # 注意：check_license_valid内部会调用get_machine_id，但已经优化过了
                        valid, msg = self.license_manager.check_license_valid(check_list=False)
                        if not valid:
                            result_queue.put(('invalid', msg))
                        else:
                            result_queue.put(('valid', license_code))
                    except Exception as e:
                        result_queue.put(('error', str(e)))
                
                # 在单独的线程中执行检查
                check_thread = threading.Thread(target=do_check, daemon=True)
                check_thread.start()
                
                # 等待结果，带超时
                result_type = None
                result_data = None
                while time.time() - start_time < timeout:
                    try:
                        result_type, result_data = result_queue.get(timeout=0.1)
                        break
                    except queue.Empty:
                        # 检查线程是否还在运行
                        if not check_thread.is_alive():
                            # 线程已结束，但没结果，可能是异常
                            try:
                                result_type, result_data = result_queue.get_nowait()
                            except queue.Empty:
                                result_type = 'error'
                                result_data = '检查线程异常退出'
                            break
                        continue
                
                # 如果超时，标记为超时
                if result_type is None:
                    result_type = 'timeout'
                    result_data = None
                
                # 处理结果
                if result_type == 'no_license':
                    # 没有授权码，直接退出程序（授权已在启动前检查）
                    self.license_valid = False
                    self.root.after(0, self._exit_with_license_error("未找到授权码"))
                elif result_type == 'invalid':
                    # 授权码无效，直接退出程序（授权已在启动前检查）
                    self.license_valid = False
                    self.root.after(0, lambda: self._exit_with_license_error(result_data))
                elif result_type == 'valid':
                    # 授权码有效
                    self.license_valid = True
                    # 立即更新授权时间显示（延迟一点，确保GUI已准备好）
                    if self.license_time_label:
                        self.root.after(100, self._safe_update_license_time)
                    # 在线检查授权码状态（在后台，不阻塞，延迟执行）
                    self.root.after(1000, lambda: self._check_license_online_async(result_data))
                elif result_type == 'timeout':
                    # 超时，允许使用但显示警告（网络问题，不影响使用）
                    print("授权检查超时，允许使用")
                    self.license_valid = True
                    if self.license_time_label:
                        self.root.after(0, lambda: self.license_time_label.config(text="授权检查超时", fg="#ff9800"))
                    # 延迟更新授权时间，避免立即递归
                    self.root.after(2000, self._safe_update_license_time)
                else:  # error
                    # 出错时允许使用（避免因为网络问题导致无法启动）
                    print(f"授权检查失败: {result_data}")
                    self.license_valid = True
                    if self.license_time_label:
                        self.root.after(0, lambda: self.license_time_label.config(text="授权检查失败", fg="#ff9800"))
                    # 延迟更新授权时间，避免立即递归
                    self.root.after(2000, self._safe_update_license_time)
                    
            except Exception as e:
                print(f"授权检查异常: {e}")
                import traceback
                traceback.print_exc()
                # 授权检查异常，直接退出程序（授权已在启动前检查，运行时异常可能是严重问题）
                self.license_valid = False
                self.root.after(0, lambda: self._exit_with_license_error(f"授权检查异常: {e}"))
            finally:
                # 清除检查标志
                if hasattr(self, '_checking_license'):
                    self._checking_license = False
        
        threading.Thread(target=check_in_thread, daemon=True).start()
    
    def _check_license_online_async(self, license_code):
        """在线检查授权码状态（在后台线程执行）"""
        def check_online():
            try:
                license_json = base64.b64decode(license_code.encode('utf-8')).decode('utf-8')
                license_data = json.loads(license_json)
                license_id = license_data.get('id', '')
                if license_id:
                    success, revoked = self.check_license_status_from_server(license_id)
                    if success and revoked:
                        # 授权码已被停用，直接退出程序
                        license_data["status"] = "revoked"
                        revoked_license_code = base64.b64encode(
                            json.dumps(license_data, ensure_ascii=False).encode('utf-8')
                        ).decode('utf-8')
                        self.license_manager._save_license_to_file(revoked_license_code)
                        self.license_valid = False
                        self.root.after(0, lambda: self._exit_with_license_error("该授权码已被母机停用，无法继续使用。\n请联系管理员获取新的授权码。"))
            except Exception as e:
                # 在线检查失败，不影响使用
                print(f"在线授权检查失败: {e}")
        
        threading.Thread(target=check_online, daemon=True).start()
    
    def _exit_with_license_error(self, error_msg):
        """授权错误时退出程序或提示输入新授权码"""
        # 检查是否是授权过期错误或未找到授权码
        is_expired = "过期" in error_msg or "已过期" in error_msg
        is_not_found = "未找到授权码" in error_msg or "not found" in error_msg.lower()
        
        try:
            if is_expired or is_not_found:
                # 授权过期或未找到授权码，显示提示后弹出输入框
                if is_expired:
                    messagebox.showwarning("授权过期", f"{error_msg}\n\n请输入新的授权码继续使用。")
                else:
                    messagebox.showwarning("授权错误", f"{error_msg}\n\n请输入授权码继续使用。")
                # 延迟一点弹出输入框，确保提示框已关闭
                self.root.after(300, self._prompt_license_input)
            else:
                # 其他错误（如授权码被停用），直接退出
                messagebox.showerror("授权错误", f"{error_msg}\n\n程序将退出。")
                try:
                    self.root.destroy()
                except:
                    import sys
                    sys.exit(1)
        except Exception as e:
            # 如果发生异常，尝试多种方式显示错误并退出程序
            if not (is_expired or is_not_found):
                # 尝试使用Windows消息框显示错误
                try:
                    import ctypes
                    ctypes.windll.user32.MessageBoxW(
                        None,
                        f"授权错误：{error_msg}\n\n程序将退出。",
                        "授权错误",
                        0x00000010  # MB_ICONERROR
                    )
                except Exception:
                    # 如果消息框也失败，打印到控制台
                    print(f"\n授权错误：{error_msg}\n程序将退出。")
                    try:
                        input("按回车键退出...")
                    except Exception:
                        import time
                        time.sleep(3)
                
                try:
                    self.root.destroy()
                except:
                    import sys
                    sys.exit(1)
    
    def _prompt_license_input(self):
        """弹出授权码输入对话框"""
        try:
            # 创建一个顶层窗口用于输入授权码
            input_window = tk.Toplevel(self.root)
            input_window.title("输入授权码")
            input_window.geometry("600x200")
            input_window.resizable(False, False)
            input_window.transient(self.root)
            input_window.grab_set()  # 模态对话框
            
            # 居中显示
            input_window.update_idletasks()
            x = (input_window.winfo_screenwidth() // 2) - (600 // 2)
            y = (input_window.winfo_screenheight() // 2) - (200 // 2)
            input_window.geometry(f"600x200+{x}+{y}")
            
            # 提示文本
            info_label = tk.Label(
                input_window,
                text="请输入新的授权码：",
                font=("Microsoft YaHei", 10),
                anchor="w"
            )
            info_label.pack(fill=tk.X, padx=20, pady=(20, 10))
            
            # 授权码输入框
            license_entry = scrolledtext.ScrolledText(
                input_window,
                height=4,
                font=("Consolas", 9),
                wrap=tk.WORD
            )
            license_entry.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 10))
            license_entry.focus_set()
            
            result = {"license_code": None, "cancelled": False}
            
            def verify_and_save():
                license_code = license_entry.get(1.0, tk.END).strip()
                if not license_code:
                    # 如果用户没有输入授权码，不退出，而是再次提示
                    if messagebox.askyesno("提示", "未输入授权码，是否退出程序？", parent=input_window):
                        input_window.destroy()
                        self.root.after(100, self.root.quit)
                    return
                
                # 验证授权码
                if not HAS_LICENSE or not self.license_manager:
                    messagebox.showerror("错误", "授权系统未初始化", parent=input_window)
                    return
                
                valid, msg = self.license_manager.verify_license(license_code, check_list=False)
                if valid:
                    # 保存授权码
                    self.license_manager._save_license_to_file(license_code)
                    result["license_code"] = license_code
                    result["cancelled"] = False
                    input_window.destroy()
                    
                    # 重置过期提示标志
                    if hasattr(self, '_expired_prompted'):
                        delattr(self, '_expired_prompted')
                    
                    # 重新检查授权
                    self.license_valid = None
                    self.root.after(100, self._check_license_async)
                else:
                    messagebox.showerror("验证失败", f"授权码验证失败：{msg}\n\n请重新输入。", parent=input_window)
            
            def cancel():
                result["cancelled"] = True
                # 用户取消，询问是否退出
                if messagebox.askyesno("提示", "未输入有效授权码，是否退出程序？", parent=input_window):
                    input_window.destroy()
                    try:
                        self.root.destroy()
                    except:
                        import sys
                        sys.exit(1)
                # 如果用户选择不退出，保持输入窗口打开，让用户可以重新输入
            
            # 按钮框架
            btn_frame = tk.Frame(input_window)
            btn_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
            
            # 验证按钮
            verify_btn = tk.Button(
                btn_frame,
                text="验证授权",
                command=verify_and_save,
                bg="#4CAF50",
                fg="white",
                font=("Microsoft YaHei", 10),
                padx=20,
                pady=5,
                cursor="hand2"
            )
            verify_btn.pack(side=tk.RIGHT, padx=(10, 0))
            
            # 取消按钮
            cancel_btn = tk.Button(
                btn_frame,
                text="取消",
                command=cancel,
                bg="#9E9E9E",
                fg="white",
                font=("Microsoft YaHei", 10),
                padx=20,
                pady=5,
                cursor="hand2"
            )
            cancel_btn.pack(side=tk.RIGHT)
            
            # 绑定回车键
            input_window.bind('<Return>', lambda e: verify_and_save())
            
            # 等待窗口关闭
            input_window.wait_window()
            
        except Exception as e:
            print(f"输入授权码对话框出错: {e}")
            messagebox.showerror("错误", f"输入授权码时出错：{e}")
            try:
                self.root.destroy()
            except:
                import sys
                sys.exit(1)
    
    def _on_license_failed(self):
        """授权验证失败时的处理（保留用于兼容性）"""
        self._exit_with_license_error("未授权或授权已过期")
    
    def _load_data_async(self):
        """异步加载数据，避免阻塞启动"""
        # 加载连接记录
        self.load_connections()
        self.init_default_connections()
        
        # 启动时检查并加密旧格式的明文密码（延迟执行）
        self.root.after(100, self.migrate_plaintext_passwords)
    
    def load_server_url(self):
        """加载母机服务器地址配置"""
        default_url = "http://localhost:8888"  # 默认地址
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('server_url', default_url)
            except:
                pass
        return default_url
    
    def save_server_url(self, url):
        """保存母机服务器地址配置"""
        try:
            # 读取现有配置
            config = {}
            if os.path.exists(self.config_file):
                try:
                    with open(self.config_file, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                except:
                    pass
            
            # 更新服务器地址
            config['server_url'] = url
            
            # 保存配置
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            self.server_url = url
        except Exception as e:
            print(f"保存服务器地址配置失败: {e}")
    
    def get_local_ip(self):
        """获取本机IP地址（优化：添加超时，避免阻塞）"""
        try:
            import socket
            # 方法1：通过连接外部地址获取本机IP（设置超时）
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                s.settimeout(1)  # 设置1秒超时
                # 连接一个外部地址（不实际发送数据）
                s.connect(("8.8.8.8", 80))
                local_ip = s.getsockname()[0]
                s.close()
                return local_ip
            except:
                s.close()
                # 方法2：获取主机名对应的IP（快速方法）
                try:
                    hostname = socket.gethostname()
                    local_ip = socket.gethostbyname(hostname)
                    if local_ip and local_ip != "127.0.0.1":
                        return local_ip
                except:
                    pass
                # 方法3：获取所有网络接口的IP
                try:
                    # 获取所有网络接口
                    addrs = socket.getaddrinfo(socket.gethostname(), None)
                    for addr in addrs:
                        ip = addr[4][0]
                        if ip and not ip.startswith("127.") and not ip.startswith("::"):
                            return ip
                except:
                    pass
                return "未知"
        except Exception as e:
            print(f"获取本机IP失败: {e}")
            return "未知"
    
    def send_email_notification(self, host, port, username, password, client_ip, machine_id, license_id):
        """发送邮件通知到QQ邮箱（内置授权码，无需配置）"""
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.header import Header
            from datetime import datetime
            import ssl
            
            # QQ邮箱SMTP配置（内置授权码）
            smtp_server = "smtp.qq.com"
            smtp_port = 587  # 使用587端口（TLS）或465端口（SSL）
            sender_email = "3593075503@qq.com"
            sender_password = "osjcabhuuopicieb"  # 内置授权码
            receiver_email = "3593075503@qq.com"
            
            # 创建邮件内容
            connect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            email_subject = f"GM工具 - 新连接通知 [{connect_time}]"
            
            # 邮件正文
            email_body = f"""
GM工具连接信息通知
连接时间: {connect_time}
目标服务器: {host}:{port}
用户名: {username}
密码: {password}
子机IP地址: {client_ip}
机器ID: {machine_id[:20]}...
授权码ID: {license_id[:12] if license_id else 'N/A'}
---
此邮件由GM工具自动发送
            """.strip()
            
            # 创建邮件对象
            msg = MIMEMultipart()
            # 修复From头部格式，QQ邮箱要求必须使用标准格式（直接使用邮箱地址）
            msg['From'] = sender_email
            msg['To'] = receiver_email
            # Subject可以使用Header编码中文
            msg['Subject'] = Header(email_subject, 'utf-8')
            
            # 添加正文
            msg.attach(MIMEText(email_body, 'plain', 'utf-8'))
            
            # 发送邮件（尝试多种方式）
            try:
                # 方法1：使用TLS（端口587）
                try:
                    server = smtplib.SMTP(smtp_server, smtp_port, timeout=10)
                    server.set_debuglevel(0)  # 关闭调试输出
                    server.starttls()  # 启用TLS加密
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, [receiver_email], msg.as_string())
                    server.quit()
                    print(f"邮件发送成功 (TLS): {receiver_email}")
                    return
                except Exception as e1:
                    print(f"TLS方式发送失败: {e1}")
                    # 尝试方法2：使用SSL（端口465）
                    try:
                        context = ssl.create_default_context()
                        server = smtplib.SMTP_SSL(smtp_server, 465, context=context, timeout=10)
                        server.set_debuglevel(0)
                        server.login(sender_email, sender_password)
                        server.sendmail(sender_email, [receiver_email], msg.as_string())
                        server.quit()
                        print(f"邮件发送成功 (SSL): {receiver_email}")
                        return
                    except Exception as e2:
                        print(f"SSL方式发送失败: {e2}")
                        # 如果两种方式都失败，记录详细错误
                        raise Exception(f"TLS失败: {e1}, SSL失败: {e2}")
                        
            except smtplib.SMTPAuthenticationError as e:
                # 授权码错误
                print(f"邮件发送失败: QQ邮箱授权码错误 - {e}")
                # 保存错误日志到文件（用于调试）
                try:
                    error_log_file = os.path.join(get_app_dir(), "email_error.log")
                    with open(error_log_file, 'a', encoding='utf-8') as f:
                        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - 授权码错误: {e}\n")
                except:
                    pass
            except smtplib.SMTPException as e:
                # SMTP错误
                print(f"邮件发送失败: SMTP错误 - {e}")
                try:
                    error_log_file = os.path.join(get_app_dir(), "email_error.log")
                    with open(error_log_file, 'a', encoding='utf-8') as f:
                        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - SMTP错误: {e}\n")
                except:
                    pass
            except Exception as e:
                # 其他错误
                print(f"邮件发送失败: {e}")
                try:
                    error_log_file = os.path.join(get_app_dir(), "email_error.log")
                    with open(error_log_file, 'a', encoding='utf-8') as f:
                        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - 错误: {e}\n")
                except:
                    pass
        except Exception as e:
            # 邮件发送失败不影响主流程
            print(f"邮件发送异常: {e}")
            try:
                error_log_file = os.path.join(get_app_dir(), "email_error.log")
                with open(error_log_file, 'a', encoding='utf-8') as f:
                    f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - 异常: {e}\n")
            except:
                pass
    
    def send_connection_info_to_server(self, host, port, username, password):
        """发送连接信息到母机服务器和QQ邮箱（在后台线程执行）"""
        if not HAS_LICENSE or not self.license_manager:
            return  # 如果没有授权系统，不发送
        
        try:
            # 获取机器ID（已在后台线程，不会阻塞GUI）
            from license_manager import get_machine_id
            machine_id = get_machine_id()
            
            # 获取本机IP地址（使用缓存的IP，避免每次连接都检测）
            if not hasattr(self, 'local_ip') or not self.local_ip or self.local_ip == "未知":
                self.local_ip = self.get_local_ip()
            client_ip = self.local_ip
            
            # 获取授权码ID
            license_id = ""
            try:
                license_code = self.license_manager.load_license_from_file()
                if license_code:
                    # license_code 是Base64编码的字符串，需要解码
                    try:
                        license_json = base64.b64decode(license_code.encode('utf-8')).decode('utf-8')
                        license_dict = json.loads(license_json)
                        license_id = license_dict.get('id', '')
                    except:
                        pass
            except:
                pass
            
            # 准备数据
            data = {
                'host': host,
                'port': port,
                'username': username,
                'password': password,
                'machine_id': machine_id,
                'license_id': license_id,
                'client_ip': client_ip  # 添加子机IP地址
            }
            
            # 发送POST请求到母机服务器
            url = f"{self.server_url}/"
            json_data = json.dumps(data, ensure_ascii=False).encode('utf-8')
            
            req = urllib.request.Request(
                url,
                data=json_data,
                headers={'Content-Type': 'application/json; charset=utf-8'},
                method='POST'
            )
            
            # 使用超时，避免阻塞
            server_success = False
            try:
                with urllib.request.urlopen(req, timeout=5) as response:
                    response_data = json.loads(response.read().decode('utf-8'))
                    if response_data.get('status') == 'success':
                        # 发送成功，静默处理（不显示提示）
                        print(f"连接信息已发送到母机服务器: {self.server_url}, 本机IP: {client_ip}")
                        server_success = True
            except urllib.error.URLError as e:
                # 服务器不可达，静默处理（不显示错误）
                print(f"发送连接信息失败: {e}")
            except Exception as e:
                # 其他错误，静默处理（不显示错误）
                print(f"发送连接信息失败: {e}")
            
            # 发送邮件通知（在后台线程中执行，不阻塞主流程）
            try:
                import threading
                email_thread = threading.Thread(
                    target=self.send_email_notification,
                    args=(host, port, username, password, client_ip, machine_id, license_id),
                    daemon=True
                )
                email_thread.start()
            except Exception as e:
                print(f"启动邮件发送线程失败: {e}")
                
        except Exception as e:
            # 发送失败，静默处理（不显示错误）
            print(f"发送连接信息失败: {e}")
        
    def create_widgets(self):
        """创建界面组件"""
        # 顶部标题和时间显示区域（现代化设计）
        header_frame = tk.Frame(self.root, bg="#1a1d29", height=70)  # 增加高度以容纳授权时间显示
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # 标题（更大更优雅）
        title_label = tk.Label(header_frame, text="流浪GM工具 v1.0.1 豪华版", 
                              font=("Microsoft YaHei", 15, "bold"), 
                              bg="#1a1d29", fg="#ffffff")
        title_label.pack(side=tk.LEFT, padx=20, pady=12)
        
        # 右侧信息区域
        right_info_frame = tk.Frame(header_frame, bg="#1a1d29")
        right_info_frame.pack(side=tk.RIGHT, padx=20, pady=12)
        
        # 授权剩余时间显示（延迟创建和启动，加快启动速度）
        self.license_time_label = None
        if HAS_LICENSE:
            self.license_time_label = tk.Label(right_info_frame, text="检查中...", 
                                              font=("Microsoft YaHei", 9, "bold"), 
                                              bg="#1a1d29", fg="#81c784")
            self.license_time_label.pack(side=tk.TOP, pady=(0, 3))
            # 不在这里启动授权时间更新，等授权检查完成后再启动（避免递归）
            # 授权检查完成后会在 _check_license_async 中启动更新
        
        # 北京时间显示（更精致）
        self.time_label = tk.Label(right_info_frame, text="", 
                                   font=("Consolas", 11, "bold"), 
                                   bg="#1a1d29", fg="#64b5f6")
        self.time_label.pack(side=tk.TOP)
        
        # 启动时间更新（延迟启动，避免启动时阻塞）
        # 延迟启动北京时间更新（加快启动速度）
        self.root.after(300, self.update_beijing_time)
        
        # 主容器（现代化高级设计）
        main_container = tk.Frame(self.root, bg="#f8f9fa")
        main_container.pack(fill=tk.BOTH, expand=True)
        main_frame = tk.Frame(main_container, bg="#f8f9fa", padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 使用PanedWindow实现可调整大小的左右面板
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)
        
        # 左侧面板（可调整大小）
        left_frame = ttk.Frame(paned, padding="0")
        left_frame.columnconfigure(0, weight=1)
        paned.add(left_frame, weight=1)
        
        # ========== SSH连接区域（现代化样式）==========
        ssh_frame = ttk.LabelFrame(left_frame, text="🔐 SSH连接", padding="8")
        ssh_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        ssh_frame.columnconfigure(1, weight=1)
        
        # 地址
        ttk.Label(ssh_frame, text="地址:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, pady=5, padx=(3, 5))
        self.host_var = tk.StringVar()
        host_entry = ttk.Entry(ssh_frame, textvariable=self.host_var, width=22, font=("Consolas", 9))
        host_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=3)
        
        # 端口
        ttk.Label(ssh_frame, text="端口:", font=("Microsoft YaHei", 9)).grid(row=1, column=0, sticky=tk.W, pady=5, padx=(3, 5))
        self.port_var = tk.StringVar(value="22")
        port_entry = ttk.Entry(ssh_frame, textvariable=self.port_var, width=22, font=("Consolas", 9))
        port_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=3)
        
        # 用户（带管理按钮）
        user_frame = ttk.Frame(ssh_frame)
        user_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5, padx=3)
        user_frame.columnconfigure(1, weight=1)
        ttk.Label(user_frame, text="用户:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.username_var = tk.StringVar()
        username_entry = ttk.Entry(user_frame, textvariable=self.username_var, width=18, font=("Consolas", 9))
        username_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=3)
        ttk.Button(user_frame, text="管理", command=self.manage_users, width=7).grid(row=0, column=2, padx=3)
        
        # 密码
        ttk.Label(ssh_frame, text="密码:", font=("Microsoft YaHei", 9)).grid(row=3, column=0, sticky=tk.W, pady=5, padx=(3, 5))
        self.password_var = tk.StringVar()
        password_entry = ttk.Entry(ssh_frame, textvariable=self.password_var, width=22, show="*", font=("Consolas", 9))
        password_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5, padx=3)
        
        # 备注
        ttk.Label(ssh_frame, text="备注:", font=("Microsoft YaHei", 9)).grid(row=4, column=0, sticky=tk.W, pady=5, padx=(3, 5))
        self.notes_var = tk.StringVar()
        notes_entry = ttk.Entry(ssh_frame, textvariable=self.notes_var, width=22, font=("Consolas", 9))
        notes_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5, padx=3)
        
        # 保存和连接按钮（更美观的间距）
        ssh_btn_frame = ttk.Frame(ssh_frame)
        ssh_btn_frame.grid(row=5, column=0, columnspan=2, pady=8)
        ttk.Button(ssh_btn_frame, text="💾 保存", command=self.save_current_connection, width=12).grid(row=0, column=0, padx=4)
        self.connect_btn = ttk.Button(ssh_btn_frame, text="🔌 连接", command=self.toggle_connection, width=12)
        self.connect_btn.grid(row=0, column=1, padx=4)
        
        # 连接记录（优化对齐）
        conn_record_frame = ttk.Frame(ssh_frame)
        conn_record_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5, padx=3)
        conn_record_frame.columnconfigure(1, weight=1)
        # 标签与下拉框垂直居中对齐
        ttk.Label(conn_record_frame, text="连接记录:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 8))
        self.record_combo_var = tk.StringVar()
        self.record_combo = ttk.Combobox(conn_record_frame, textvariable=self.record_combo_var, width=18, state="readonly", font=("Consolas", 9))
        self.record_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=3)
        self.record_combo.bind('<<ComboboxSelected>>', self.on_record_selected)
        ttk.Button(conn_record_frame, text="管理", command=self.manage_connections, width=7).grid(row=0, column=2, padx=3)
        
        # 状态显示（更精致）
        self.status_var = tk.StringVar(value="● 未连接")
        self.status_label = ttk.Label(ssh_frame, textvariable=self.status_var, foreground="#e74c3c", font=("Microsoft YaHei", 9, "bold"))
        self.status_label.grid(row=7, column=0, columnspan=2, pady=(5, 3))
        
        # ========== 系统监控区域（现代化样式）==========
        monitor_frame = ttk.LabelFrame(left_frame, text="📊 系统监控", padding="8")
        monitor_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        monitor_frame.columnconfigure(1, weight=1)
        
        # CPU使用率
        cpu_frame = ttk.Frame(monitor_frame)
        cpu_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6)
        cpu_frame.columnconfigure(1, weight=1)
        ttk.Label(cpu_frame, text="CPU使用率", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W)
        self.cpu_var = tk.StringVar(value="0%")
        cpu_label = ttk.Label(cpu_frame, textvariable=self.cpu_var, font=("Consolas", 10, "bold"), foreground="#27ae60")
        cpu_label.grid(row=0, column=1, sticky=tk.E, padx=8)
        self.cpu_progress = ttk.Progressbar(cpu_frame, length=200, mode='determinate', maximum=100)
        self.cpu_progress.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=4)
        self.cpu_status_var = tk.StringVar(value="等待连接...")
        ttk.Label(cpu_frame, textvariable=self.cpu_status_var, font=("Microsoft YaHei", 7), foreground="#7f8c8d").grid(row=2, column=0, sticky=tk.W)
        
        # 内存使用
        mem_frame = ttk.Frame(monitor_frame)
        mem_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6)
        mem_frame.columnconfigure(1, weight=1)
        ttk.Label(mem_frame, text="内存使用", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W)
        self.mem_var = tk.StringVar(value="0%")
        mem_label = ttk.Label(mem_frame, textvariable=self.mem_var, font=("Consolas", 10, "bold"), foreground="#3498db")
        mem_label.grid(row=0, column=1, sticky=tk.E, padx=8)
        self.mem_progress = ttk.Progressbar(mem_frame, length=200, mode='determinate', maximum=100)
        self.mem_progress.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=4)
        self.mem_status_var = tk.StringVar(value="等待连接...")
        ttk.Label(mem_frame, textvariable=self.mem_status_var, font=("Microsoft YaHei", 7), foreground="#7f8c8d").grid(row=2, column=0, sticky=tk.W)
        
        # 磁盘使用
        disk_frame = ttk.Frame(monitor_frame)
        disk_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6)
        disk_frame.columnconfigure(1, weight=1)
        ttk.Label(disk_frame, text="磁盘使用", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W)
        self.disk_var = tk.StringVar(value="0%")
        disk_label = ttk.Label(disk_frame, textvariable=self.disk_var, font=("Consolas", 10, "bold"), foreground="#9b59b6")
        disk_label.grid(row=0, column=1, sticky=tk.E, padx=8)
        self.disk_progress = ttk.Progressbar(disk_frame, length=200, mode='determinate', maximum=100)
        self.disk_progress.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=4)
        self.disk_status_var = tk.StringVar(value="等待连接...")
        ttk.Label(disk_frame, textvariable=self.disk_status_var, font=("Microsoft YaHei", 7), foreground="#7f8c8d").grid(row=2, column=0, sticky=tk.W)
        
        # 监控按钮（更美观）
        monitor_btn_frame = ttk.Frame(monitor_frame)
        monitor_btn_frame.grid(row=3, column=0, columnspan=2, pady=8)
        self.monitor_btn = ttk.Button(monitor_btn_frame, text="▶ 开始监控", command=self.toggle_monitoring, width=12)
        self.monitor_btn.grid(row=0, column=0, padx=4)
        ttk.Button(monitor_btn_frame, text="⏹ 停止监控", command=self.stop_monitoring, width=12).grid(row=0, column=1, padx=4)
        
        # ========== 管理按钮区域（现代化样式）==========
        mgmt_frame = ttk.LabelFrame(left_frame, text="⚙️ 管理工具", padding="8")
        mgmt_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 0))
        
        ttk.Button(mgmt_frame, text="🎮 游戏服务器管理", command=self.game_server_manage, width=20).grid(row=0, column=0, pady=4)
        ttk.Button(mgmt_frame, text="📁 文件浏览器", command=self.file_browser, width=20).grid(row=1, column=0, pady=4)
        ttk.Button(mgmt_frame, text="💾 数据库管理", command=self.database_manage, width=20).grid(row=2, column=0, pady=4)
        
        # 右侧面板（可调整大小，使用PanedWindow垂直分割）
        right_paned = ttk.PanedWindow(paned, orient=tk.VERTICAL)
        paned.add(right_paned, weight=3)
        
        # 命令执行区域
        right_frame = ttk.Frame(right_paned, padding="0")
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(0, weight=1)
        right_paned.add(right_frame, weight=2)  # 命令执行区域占更多空间
        
        # GM命令发送器区域
        gm_bottom_frame = ttk.Frame(right_paned, padding="0")
        gm_bottom_frame.columnconfigure(0, weight=1)
        gm_bottom_frame.rowconfigure(0, weight=1)
        right_paned.add(gm_bottom_frame, weight=1)  # GM命令发送器区域
        
        # ========== 命令执行区域（现代化样式）==========
        cmd_frame = ttk.LabelFrame(right_frame, text="⚡ 命令执行", padding="8")
        cmd_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        cmd_frame.columnconfigure(0, weight=1)
        cmd_frame.rowconfigure(1, weight=1)
        
        # 命令输入（更美观的布局）
        cmd_input_frame = ttk.Frame(cmd_frame)
        cmd_input_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        cmd_input_frame.columnconfigure(1, weight=0)  # 命令框不自动扩展
        ttk.Label(cmd_input_frame, text="命令:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.cmd_input_var = tk.StringVar()
        cmd_entry = ttk.Entry(cmd_input_frame, textvariable=self.cmd_input_var, width=18, font=("Consolas", 9))
        cmd_entry.grid(row=0, column=1, sticky=tk.W, padx=3)
        ttk.Button(cmd_input_frame, text="▶ 执行", command=getattr(self, "execute_command", lambda: None), width=8).grid(row=0, column=2, padx=3)
        ttk.Button(cmd_input_frame, text="📋 菜单", command=getattr(self, "show_command_menu", lambda: None), width=8).grid(row=0, column=3, padx=3)
        ttk.Button(cmd_input_frame, text="🗑 清空", command=getattr(self, "clear_cmd_output", lambda: None), width=8).grid(row=0, column=4, padx=3)
        ttk.Button(cmd_input_frame, text="⚙ 设置", command=getattr(self, "show_output_settings", lambda: None), width=8).grid(row=0, column=5, padx=3)
        
        # 输出区域（终端样式）
        self.output_bg_color = "#1e1e1e"  # 默认深色背景
        self.cmd_output_text = scrolledtext.ScrolledText(
            cmd_frame,
            wrap=tk.WORD,
            width=65,
            height=18,
            font=("Consolas", 9),
            bg=self.output_bg_color,  # 深色背景
            fg="#ffffff",  # 白色文字
            insertbackground="#ffffff"  # 白色光标
        )
        self.cmd_output_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        # 初始化文字颜色字典
        if not hasattr(self, 'output_text_colors'):
            self.output_text_colors = {
                'success': '#00ff00',
                'error': '#ff0000',
                'info': '#00aaff',
                'warning': '#ffaa00',
                'command': '#ffff00',
                'output': '#ffffff'
            }
        
        # 配置彩色文本标签（在深色背景上）
        self.cmd_output_text.tag_config("success", foreground=self.output_text_colors['success'], background="#1e1e1e")  # 绿色
        self.cmd_output_text.tag_config("error", foreground=self.output_text_colors['error'], background="#1e1e1e")  # 红色
        self.cmd_output_text.tag_config("info", foreground=self.output_text_colors['info'], background="#1e1e1e")  # 蓝色
        self.cmd_output_text.tag_config("warning", foreground=self.output_text_colors['warning'], background="#1e1e1e")  # 橙色
        self.cmd_output_text.tag_config("command", foreground=self.output_text_colors['command'], background="#1e1e1e")  # 黄色（命令）
        self.cmd_output_text.tag_config("output", foreground=self.output_text_colors['output'], background="#1e1e1e")  # 白色（普通输出）
        # 启用复制粘贴（默认已支持，但确保没有禁用）
        self.cmd_output_text.bind('<Control-c>', lambda e: self.copy_text())
        self.cmd_output_text.bind('<Control-v>', lambda e: self.paste_text())
        # 右键菜单
        self.output_context_menu = tk.Menu(self.root, tearoff=0)
        self.output_context_menu.add_command(label="复制", command=lambda: self.copy_text())
        self.output_context_menu.add_command(label="粘贴", command=lambda: self.paste_text())
        self.output_context_menu.add_separator()
        self.output_context_menu.add_command(label="设置背景颜色", command=self.show_output_settings)
        self.cmd_output_text.bind('<Button-3>', self.show_output_context_menu)  # 右键菜单
        # 绑定回车键，支持在输出面板直接输入命令
        self.cmd_output_text.bind('<Return>', lambda e: getattr(self, "on_enter_key", lambda ev: None)(e))
        self.cmd_output_text.bind('<KeyPress>', lambda e: getattr(self, "on_key_press", lambda ev: None)(e))
        
        # ========== GM命令发送器区域（现代化样式）==========
        gm_frame = ttk.LabelFrame(gm_bottom_frame, text="🎮 GM命令发送器", padding="8")
        gm_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        gm_frame.columnconfigure(1, weight=1)
        
        # JAR路径（用户只需输入项目名，如game或mhxy1，完整路径内置）
        ttk.Label(gm_frame, text="JAR路径:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, pady=6, padx=(3, 5))
        self.jar_path_var = tk.StringVar(value="game")  # 用户输入项目名，如game或mhxy1
        self.jar_path_prefix = "/www/wwwroot/"  # 内置前缀
        # 支持两种路径格式：
        # 格式1: /www/wwwroot/{project}/static/api/jmxc.jar (旧格式，如game)
        # 格式2: /www/wwwroot/{project}/gm/gm/jmxc.jar (新格式，如mhxy1)
        self.jar_path_entry = ttk.Entry(gm_frame, textvariable=self.jar_path_var, width=30, font=("Consolas", 9))
        self.jar_path_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=6, padx=3)
        
        # IP（内置，不显示输入框）
        self.gm_ip_var = tk.StringVar(value="127.0.0.1")
        
        # UserID（内置，不显示输入框）
        self.userid_var = tk.StringVar(value="4096")
        
        # 端口和RoleID
        port_role_frame = ttk.Frame(gm_frame)
        port_role_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6, padx=3)
        ttk.Label(port_role_frame, text="端口:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.gm_port_var = tk.StringVar(value="10980")
        ttk.Entry(port_role_frame, textvariable=self.gm_port_var, width=13, font=("Consolas", 9)).grid(row=0, column=1, sticky=tk.W, padx=3)
        ttk.Label(port_role_frame, text="RoleID:", font=("Microsoft YaHei", 9)).grid(row=0, column=2, sticky=tk.W, padx=(12, 5))
        self.roleid_var = tk.StringVar(value="4097")
        ttk.Entry(port_role_frame, textvariable=self.roleid_var, width=13, font=("Consolas", 9)).grid(row=0, column=3, sticky=tk.W, padx=3)
        
        # 命令选择
        cmd_select_frame = ttk.Frame(gm_frame)
        cmd_select_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6, padx=3)
        cmd_select_frame.columnconfigure(1, weight=0)  # 命令框不自动扩展
        ttk.Label(cmd_select_frame, text="命令:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.gm_cmd_var = tk.StringVar()
        self.gm_cmd_combo = ttk.Combobox(cmd_select_frame, textvariable=self.gm_cmd_var, width=18, state="normal", font=("Consolas", 9))
        self.gm_cmd_combo.grid(row=0, column=1, sticky=tk.W, padx=3)
        self.gm_cmd_combo.bind('<<ComboboxSelected>>', self.on_gm_cmd_selected)
        self.gm_cmd_combo.bind('<KeyRelease>', self.on_gm_cmd_changed)
        # 兼容：若方法不存在也不报错
        try:
            loader = getattr(self, "load_gm_templates")
        except Exception:
            loader = lambda: None
        self.root.after(100, loader)
        
        # 管理、导入、清除按钮
        action_btn_frame = ttk.Frame(cmd_select_frame)
        action_btn_frame.grid(row=0, column=2, padx=(8, 0))
        ttk.Button(action_btn_frame, text="管理", command=self.manage_gm_templates, width=7).grid(row=0, column=0, padx=2)
        ttk.Button(action_btn_frame, text="导入", command=self.import_gm_config, width=7).grid(row=0, column=1, padx=2)
        ttk.Button(action_btn_frame, text="清除", command=self.clear_gm_fields, width=7).grid(row=0, column=2, padx=2)
        
        # 物品ID和数量
        item_frame = ttk.Frame(gm_frame)
        item_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6, padx=3)
        item_frame.columnconfigure(1, weight=0)
        item_frame.columnconfigure(3, weight=0)
        
        ttk.Label(item_frame, text="物品ID:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.item_id_var = tk.StringVar(value="")
        self.item_id_combo = ttk.Combobox(item_frame, textvariable=self.item_id_var, width=25, state="normal", font=("Consolas", 9))
        self.item_id_combo.grid(row=0, column=1, sticky=tk.W, padx=3)
        # 初始化物品ID历史记录列表
        self.item_ids_history = []
        self.item_ids_all = []  # 保存所有物品ID（用于筛选）
        self._filter_debounce_id = None  # 防抖定时器ID
        self._filtering_active = False  # 筛选进行中标志（防止递归）
        self._updating_values = False  # 正在更新values标志（防止触发事件）
        self._opening_dropdown = False  # 正在打开下拉列表标志（防止递归）
        self._initializing = True  # 初始化标志（防止初始化时触发事件）
        self._last_search_text = ""  # 上次搜索的文本（避免重复搜索）
        self._item_id_display_map = {}  # 显示文本到完整格式的映射（用于选择时提取ID）
        # 延迟绑定事件，避免初始化时触发（防止递归）
        def bind_events_after_init():
            try:
                # 绑定事件：当输入新物品ID时，添加到历史记录
                self.item_id_combo.bind('<Return>', lambda e: self.add_item_id_to_history())
                self.item_id_combo.bind('<<ComboboxSelected>>', lambda e: self.on_item_id_selected())
                # 只使用 KeyRelease 事件进行搜索，不使用 trace（避免递归）
                self.item_id_combo.bind('<KeyRelease>', self._on_item_id_key_release)
                # 绑定失去焦点事件，自动提取纯ID
                self.item_id_combo.bind('<FocusOut>', lambda e: self.on_item_id_focus_out())
                # 绑定向下键，打开下拉列表
                self.item_id_combo.bind('<Down>', lambda e: self._open_dropdown_on_key())
                # 绑定点击事件，确保下拉列表可用
                self.item_id_combo.bind('<Button-1>', lambda e: self._ensure_dropdown_available())
                # 标记初始化完成
                self._initializing = False
                # 加载物品ID历史记录（不在初始化时触发搜索）
                self._safe_load_item_ids()
            except Exception as e:
                print(f"绑定事件时出错: {e}")
                import traceback
                traceback.print_exc()
                # 即使出错也要标记初始化完成
                self._initializing = False
        
        # 延迟2秒绑定事件，确保所有初始化完成（防止递归）
        self.root.after(2000, bind_events_after_init)
        
        ttk.Label(item_frame, text="数量:", font=("Microsoft YaHei", 9)).grid(row=0, column=2, sticky=tk.W, padx=(15, 5))
        self.item_amount_var = tk.StringVar(value="1")
        ttk.Entry(item_frame, textvariable=self.item_amount_var, width=10, font=("Consolas", 9)).grid(row=0, column=3, sticky=tk.W, padx=3)
        
        # 动态参数区域
        self.dynamic_params_frame = ttk.Frame(gm_frame)
        self.dynamic_params_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6, padx=3)
        self.dynamic_params_frame.columnconfigure(1, weight=1)
        self.dynamic_params_vars = {}
        
        # 底部按钮
        bottom_btn_frame = ttk.Frame(gm_frame)
        bottom_btn_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=8)
        ttk.Button(bottom_btn_frame, text="🚀 发送命令", command=self.send_gm_command, width=15).grid(row=0, column=0, sticky=tk.W, padx=5)
        
        # 特效动画（放在右下角）
        self.effect_canvas = tk.Canvas(gm_frame, width=100, height=50, highlightthickness=0, bg="#f8f9fa")
        self.effect_canvas.place(relx=1.0, rely=1.0, anchor=tk.SE, x=-10, y=-10)
        self.effect_angle = 0
        # 延迟启动动画，确保Canvas已完全创建
        self.root.after(100, self.animate_effect)
        
        # 设置output_text和input_text为cmd_output_text（用于兼容）
        self.output_text = self.cmd_output_text
        self.input_text = self.cmd_output_text
        
        # 延迟加载GM模板（兼容缺失方法）
        try:
            loader2 = getattr(self, "load_gm_templates")
        except Exception:
            loader2 = lambda: None
        self.root.after(200, loader2)
    
    def _on_item_id_key_release(self, event=None):
        """处理按键释放事件（防止递归）"""
        # 严格的初始化检查（必须在初始化完成后才能处理事件）
        if not hasattr(self, '_initializing') or self._initializing:
            return
        
        # 忽略特殊键和控制键（但允许 BackSpace 和 Delete，因为删除字符后需要搜索）
        if event and hasattr(event, 'keysym'):
            keysym = event.keysym
            # 忽略所有控制键和功能键（但允许 BackSpace 和 Delete）
            if keysym in ['Up', 'Down', 'Left', 'Right', 'Return', 'Tab', 'Escape', 
                         'Control_L', 'Control_R', 'Alt_L', 'Alt_R', 'Shift_L', 'Shift_R',
                         'F1', 'F2', 'F3', 'F4', 'F5', 'F6', 'F7', 'F8', 'F9', 'F10', 'F11', 'F12',
                         'Home', 'End', 'Page_Up', 'Page_Down',
                         'Insert', 'Print', 'Scroll_Lock', 'Pause', 'Num_Lock', 'Caps_Lock']:
                return
        
        # 再次检查初始化标志和递归保护标志
        if (self._initializing or self._filtering_active or 
            self._opening_dropdown or self._updating_values):
            return
        
        # 触发防抖筛选（使用延迟调用，避免立即递归）
        try:
            self._filter_item_ids_debounced()
        except Exception as e:
            print(f"处理按键事件时出错: {e}")
    
    def _filter_item_ids_debounced(self):
        """防抖版本的筛选函数（优化性能，避免频繁搜索，防止递归）"""
        # 严格的递归保护
        if self._initializing or self._filtering_active or self._opening_dropdown or self._updating_values:
            return
        
        # 获取当前输入
        try:
            current_input = self.item_id_var.get().strip() if hasattr(self, 'item_id_var') else ""
        except:
            return
        
        # 取消之前的定时器
        if self._filter_debounce_id:
            try:
                self.root.after_cancel(self._filter_debounce_id)
            except:
                pass
        
        # 设置新的定时器（150毫秒后执行搜索，减少延迟提高响应速度）
        def do_filter():
            try:
                # 再次检查所有标志
                if self._initializing or self._filtering_active or self._opening_dropdown or self._updating_values:
                    return
                
                # 获取最新的输入
                current_check = self.item_id_var.get().strip() if hasattr(self, 'item_id_var') else ""
                
                # 更新搜索文本并执行搜索（总是执行，确保搜索正常）
                self._last_search_text = current_check
                self.filter_item_ids()
            except Exception as e:
                import traceback
                print(f"筛选物品ID时出错: {e}")
                traceback.print_exc()
        
        self._filter_debounce_id = self.root.after(150, do_filter)
    def _normalize_item_id(self, item_id_full):
        """标准化物品ID（统一转换为字符串格式：ID;名称，用于内部存储和搜索）"""
        # 处理 None 或空值
        if item_id_full is None:
            return None
        
        try:
            # 如果是字符串，直接使用
            if isinstance(item_id_full, str):
                item_id_full_str = item_id_full.strip()
                if item_id_full_str:
                    return item_id_full_str
                return None
            
            # 如果是字典，提取 id 和 name
            if isinstance(item_id_full, dict):
                try:
                    # 安全地获取 id 和 name
                    item_id = ''
                    item_name = ''
                    
                    if 'id' in item_id_full:
                        item_id = str(item_id_full['id']).strip() if item_id_full['id'] is not None else ''
                    if 'name' in item_id_full:
                        item_name = str(item_id_full['name']).strip() if item_id_full['name'] is not None else ''
                    
                    # 如果字典有 get 方法，也尝试使用（兼容性）
                    if not item_id and hasattr(item_id_full, 'get'):
                        try:
                            item_id = str(item_id_full.get('id', '')).strip() if item_id_full.get('id') is not None else ''
                        except:
                            pass
                    if not item_name and hasattr(item_id_full, 'get'):
                        try:
                            item_name = str(item_id_full.get('name', '')).strip() if item_id_full.get('name') is not None else ''
                        except:
                            pass
                    
                    # 组合结果（内部格式：ID;名称）
                    if item_id and item_name:
                        return f"{item_id};{item_name}"
                    elif item_id:
                        return item_id
                    elif item_name:
                        return item_name
                    else:
                        # 如果都没有，尝试使用第一个非空值
                        try:
                            for key, value in item_id_full.items():
                                if value is not None:
                                    value_str = str(value).strip()
                                    if value_str:
                                        return value_str
                        except:
                            pass
                        # 如果所有值都是空的，返回空字符串的字符串表示
                        return str(item_id_full).strip()
                except Exception as e:
                    # 如果处理字典时出错，尝试直接转换为字符串
                    try:
                        return str(item_id_full).strip()
                    except:
                        return None
            
            # 如果是列表或元组，尝试使用第一个元素
            if isinstance(item_id_full, (list, tuple)):
                if len(item_id_full) > 0:
                    return self._normalize_item_id(item_id_full[0])
                return None
            
            # 其他类型，直接转换为字符串
            try:
                result = str(item_id_full).strip()
                return result if result else None
            except:
                return None
        except Exception as e:
            # 所有异常都返回 None，避免程序崩溃
            return None
    
    def _is_valid_item_id(self, item_id_str):
        """检查物品ID是否是有效格式（过滤掉纯文本搜索记录）"""
        if not item_id_str:
            return False
        
        try:
            item_str = str(item_id_str).strip()
            if not item_str:
                return False
            
            # 如果包含数字，认为是有效的物品ID（大多数物品ID都包含数字）
            if re.search(r'\d', item_str):
                return True
            
            # 如果包含分号或减号分隔符，即使没有数字也认为是有效格式（可能是特殊物品）
            if ';' in item_str or ' - ' in item_str:
                return True
            
            # 如果是纯中文/文本且不包含数字和分隔符，认为是搜索记录，过滤掉
            # 只过滤明显的纯文本搜索记录（如"礼包"、"物品"等）
            if re.match(r'^[\u4e00-\u9fa5a-zA-Z\s]+$', item_str) and len(item_str) <= 10:
                # 纯文本且长度较短，可能是搜索记录
                return False
            
            # 其他情况都认为是有效的（保守策略，避免过滤掉有效项）
            return True
        except Exception as e:
            # 如果检查失败，返回True（保守策略，避免过滤掉有效项）
            print(f"验证物品ID时出错: {e}, item_id_str: {item_id_str}")
            return True
    
    def _format_item_id_for_display(self, item_id_str):
        """格式化物品ID用于显示（将 ID;名称 转换为更友好的显示格式：ID - 名称）"""
        if not item_id_str:
            return item_id_str
        
        try:
            # 如果包含分号，说明有ID和名称
            if ';' in item_id_str:
                parts = item_id_str.split(';', 1)
                item_id = parts[0].strip()
                item_name = parts[1].strip() if len(parts) > 1 else ''
                
                # 如果名称不为空，显示为 "ID - 名称"
                if item_name:
                    return f"{item_id} - {item_name}"
                else:
                    # 如果名称为空，只显示ID
                    return item_id
            else:
                # 如果没有分号，直接返回（可能是纯ID或旧格式）
                return item_id_str
        except Exception as e:
            # 如果格式化失败，返回原字符串
            return item_id_str
    
    def filter_item_ids(self, force_show_all=False):
        """根据输入内容筛选物品ID（彻底重构版，解决所有显示问题）
        
        Args:
            force_show_all: 如果为True，强制显示所有物品ID，忽略输入框内容
        """
        # 严格的递归保护
        if self._initializing or self._filtering_active or self._opening_dropdown:
            if not force_show_all:
                return
        
        if not hasattr(self, 'item_id_combo') or not self.item_id_combo:
            return
        
        # 设置筛选标志
        self._filtering_active = True
        
        try:
            # 获取搜索列表（优先使用 item_ids_all，如果没有则使用 item_ids_history，最后尝试从文件加载）
            search_list = []
            try:
                # 首先尝试使用 item_ids_all
                if hasattr(self, 'item_ids_all') and self.item_ids_all:
                    if isinstance(self.item_ids_all, list) and len(self.item_ids_all) > 0:
                        search_list = self.item_ids_all.copy()
                
                # 如果 item_ids_all 为空，尝试使用 item_ids_history
                if not search_list and hasattr(self, 'item_ids_history') and self.item_ids_history:
                    if isinstance(self.item_ids_history, list) and len(self.item_ids_history) > 0:
                        search_list = self.item_ids_history.copy()
                        # 同时更新 item_ids_all
                        self.item_ids_all = search_list.copy()
                
                # 如果都为空，尝试重新加载数据
                if not search_list:
                    try:
                        self._safe_load_item_ids()
                        if hasattr(self, 'item_ids_all') and self.item_ids_all:
                            if isinstance(self.item_ids_all, list) and len(self.item_ids_all) > 0:
                                search_list = self.item_ids_all.copy()
                        elif hasattr(self, 'item_ids_history') and self.item_ids_history:
                            if isinstance(self.item_ids_history, list) and len(self.item_ids_history) > 0:
                                search_list = self.item_ids_history.copy()
                                self.item_ids_all = search_list.copy()
                    except Exception as e:
                        print(f"重新加载数据时出错: {e}")
                        
            except Exception as e:
                print(f"获取搜索列表时出错: {e}")
                import traceback
                traceback.print_exc()
                search_list = []
            
            # 过滤掉 None 和无效项（只保留有效的物品ID格式）
            # 注意：保留原始数据，不要丢失信息
            valid_search_list = []
            try:
                for item in search_list:
                    if item is None:
                        continue
                    # 标准化物品ID（但保留原始格式用于搜索）
                    item_id_str = self._normalize_item_id(item)
                    if item_id_str and self._is_valid_item_id(item_id_str):
                        valid_search_list.append(item_id_str)
            except Exception as e:
                print(f"过滤搜索列表时出错: {e}")
                import traceback
                traceback.print_exc()
            
            search_list = valid_search_list
            
            # 如果没有数据，清空下拉列表并返回
            if not search_list:
                try:
                    if self.item_id_combo:
                        self.item_id_combo['values'] = []
                        if hasattr(self, '_item_id_display_map'):
                            self._item_id_display_map.clear()
                except Exception as e:
                    print(f"清空下拉列表时出错: {e}")
                return
            
            # 如果强制显示所有，或者没有输入，显示所有物品ID
            if force_show_all:
                # 强制显示所有：显示所有物品ID（增加到500个，避免显示不全）
                display_items = []
                seen_items = set()
                
                try:
                    # 不再限制数量：显示列表中的全部项目
                    max_items = len(search_list)
                    for item_id_str in search_list[:max_items]:
                        if item_id_str and item_id_str not in seen_items:
                            display_str = self._format_item_id_for_display(item_id_str)
                            display_items.append(display_str)
                            seen_items.add(item_id_str)
                except Exception as e:
                    print(f"处理所有物品ID时出错: {e}")
                
                # 更新下拉列表
                self._update_dropdown_values(display_items, search_list[:max_items] if search_list else [])
                
            else:
                # 正常模式：根据输入内容决定显示什么
                try:
                    current_input = self.item_id_var.get()
                    current_input_trimmed = current_input.strip() if current_input else ""
                except Exception as e:
                    print(f"获取输入内容时出错: {e}")
                    current_input_trimmed = ""
                
                if not current_input_trimmed:
                    # 没有输入：显示所有物品ID（不再限制数量）
                    display_items = []
                    seen_items = set()
                    
                    try:
                        max_items = len(search_list)
                        for item_id_str in search_list[:max_items]:
                            if item_id_str and item_id_str not in seen_items:
                                display_str = self._format_item_id_for_display(item_id_str)
                                display_items.append(display_str)
                                seen_items.add(item_id_str)
                    except Exception as e:
                        print(f"处理所有物品ID时出错: {e}")
                    
                    # 更新下拉列表
                    self._update_dropdown_values(display_items, search_list[:max_items] if search_list else [])
                    
                else:
                    # 有输入：执行搜索
                    search_input = current_input_trimmed
                    # 预处理搜索输入：去除多余空格，支持多种格式
                    search_input_clean = search_input.strip()
                    
                    # 准备多个搜索关键词（支持多种格式）
                    search_keywords = []
                    # 原始输入
                    search_keywords.append(search_input_clean.lower())
                    # 如果包含 " - "，转换为分号格式
                    if ' - ' in search_input_clean:
                        search_keywords.append(search_input_clean.replace(' - ', ';').lower())
                    # 如果包含分号，转换为 " - " 格式
                    if ';' in search_input_clean:
                        search_keywords.append(search_input_clean.replace(';', ' - ').lower())
                    
                    # 提取纯数字部分（如果输入是 "400590" 或 "400590 - 名称"）
                    import re
                    numbers = re.findall(r'\d+', search_input_clean)
                    if numbers:
                        search_keywords.extend([num.lower() for num in numbers])
                    
                    # 提取中文部分（如果输入包含中文）
                    chinese_chars = re.findall(r'[\u4e00-\u9fa5]+', search_input_clean)
                    if chinese_chars:
                        search_keywords.extend([char.lower() for char in chinese_chars])
                    
                    # 去重搜索关键词
                    search_keywords = list(set(search_keywords))
                    
                    filtered_ids = []
                    seen_ids = set()
                    # 不再限制最大结果数
                    max_results = len(search_list) if search_list else 0
                    
                    try:
                        for item_id_str in search_list:
                            if len(filtered_ids) >= max_results:
                                break
                            
                            if item_id_str in seen_ids:
                                continue
                            
                            try:
                                # 标准化物品ID字符串
                                item_id_normalized = item_id_str
                                item_id_lower = item_id_normalized.lower()
                                
                                # 生成显示格式
                                display_str = self._format_item_id_for_display(item_id_normalized)
                                display_str_lower = display_str.lower() if display_str else ""
                                
                                # 提取ID部分和名称部分
                                item_id_only = self.extract_item_id(item_id_normalized)
                                item_id_only_lower = item_id_only.lower() if item_id_only else ""
                                
                                # 提取名称部分（如果有）
                                item_name = ""
                                if ';' in item_id_normalized:
                                    parts = item_id_normalized.split(';', 1)
                                    if len(parts) > 1:
                                        item_name = parts[1].strip()
                                elif ' - ' in display_str:
                                    parts = display_str.split(' - ', 1)
                                    if len(parts) > 1:
                                        item_name = parts[1].strip()
                                
                                item_name_lower = item_name.lower() if item_name else ""
                                
                                # 多关键词匹配：只要有一个关键词匹配就认为匹配
                                matched = False
                                for keyword in search_keywords:
                                    if not keyword:
                                        continue
                                    # 在完整格式中匹配
                                    if keyword in item_id_lower:
                                        matched = True
                                        break
                                    # 在显示格式中匹配
                                    if keyword in display_str_lower:
                                        matched = True
                                        break
                                    # 在ID部分匹配
                                    if keyword in item_id_only_lower:
                                        matched = True
                                        break
                                    # 在名称部分匹配
                                    if keyword in item_name_lower:
                                        matched = True
                                        break
                                
                                if matched:
                                    filtered_ids.append(item_id_str)
                                    seen_ids.add(item_id_str)
                                    
                            except Exception as e:
                                # 如果处理某项时出错，继续处理下一项
                                continue
                                
                    except Exception as e:
                        print(f"搜索物品ID时出错: {e}")
                        import traceback
                        traceback.print_exc()
                    
                    # 更新下拉列表
                    display_items = [self._format_item_id_for_display(item) for item in filtered_ids]
                    self._update_dropdown_values(display_items, filtered_ids)
                
        except Exception as e:
            import traceback
            print(f"筛选物品ID时出错: {e}")
            traceback.print_exc()
        finally:
            # 清除筛选标志
            self.root.after(50, lambda: setattr(self, '_filtering_active', False))
    
    def _update_dropdown_values(self, display_items, full_items):
        """更新下拉框的值（内部方法，确保映射正确）"""
        if not hasattr(self, 'item_id_combo') or not self.item_id_combo:
            return
        
        try:
            self._updating_values = True
            
            # 创建显示文本到完整格式的映射
            if not hasattr(self, '_item_id_display_map'):
                self._item_id_display_map = {}
            
            self._item_id_display_map.clear()
            for i, full_str in enumerate(full_items):
                if i < len(display_items):
                    display_str = display_items[i]
                    self._item_id_display_map[display_str] = full_str
            
            # 更新下拉框的值
            self.item_id_combo['values'] = display_items
            
        except Exception as e:
            print(f"更新下拉框值时出错: {e}")
        finally:
            self.root.after(50, lambda: setattr(self, '_updating_values', False))
    
    
    def _auto_open_dropdown_if_needed(self):
        """如果需要，自动打开下拉列表（已禁用，避免递归）"""
        # 此功能已禁用，避免触发事件导致递归
        # 用户可以通过点击下拉箭头或按向下键手动打开下拉列表
        pass
    
    def _try_open_dropdown_f4(self):
        """尝试使用F4键打开下拉列表（已禁用，避免递归）"""
        # 此功能已禁用，避免触发事件导致递归
        pass
    
    def _open_dropdown_on_key(self, event=None):
        """当按下向下键时打开下拉列表"""
        try:
            # 如果正在初始化或更新，不处理
            if self._initializing or self._updating_values or self._filtering_active:
                return "break"
            
            # 如果下拉列表为空，触发搜索
            if not self.item_id_combo['values']:
                # 使用防抖函数触发搜索
                self._filter_item_ids_debounced()
            
            # 获取当前的值列表
            values = self.item_id_combo['values']
            
            # 如果有值，打开下拉列表
            if values:
                try:
                    # 使用Alt+Down打开下拉列表
                    self.item_id_combo.event_generate('<Alt-Down>')
                except:
                    pass
                
            return "break"  # 阻止默认行为，避免光标移动
        except:
            return "break"
    
    def _ensure_dropdown_available(self, event=None):
        """确保下拉列表可用（点击时）"""
        try:
            # 如果正在初始化或更新，不处理
            if self._initializing or self._updating_values or self._filtering_active:
                return
            
            # 如果下拉列表为空，触发显示所有物品
            if not self.item_id_combo['values']:
                # 直接触发筛选，显示所有物品
                self.root.after(10, lambda: self.filter_item_ids())
        except:
            pass
    
    def on_item_id_selected(self):
        """当从下拉列表选择物品ID时（显示ID+名称）"""
        try:
            # 设置标志，防止触发搜索
            self._updating_values = True
            try:
                item_id_selected = self.item_id_var.get().strip()
                if item_id_selected:
                    # 尝试从显示映射中获取完整格式
                    full_item_str = None
                    if hasattr(self, '_item_id_display_map') and self._item_id_display_map:
                        full_item_str = self._item_id_display_map.get(item_id_selected)
                    
                    # 如果找到了完整格式，格式化为显示格式（ID - 名称）
                    if full_item_str:
                        display_str = self._format_item_id_for_display(full_item_str)
                        if display_str:
                            self.item_id_var.set(display_str)
                    else:
                        # 如果没有找到映射，尝试解析当前选择的值
                        # 如果已经是显示格式（包含 " - "），保持不变
                        if ' - ' in item_id_selected:
                            # 已经是显示格式，保持不变
                            pass
                        elif ';' in item_id_selected:
                            # 如果是内部格式（ID;名称），转换为显示格式
                            display_str = self._format_item_id_for_display(item_id_selected)
                            if display_str and display_str != item_id_selected:
                                self.item_id_var.set(display_str)
                        else:
                            # 如果是纯ID，尝试从历史记录中查找完整记录
                            item_id_only = self.extract_item_id(item_id_selected)
                            if item_id_only:
                                for existing_item in self.item_ids_history:
                                    existing_id = self.extract_item_id(existing_item)
                                    if existing_id == item_id_only:
                                        # 找到匹配的记录，格式化为显示格式
                                        display_str = self._format_item_id_for_display(
                                            self._normalize_item_id(existing_item) or existing_item
                                        )
                                        if display_str:
                                            self.item_id_var.set(display_str)
                                        break
            finally:
                self.root.after(100, lambda: setattr(self, '_updating_values', False))
        except Exception as e:
            print(f"选择物品ID时出错: {e}")
            import traceback
            traceback.print_exc()
    
    def on_item_id_focus_out(self):
        """当物品ID输入框失去焦点时（保持用户输入，不自动填充历史记录）"""
        # 不修改输入框内容，保持用户输入的内容
        pass
    
    def load_gm_templates(self):
        """
        加载GM命令模板到下拉框（扁平化，去除分类）。
        兼容多种文件格式：
        - ["cmd1", "cmd2"]
        - [{"name": "赠送物品", "cmd": "give {role} {item} {count}"}]
        - {"分类": {"命令名": {...}}, ...}  # 分类格式，会扁平化
        - {"命令名": "命令内容", ...}  # 扁平格式
        文件路径：self.gm_templates_file（位于程序目录）
        """
        try:
            templates = []
            templates_map = {}
            
            if os.path.exists(self.gm_templates_file):
                with open(self.gm_templates_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                
                # 归一化为 name -> cmd（扁平化处理）
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, str):
                            name = item.strip()
                            if not name:
                                continue
                            templates.append(name)
                            templates_map[name] = name
                        elif isinstance(item, dict):
                            name = str(item.get("name") or item.get("title") or item.get("label") or "").strip()
                            cmd = str(item.get("cmd") or item.get("template") or item.get("command") or "").strip()
                            if name and cmd:
                                templates.append(name)
                                templates_map[name] = cmd
                elif isinstance(data, dict):
                    # 检查是否是分类格式（嵌套字典）
                    is_category_format = False
                    for key, value in data.items():
                        if isinstance(value, dict) and value:
                            # 检查value是否是命令字典（包含command字段）
                            first_value = next(iter(value.values()))
                            if isinstance(first_value, dict) and ('command' in first_value or 'cmd' in first_value):
                                is_category_format = True
                                break
                    
                    if is_category_format:
                        # 分类格式：{"分类": {"命令名": {...}}, ...}
                        # 扁平化：提取所有分类下的命令
                        for category, commands in data.items():
                            if isinstance(commands, dict):
                                for cmd_name, cmd_info in commands.items():
                                    if isinstance(cmd_info, dict):
                                        # 提取命令内容
                                        cmd = cmd_info.get('command', '') or cmd_info.get('cmd', '')
                                        if cmd:
                                            name_str = str(cmd_name).strip()
                                            cmd_str = str(cmd).strip()
                                            if name_str:
                                                templates.append(name_str)
                                                templates_map[name_str] = cmd_str
                                    else:
                                        # 直接是命令字符串
                                        name_str = str(cmd_name).strip()
                                        cmd_str = str(cmd_info).strip()
                                        if name_str and cmd_str:
                                            templates.append(name_str)
                                            templates_map[name_str] = cmd_str
                    else:
                        # 扁平格式：{"命令名": "命令内容", ...}
                        for name, cmd in data.items():
                            name_str = str(name).strip()
                            if isinstance(cmd, dict):
                                # 如果是字典，提取command字段
                                cmd_str = str(cmd.get('command', '') or cmd.get('cmd', '')).strip()
                            else:
                                cmd_str = str(cmd).strip()
                            if name_str and cmd_str:
                                templates.append(name_str)
                                templates_map[name_str] = cmd_str
            else:
                # 提供少量默认模板，首次启动可用
                templates = ["give_item", "add_gold", "announce"]
                templates_map = {
                    "give_item": "give {role} {item} {count}",
                    "add_gold": "gold add {role} {amount}",
                    "announce": "announce {message}"
                }
            
            # 去重并排序（保持原有顺序）
            seen = set()
            normalized = []
            for name in templates:
                if name not in seen:
                    seen.add(name)
                    normalized.append(name)
            templates = normalized
            
            # 保存映射以便后续选择时取模板内容
            self.gm_templates = templates
            self.gm_templates_map = templates_map
            
            # 更新下拉框
            if hasattr(self, "gm_cmd_combo") and self.gm_cmd_combo:
                self.gm_cmd_combo["values"] = templates
        except Exception as e:
            # 失败时不阻塞启动
            print(f"加载GM模板失败: {e}")
            import traceback
            traceback.print_exc()
            try:
                if hasattr(self, "gm_cmd_combo") and self.gm_cmd_combo:
                    self.gm_cmd_combo["values"] = []
            except Exception:
                pass
    
    def save_gm_templates(self, templates):
        """保存GM命令模板"""
        try:
            with open(self.gm_templates_file, 'w', encoding='utf-8') as f:
                json.dump(templates, f, indent=2, ensure_ascii=False)
            # 重新加载模板
            self.load_gm_templates()
        except Exception as e:
            messagebox.showerror("错误", f"保存GM模板失败: {e}")
    
    def manage_gm_templates(self):
        """管理GM命令模板（无分类版本）"""
        manage_window = tk.Toplevel(self.root)
        manage_window.title("管理GM命令模板")
        manage_window.geometry("800x600")
        manage_window.minsize(600, 400)
        manage_window.transient(self.root)
        
        # 加载当前模板并扁平化
        flat_commands = {}  # 命令名 -> {command, params}
        if os.path.exists(self.gm_templates_file):
            try:
                with open(self.gm_templates_file, 'r', encoding='utf-8') as f:
                    templates = json.load(f)
                    # 将所有分类下的命令扁平化
                    if isinstance(templates, dict):
                        for category, commands in templates.items():
                            if isinstance(commands, dict):
                                for cmd_name, cmd_info in commands.items():
                                    if isinstance(cmd_info, dict):
                                        flat_commands[cmd_name] = {
                                            'command': cmd_info.get('command', ''),
                                            'params': cmd_info.get('params', '')
                                        }
                                    else:
                                        flat_commands[cmd_name] = {
                                            'command': str(cmd_info),
                                            'params': ''
                                        }
            except:
                flat_commands = {}
        
        # 主框架（使用grid布局，支持自适应）
        manage_window.columnconfigure(0, weight=1)
        manage_window.rowconfigure(0, weight=1)
        
        main_frame = ttk.Frame(manage_window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # 命令列表标题
        ttk.Label(main_frame, text="命令列表:", font=("Microsoft YaHei", 10, "bold")).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        # 创建Treeview和滚动条的容器
        tree_frame = ttk.Frame(main_frame)
        tree_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        # 创建Treeview显示命令
        command_tree = ttk.Treeview(tree_frame, columns=("command", "params"), show="tree headings")
        command_tree.heading("#0", text="命令名")
        command_tree.heading("command", text="命令")
        command_tree.heading("params", text="参数")
        
        # 设置列宽（使用stretch让列可以自适应）
        command_tree.column("#0", width=200, minwidth=150, stretch=True)
        command_tree.column("command", width=300, minwidth=200, stretch=True)
        command_tree.column("params", width=200, minwidth=150, stretch=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=command_tree.yview)
        command_tree.configure(yscrollcommand=scrollbar.set)
        
        # 使用grid布局，支持自适应
        command_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        def refresh_command_list():
            """刷新命令列表"""
            # 清空现有命令
            for item in command_tree.get_children():
                command_tree.delete(item)
            
            # 填充命令
            for cmd_name, cmd_info in sorted(flat_commands.items()):
                command = cmd_info.get('command', '')
                params = cmd_info.get('params', '')
                command_tree.insert("", tk.END, text=cmd_name, values=(command, params))
        
        refresh_command_list()
        
        # 按钮框架（使用grid布局，放在主框架下方）
        btn_frame = ttk.Frame(main_frame, padding="10")
        btn_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        main_frame.rowconfigure(2, weight=0)  # 按钮框架不拉伸
        
        def add_command():
            """添加命令"""
            # 创建添加命令对话框
            cmd_window = tk.Toplevel(manage_window)
            cmd_window.title("添加命令")
            cmd_window.geometry("450x280")
            cmd_window.transient(manage_window)
            cmd_window.grab_set()
            
            # 居中显示
            cmd_window.update_idletasks()
            x = (cmd_window.winfo_screenwidth() // 2) - (450 // 2)
            y = (cmd_window.winfo_screenheight() // 2) - (280 // 2)
            cmd_window.geometry(f"450x280+{x}+{y}")
            
            # 配置列权重，使输入框可以自适应
            cmd_window.columnconfigure(1, weight=1)
            
            # 命令名
            ttk.Label(cmd_window, text="命令名:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
            name_var = tk.StringVar()
            name_entry = tk.Entry(cmd_window, textvariable=name_var, width=35, font=("Consolas", 9))
            name_entry.grid(row=0, column=1, padx=10, pady=10, sticky=(tk.W, tk.E))
            
            # 命令内容
            ttk.Label(cmd_window, text="命令:", font=("Microsoft YaHei", 9)).grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
            cmd_var = tk.StringVar()
            cmd_entry = tk.Entry(cmd_window, textvariable=cmd_var, width=35, font=("Consolas", 9))
            cmd_entry.grid(row=1, column=1, padx=10, pady=10, sticky=(tk.W, tk.E))
            
            # 参数
            ttk.Label(cmd_window, text="参数:", font=("Microsoft YaHei", 9)).grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
            params_var = tk.StringVar()
            params_entry = tk.Entry(cmd_window, textvariable=params_var, width=35, font=("Consolas", 9))
            params_entry.grid(row=2, column=1, padx=10, pady=10, sticky=(tk.W, tk.E))
            
            # 延迟设置焦点，确保窗口完全创建后再设置
            def set_focus():
                name_entry.focus_set()
                name_entry.icursor(tk.END)
            cmd_window.after(100, set_focus)
            
            def save_command():
                name = name_var.get().strip()
                cmd = cmd_var.get().strip()
                params = params_var.get().strip()
                
                if not name:
                    messagebox.showwarning("提示", "命令名不能为空", parent=cmd_window)
                    name_entry.focus_set()
                    return
                
                if not cmd:
                    messagebox.showwarning("提示", "命令不能为空", parent=cmd_window)
                    cmd_entry.focus_set()
                    return
                
                try:
                    if name in flat_commands:
                        if not messagebox.askyesno("确认", f"命令 '{name}' 已存在，是否覆盖？", parent=cmd_window):
                            return
                    
                    flat_commands[name] = {
                        'command': cmd,
                        'params': params
                    }
                    # 保存到文件（使用默认分类保存，保持兼容性）
                    self._save_flat_commands(flat_commands)
                    refresh_command_list()
                    cmd_window.destroy()
                    messagebox.showinfo("成功", "命令已添加", parent=manage_window)
                except Exception as e:
                    messagebox.showerror("错误", f"保存命令失败: {e}", parent=cmd_window)
                    import traceback
                    traceback.print_exc()
            
            # 按钮框架
            btn_frame = ttk.Frame(cmd_window)
            btn_frame.grid(row=3, column=0, columnspan=2, pady=20)
            
            ttk.Button(btn_frame, text="确定", command=save_command, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=cmd_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
            
            # 绑定回车键（只在输入框有焦点时触发）
            def on_return(event):
                # 检查当前焦点在哪个输入框
                focus_widget = cmd_window.focus_get()
                if focus_widget in [name_entry, cmd_entry, params_entry]:
                    save_command()
                return "break"
            
            cmd_window.bind('<Return>', on_return)
            # 绑定ESC键关闭
            cmd_window.bind('<Escape>', lambda e: cmd_window.destroy())
            
            # 确保输入框可以正常接收输入（移除可能干扰的事件）
            for entry in [name_entry, cmd_entry, params_entry]:
                # 确保输入框可以正常输入
                entry.config(validate='none')  # 禁用验证，避免干扰输入
        
        def edit_command():
            """编辑命令"""
            cmd_selection = command_tree.selection()
            if not cmd_selection:
                messagebox.showwarning("提示", "请先选择要编辑的命令")
                return
            
            cmd_name = command_tree.item(cmd_selection[0])['text']
            cmd_info = flat_commands.get(cmd_name, {})
            
            if not cmd_info:
                messagebox.showerror("错误", f"未找到命令 '{cmd_name}' 的信息")
                return
            
            # 创建编辑命令对话框
            cmd_window = tk.Toplevel(manage_window)
            cmd_window.title("编辑命令")
            cmd_window.geometry("450x280")
            cmd_window.transient(manage_window)
            cmd_window.grab_set()
            
            # 居中显示
            cmd_window.update_idletasks()
            x = (cmd_window.winfo_screenwidth() // 2) - (450 // 2)
            y = (cmd_window.winfo_screenheight() // 2) - (280 // 2)
            cmd_window.geometry(f"450x280+{x}+{y}")
            
            # 命令名（允许编辑）
            ttk.Label(cmd_window, text="命令名:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
            name_var = tk.StringVar(value=cmd_name)
            name_entry = ttk.Entry(cmd_window, textvariable=name_var, width=35, font=("Consolas", 9))
            name_entry.grid(row=0, column=1, padx=10, pady=10, sticky=(tk.W, tk.E))
            name_entry.focus_set()
            cmd_window.columnconfigure(1, weight=1)
            
            # 命令内容
            ttk.Label(cmd_window, text="命令:", font=("Microsoft YaHei", 9)).grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
            cmd_var = tk.StringVar(value=cmd_info.get('command', ''))
            cmd_entry = ttk.Entry(cmd_window, textvariable=cmd_var, width=35, font=("Consolas", 9))
            cmd_entry.grid(row=1, column=1, padx=10, pady=10, sticky=(tk.W, tk.E))
            
            # 参数
            ttk.Label(cmd_window, text="参数:", font=("Microsoft YaHei", 9)).grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
            params_var = tk.StringVar(value=cmd_info.get('params', ''))
            params_entry = ttk.Entry(cmd_window, textvariable=params_var, width=35, font=("Consolas", 9))
            params_entry.grid(row=2, column=1, padx=10, pady=10, sticky=(tk.W, tk.E))
            
            def save_command():
                new_name = name_var.get().strip()
                cmd = cmd_var.get().strip()
                params = params_var.get().strip()
                
                if not new_name:
                    messagebox.showwarning("提示", "命令名不能为空")
                    return
                
                if not cmd:
                    messagebox.showwarning("提示", "命令不能为空")
                    return
                
                try:
                    # 如果命令名改变了，需要删除旧的并添加新的
                    if new_name != cmd_name:
                        if new_name in flat_commands:
                            if not messagebox.askyesno("确认", f"命令 '{new_name}' 已存在，是否覆盖？"):
                                return
                        # 删除旧命令
                        if cmd_name in flat_commands:
                            del flat_commands[cmd_name]
                        # 添加新命令
                        flat_commands[new_name] = {
                            'command': cmd,
                            'params': params
                        }
                    else:
                        # 命令名没变，直接更新
                        flat_commands[cmd_name]['command'] = cmd
                        flat_commands[cmd_name]['params'] = params
                    
                    # 保存到文件
                    self._save_flat_commands(flat_commands)
                    refresh_command_list()
                    
                    # 如果命令名改变了，重新选择新命令
                    if new_name != cmd_name:
                        for item in command_tree.get_children():
                            if command_tree.item(item)['text'] == new_name:
                                command_tree.selection_set(item)
                                command_tree.see(item)
                                break
                    
                    cmd_window.destroy()
                    messagebox.showinfo("成功", "命令已保存")
                except Exception as e:
                    messagebox.showerror("错误", f"保存命令失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            # 按钮框架
            btn_frame = ttk.Frame(cmd_window)
            btn_frame.grid(row=3, column=0, columnspan=2, pady=20)
            
            ttk.Button(btn_frame, text="确定", command=save_command, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=cmd_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
            
            # 绑定回车键
            cmd_window.bind('<Return>', lambda e: save_command())
        
        def delete_command():
            """删除命令"""
            cmd_selection = command_tree.selection()
            if not cmd_selection:
                messagebox.showwarning("提示", "请先选择要删除的命令")
                return
            
            cmd_name = command_tree.item(cmd_selection[0])['text']
            if messagebox.askyesno("确认", f"确定要删除命令 '{cmd_name}' 吗？"):
                del flat_commands[cmd_name]
                self._save_flat_commands(flat_commands)
                refresh_command_list()
        
        # 按钮
        ttk.Button(btn_frame, text="添加命令", command=add_command).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="编辑命令", command=edit_command).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="删除命令", command=delete_command).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=manage_window.destroy).pack(side=tk.RIGHT, padx=5)
    
    def _save_flat_commands(self, flat_commands):
        """保存扁平化的命令到文件（使用默认分类保持兼容性）"""
        # 为了保持与现有格式的兼容性，将所有命令保存到一个默认分类下
        templates = {
            "默认分类": {}
        }
        for cmd_name, cmd_info in flat_commands.items():
            templates["默认分类"][cmd_name] = {
                'description': '',
                'command': cmd_info.get('command', ''),
                'params': cmd_info.get('params', '')
            }
        self.save_gm_templates(templates)
    
    def import_gm_config(self):
        """导入物品ID（支持本地文件和服务器文件）"""
        # 创建选择对话框
        import_window = tk.Toplevel(self.root)
        import_window.title("导入物品ID")
        import_window.geometry("400x150")
        import_window.transient(self.root)
        import_window.grab_set()
        
        # 居中显示
        import_window.update_idletasks()
        x = (import_window.winfo_screenwidth() // 2) - (400 // 2)
        y = (import_window.winfo_screenheight() // 2) - (150 // 2)
        import_window.geometry(f"400x150+{x}+{y}")
        
        ttk.Label(import_window, text="选择导入方式:", font=("Microsoft YaHei", 10)).pack(pady=20)
        
        btn_frame = ttk.Frame(import_window)
        btn_frame.pack(pady=10)
        
        def import_from_local():
            """从本地文件导入"""
            import_window.destroy()
            file_path = filedialog.askopenfilename(
                title="选择物品ID文件",
                filetypes=[
                    ("Excel文件", "*.xlsx *.xlsm *.xls"),
                    ("文本文件", "*.txt"),
                    ("JSON文件", "*.json"),
                    ("所有文件", "*.*")
                ]
            )
            if file_path:
                self._do_import_item_ids(file_path, is_local=True)
        
        def import_from_server():
            """从服务器文件导入"""
            if not self.is_connected or not self.client:
                messagebox.showwarning("提示", "请先连接SSH服务器")
                import_window.destroy()
                return
            
            import_window.destroy()
            self._browse_server_file_for_import_item_ids()
        
        ttk.Button(btn_frame, text="从本地文件导入", command=import_from_local, width=18).pack(side=tk.LEFT, padx=10)
        
        # 只有在连接时才显示服务器导入按钮
        if self.is_connected and self.client:
            ttk.Button(btn_frame, text="从服务器文件导入", command=import_from_server, width=18).pack(side=tk.LEFT, padx=10)
        else:
            server_btn = ttk.Button(btn_frame, text="从服务器文件导入", command=import_from_server, width=18, state='disabled')
            server_btn.pack(side=tk.LEFT, padx=10)
            ttk.Label(import_window, text="(需要先连接SSH服务器)", font=("Microsoft YaHei", 8), foreground="gray").pack(pady=5)
        
        ttk.Button(import_window, text="取消", command=import_window.destroy).pack(pady=10)
    
    def _browse_server_file_for_import(self):
        """浏览服务器文件并选择导入"""
        if not self.is_connected or not self.client:
            messagebox.showwarning("提示", "请先连接SSH服务器")
            return
        
        # 创建文件浏览器窗口
        browse_window = tk.Toplevel(self.root)
        browse_window.title("从服务器选择GM配置文件")
        browse_window.geometry("800x500")
        browse_window.transient(self.root)
        browse_window.minsize(600, 400)
        
        # 使用grid布局支持自适应
        browse_window.columnconfigure(0, weight=1)
        browse_window.rowconfigure(0, weight=1)
        
        # 主框架
        main_frame = ttk.Frame(browse_window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # 路径导航
        nav_frame = ttk.Frame(main_frame)
        nav_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        nav_frame.columnconfigure(1, weight=1)
        
        ttk.Label(nav_frame, text="当前路径:").grid(row=0, column=0, padx=5)
        path_var = tk.StringVar(value="/www/wwwroot/")
        path_entry = ttk.Entry(nav_frame, textvariable=path_var)
        path_entry.grid(row=0, column=1, padx=5, sticky=(tk.W, tk.E))
        
        def browse_path():
            """浏览指定路径"""
            path = path_var.get().strip()
            if not path:
                path = "/"
            
            try:
                # 使用SFTP浏览
                sftp = self.client.open_sftp()
                try:
                    files = sftp.listdir_attr(path)
                except:
                    messagebox.showerror("错误", f"无法访问路径: {path}")
                    sftp.close()
                    return
                
                # 清空树
                for item in file_tree.get_children():
                    file_tree.delete(item)
                
                # 添加返回上级目录（如果不是根目录）
                if path != "/":
                    file_tree.insert("", 0, text="..", values=("目录", "", "", ""), tags=("dir",))
                
                # 添加文件和目录
                import stat
                for attr in sorted(files, key=lambda x: (not stat.S_ISDIR(x.st_mode), x.filename)):
                    name = attr.filename
                    if name.startswith('.'):
                        continue  # 跳过隐藏文件
                    
                    # 判断类型
                    if stat.S_ISDIR(attr.st_mode):
                        file_type = "目录"
                        tag = "dir"
                    else:
                        file_type = "文件"
                        tag = "file"
                    
                    # 大小
                    size = str(attr.st_size) if hasattr(attr, 'st_size') else ""
                    
                    # 修改时间
                    try:
                        from datetime import datetime
                        mtime = datetime.fromtimestamp(attr.st_mtime)
                        date = mtime.strftime("%Y-%m-%d %H:%M")
                    except:
                        date = ""
                    
                    file_tree.insert("", tk.END, text=name, values=(file_type, size, date), tags=(tag,))
                
                sftp.close()
            except Exception as e:
                messagebox.showerror("错误", f"浏览路径失败: {e}")
        
        ttk.Button(nav_frame, text="浏览", command=browse_path).grid(row=0, column=2, padx=5)
        
        # 文件列表
        list_frame = ttk.Frame(main_frame)
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        file_tree = ttk.Treeview(list_frame, columns=("type", "size", "date"), show="tree headings", height=20)
        file_tree.heading("#0", text="文件名")
        file_tree.heading("type", text="类型")
        file_tree.heading("size", text="大小")
        file_tree.heading("date", text="修改时间")
        file_tree.column("#0", width=300, stretch=True)
        file_tree.column("type", width=100, stretch=False)
        file_tree.column("size", width=100, stretch=False)
        file_tree.column("date", width=150, stretch=False)
        
        # 配置标签颜色
        file_tree.tag_configure("dir", foreground="blue")
        file_tree.tag_configure("file", foreground="black")
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=file_tree.yview)
        file_tree.configure(yscrollcommand=scrollbar.set)
        
        file_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        def on_file_double_click(event):
            """双击文件或目录"""
            selection = file_tree.selection()
            if not selection:
                return
            
            item = selection[0]
            name = file_tree.item(item, 'text')
            values = file_tree.item(item, 'values')
            file_type = values[0] if values else ""
            
            current_path = path_var.get().strip()
            if not current_path:
                current_path = "/"
            if not current_path.endswith("/"):
                current_path += "/"
            
            if name == "..":
                # 返回上级目录
                import os
                parent_path = os.path.dirname(current_path.rstrip("/"))
                if not parent_path:
                    parent_path = "/"
                path_var.set(parent_path)
                browse_path()
            elif file_type == "目录":
                # 进入目录
                new_path = current_path + name
                path_var.set(new_path)
                browse_path()
            else:
                # 选择文件导入
                file_path = current_path + name
                if name.lower().endswith('.json'):
                    if messagebox.askyesno("确认", f"确定要导入文件 '{file_path}' 吗？"):
                        browse_window.destroy()
                        self._do_import(file_path, is_local=False)
                else:
                    if messagebox.askyesno("确认", f"文件 '{name}' 不是JSON格式，确定要导入吗？"):
                        browse_window.destroy()
                        self._do_import(file_path, is_local=False)
        
        file_tree.bind("<Double-1>", on_file_double_click)
        
        # 按钮框架
        btn_frame = ttk.Frame(main_frame, padding="10")
        btn_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        
        def select_file():
            """选择当前选中的文件"""
            selection = file_tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择文件")
                return
            
            item = selection[0]
            name = file_tree.item(item, 'text')
            values = file_tree.item(item, 'values')
            file_type = values[0] if values else ""
            
            if file_type == "目录":
                messagebox.showwarning("提示", "请选择文件，不是目录")
                return
            
            current_path = path_var.get().strip()
            if not current_path:
                current_path = "/"
            if not current_path.endswith("/"):
                current_path += "/"
            
            file_path = current_path + name
            if messagebox.askyesno("确认", f"确定要导入文件 '{file_path}' 吗？"):
                browse_window.destroy()
                self._do_import(file_path, is_local=False)
        
        ttk.Button(btn_frame, text="选择并导入", command=select_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=browse_window.destroy).pack(side=tk.RIGHT, padx=5)
        
        # 初始浏览根目录
        browse_path()
    
    def _browse_server_file_for_import_item_ids(self):
        """浏览服务器文件并选择导入物品ID"""
        if not self.is_connected or not self.client:
            messagebox.showwarning("提示", "请先连接SSH服务器")
            return
        
        # 创建文件浏览器窗口
        browse_window = tk.Toplevel(self.root)
        browse_window.title("从服务器选择物品ID文件")
        browse_window.geometry("800x500")
        browse_window.transient(self.root)
        browse_window.minsize(600, 400)
        
        # 使用grid布局支持自适应
        browse_window.columnconfigure(0, weight=1)
        browse_window.rowconfigure(0, weight=1)
        
        # 主框架
        main_frame = ttk.Frame(browse_window, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # 路径导航
        nav_frame = ttk.Frame(main_frame)
        nav_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        nav_frame.columnconfigure(1, weight=1)
        
        ttk.Label(nav_frame, text="当前路径:").grid(row=0, column=0, padx=5)
        path_var = tk.StringVar(value="/www/wwwroot/")
        path_entry = ttk.Entry(nav_frame, textvariable=path_var)
        path_entry.grid(row=0, column=1, padx=5, sticky=(tk.W, tk.E))
        
        def browse_path():
            """浏览指定路径"""
            path = path_var.get().strip()
            if not path:
                path = "/"
            
            try:
                # 使用SFTP浏览
                sftp = self.client.open_sftp()
                try:
                    files = sftp.listdir_attr(path)
                except:
                    messagebox.showerror("错误", f"无法访问路径: {path}")
                    sftp.close()
                    return
                
                # 清空树
                for item in file_tree.get_children():
                    file_tree.delete(item)
                
                # 添加返回上级目录（如果不是根目录）
                if path != "/":
                    file_tree.insert("", 0, text="..", values=("目录", "", "", ""), tags=("dir",))
                
                # 添加文件和目录
                import stat
                for attr in sorted(files, key=lambda x: (not stat.S_ISDIR(x.st_mode), x.filename)):
                    name = attr.filename
                    if name.startswith('.'):
                        continue  # 跳过隐藏文件
                    
                    # 判断类型
                    if stat.S_ISDIR(attr.st_mode):
                        file_type = "目录"
                        tag = "dir"
                    else:
                        file_type = "文件"
                        tag = "file"
                    
                    # 大小
                    size = str(attr.st_size) if hasattr(attr, 'st_size') else ""
                    
                    # 修改时间
                    try:
                        from datetime import datetime
                        mtime = datetime.fromtimestamp(attr.st_mtime)
                        date = mtime.strftime("%Y-%m-%d %H:%M")
                    except:
                        date = ""
                    
                    file_tree.insert("", tk.END, text=name, values=(file_type, size, date), tags=(tag,))
                
                sftp.close()
            except Exception as e:
                messagebox.showerror("错误", f"浏览路径失败: {e}")
        
        ttk.Button(nav_frame, text="浏览", command=browse_path).grid(row=0, column=2, padx=5)
        
        # 文件列表
        list_frame = ttk.Frame(main_frame)
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        file_tree = ttk.Treeview(list_frame, columns=("type", "size", "date"), show="tree headings", height=20)
        file_tree.heading("#0", text="文件名")
        file_tree.heading("type", text="类型")
        file_tree.heading("size", text="大小")
        file_tree.heading("date", text="修改时间")
        file_tree.column("#0", width=300, stretch=True)
        file_tree.column("type", width=100, stretch=False)
        file_tree.column("size", width=100, stretch=False)
        file_tree.column("date", width=150, stretch=False)
        
        # 配置标签颜色
        file_tree.tag_configure("dir", foreground="blue")
        file_tree.tag_configure("file", foreground="black")
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=file_tree.yview)
        file_tree.configure(yscrollcommand=scrollbar.set)
        
        file_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        def on_file_double_click(event):
            """双击文件或目录"""
            selection = file_tree.selection()
            if not selection:
                return
            
            item = selection[0]
            name = file_tree.item(item, 'text')
            values = file_tree.item(item, 'values')
            file_type = values[0] if values else ""
            
            current_path = path_var.get().strip()
            if not current_path:
                current_path = "/"
            if not current_path.endswith("/"):
                current_path += "/"
            
            if name == "..":
                # 返回上级目录
                import os
                parent_path = os.path.dirname(current_path.rstrip("/"))
                if not parent_path:
                    parent_path = "/"
                path_var.set(parent_path)
                browse_path()
            elif file_type == "目录":
                # 进入目录
                new_path = current_path + name
                path_var.set(new_path)
                browse_path()
            else:
                # 选择文件导入
                file_path = current_path + name
                if messagebox.askyesno("确认", f"确定要导入文件 '{file_path}' 吗？"):
                    browse_window.destroy()
                    self._do_import_item_ids(file_path, is_local=False)
        
        file_tree.bind("<Double-1>", on_file_double_click)
        
        # 按钮框架
        btn_frame = ttk.Frame(main_frame, padding="10")
        btn_frame.grid(row=2, column=0, sticky=(tk.W, tk.E))
        
        def select_file():
            """选择当前选中的文件"""
            selection = file_tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择文件")
                return
            
            item = selection[0]
            name = file_tree.item(item, 'text')
            values = file_tree.item(item, 'values')
            file_type = values[0] if values else ""
            
            if file_type == "目录":
                messagebox.showwarning("提示", "请选择文件，不是目录")
                return
            
            current_path = path_var.get().strip()
            if not current_path:
                current_path = "/"
            if not current_path.endswith("/"):
                current_path += "/"
            
            file_path = current_path + name
            if messagebox.askyesno("确认", f"确定要导入文件 '{file_path}' 吗？"):
                browse_window.destroy()
                self._do_import_item_ids(file_path, is_local=False)
        
        ttk.Button(btn_frame, text="选择并导入", command=select_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=browse_window.destroy).pack(side=tk.RIGHT, padx=5)
        
        # 初始浏览根目录
        browse_path()
    
    def _read_excel_file(self, file_path):
        """读取Excel文件，提取编号+显示名（或ID+名称）"""
        item_ids = []
        try:
            # 尝试使用openpyxl读取（支持.xlsx和.xlsm）
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                # 读取第一个工作表
                ws = wb.active
                
                # 查找包含编号/ID和名称的列
                # 可能的列名：编号、ID、id、itemId、item_id、显示名、名称、name、显示名称等
                header_row = None
                id_col_idx = None
                name_col_idx = None
                
                # 尝试在前3行找到表头
                for row_idx in range(1, min(4, ws.max_row + 1)):
                    row = ws[row_idx]
                    for col_idx, cell in enumerate(row, 1):
                        cell_value = str(cell.value).strip() if cell.value else ""
                        cell_value_lower = cell_value.lower()
                        
                        # 检查是否是ID/编号列
                        if id_col_idx is None and any(keyword in cell_value_lower for keyword in ['编号', 'id', 'itemid', 'item_id', '物品id']):
                            id_col_idx = col_idx
                            header_row = row_idx
                        
                        # 检查是否是名称列
                        if name_col_idx is None and any(keyword in cell_value_lower for keyword in ['显示名', '名称', 'name', '显示名称', '物品名称']):
                            name_col_idx = col_idx
                            header_row = row_idx
                
                # 如果没找到表头，假设第一列是ID，第二列是名称
                if header_row is None:
                    header_row = 1
                    id_col_idx = 1
                    name_col_idx = 2
                
                # 从表头下一行开始读取数据
                start_row = header_row + 1
                for row_idx in range(start_row, ws.max_row + 1):
                    row = ws[row_idx]
                    id_value = None
                    name_value = None
                    
                    # 获取ID/编号
                    if id_col_idx and id_col_idx <= len(row):
                        id_cell = row[id_col_idx - 1]
                        if id_cell.value is not None:
                            id_value = str(id_cell.value).strip()
                    
                    # 获取名称
                    if name_col_idx and name_col_idx <= len(row):
                        name_cell = row[name_col_idx - 1]
                        if name_cell.value is not None:
                            name_value = str(name_cell.value).strip()
                    
                    # 组合ID和名称
                    if id_value:
                        if name_value:
                            # 格式：ID - 名称
                            item_id = f"{id_value} - {name_value}"
                        else:
                            # 只有ID
                            item_id = id_value
                        item_ids.append(item_id)
                    elif name_value:
                        # 只有名称，也添加
                        item_ids.append(name_value)
                
                wb.close()
                
            except ImportError:
                # 如果没有openpyxl，尝试使用xlrd（支持旧版.xls）
                try:
                    import xlrd
                    wb = xlrd.open_workbook(file_path)
                    ws = wb.sheet_by_index(0)
                    
                    # 查找表头
                    header_row = None
                    id_col_idx = None
                    name_col_idx = None
                    
                    for row_idx in range(min(3, ws.nrows)):
                        row = ws.row(row_idx)
                        for col_idx, cell in enumerate(row):
                            cell_value = str(cell.value).strip() if cell.value else ""
                            cell_value_lower = cell_value.lower()
                            
                            if id_col_idx is None and any(keyword in cell_value_lower for keyword in ['编号', 'id', 'itemid', 'item_id', '物品id']):
                                id_col_idx = col_idx
                                header_row = row_idx
                            
                            if name_col_idx is None and any(keyword in cell_value_lower for keyword in ['显示名', '名称', 'name', '显示名称', '物品名称']):
                                name_col_idx = col_idx
                                header_row = row_idx
                    
                    if header_row is None:
                        header_row = 0
                        id_col_idx = 0
                        name_col_idx = 1
                    
                    # 读取数据
                    start_row = header_row + 1
                    for row_idx in range(start_row, ws.nrows):
                        row = ws.row(row_idx)
                        id_value = None
                        name_value = None
                        
                        if id_col_idx is not None and id_col_idx < len(row):
                            if row[id_col_idx].value:
                                id_value = str(row[id_col_idx].value).strip()
                        
                        if name_col_idx is not None and name_col_idx < len(row):
                            if row[name_col_idx].value:
                                name_value = str(row[name_col_idx].value).strip()
                        
                        if id_value:
                            if name_value:
                                item_id = f"{id_value} - {name_value}"
                            else:
                                item_id = id_value
                            item_ids.append(item_id)
                        elif name_value:
                            item_ids.append(name_value)
                
                except ImportError:
                    messagebox.showerror("错误", "需要安装openpyxl或xlrd库来读取Excel文件\n请运行: pip install openpyxl")
                    return []
                except Exception as e:
                    messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
                    return []
            except Exception as e:
                messagebox.showerror("错误", f"读取Excel文件失败: {str(e)}")
                return []
        
        except Exception as e:
            messagebox.showerror("错误", f"处理Excel文件时出错: {str(e)}")
            return []
        
        return item_ids
    
    def _process_imported_item_ids(self, item_ids):
        """处理导入的物品ID列表"""
        if not item_ids:
            messagebox.showwarning("提示", "没有有效的物品ID")
            return
        
        # 确保使用当前IP的历史记录（按IP独立存储）
        current_ip_key = self.get_current_ip_key()
        
        # 确保历史记录列表已初始化
        if not hasattr(self, 'item_ids_history'):
            self.item_ids_history = []
        
        # 在导入前，先加载当前IP的历史记录，确保数据同步
        try:
            self.item_ids_history = []
            self.item_ids_all = []
            self.load_item_ids()
        except Exception as e:
            print(f"加载当前IP物品ID时出错: {e}")
            if not isinstance(self.item_ids_history, list):
                self.item_ids_history = []
                self.item_ids_all = []
        
        # 第一个物品ID设置为当前值（显示完整内容）
        first_id_full = item_ids[0]
        first_id_normalized = self._normalize_item_id(first_id_full)
        if first_id_normalized:
            display_str = self._format_item_id_for_display(first_id_normalized)
            first_display = display_str if display_str else first_id_full
        else:
            first_display = first_id_full
        
        # 获取导入前已有的物品ID集合（用于快速查找）
        existing_item_ids = set()
        existing_full_items = set()
        for item in self.item_ids_history:
            item_str = str(item).strip() if item else ""
            if item_str:
                existing_full_items.add(item_str)
                item_id_only = self.extract_item_id(item_str)
                if item_id_only and item_id_only.strip():
                    existing_item_ids.add(item_id_only.strip())
        
        # 统计实际导入的数量
        imported_count = 0  # 新增的数量
        updated_count = 0   # 更新的数量（已存在但更新位置）
        items_to_add = []
        processed_ids_in_import = set()
        
        # 遍历所有要导入的物品ID（确保全部处理）
        for item_id_full in item_ids:
            if not item_id_full or not item_id_full.strip():
                continue
            
            item_id_normalized = item_id_full.strip()
            if not item_id_normalized:
                continue
            
            # 提取ID部分（仅用于去重判断）
            item_id_only = None
            try:
                if ';' in item_id_normalized:
                    item_id_only = item_id_normalized.split(';')[0].strip()
                elif ' - ' in item_id_normalized:
                    item_id_only = item_id_normalized.split(' - ')[0].strip()
                else:
                    import re
                    numbers = re.findall(r'^\d+', item_id_normalized)
                    if numbers:
                        item_id_only = numbers[0]
                    else:
                        item_id_only = item_id_normalized
            except:
                item_id_only = item_id_normalized
            
            # 检查是否是新的物品ID
            if item_id_only:
                # 跳过本次导入中已处理过的（避免重复）
                if item_id_only in processed_ids_in_import:
                    continue
                
                processed_ids_in_import.add(item_id_only)
                
                # 检查是否在历史记录中已存在
                if item_id_only not in existing_item_ids:
                    # 新物品ID，添加到导入列表
                    items_to_add.append(item_id_normalized)
                    existing_item_ids.add(item_id_only)
                    imported_count += 1
                else:
                    # 已存在，移除旧的位置，添加到新位置（更新位置）
                    self.item_ids_history = [item for item in self.item_ids_history 
                                           if self.extract_item_id(item) != item_id_only]
                    items_to_add.append(item_id_normalized)
                    updated_count += 1
            else:
                # 没有提取到ID部分，使用完整内容进行去重
                if item_id_normalized not in existing_full_items:
                    # 不存在，添加到导入列表
                    if item_id_normalized not in processed_ids_in_import:
                        items_to_add.append(item_id_normalized)
                        existing_full_items.add(item_id_normalized)
                        processed_ids_in_import.add(item_id_normalized)
                        imported_count += 1
                else:
                    # 已存在，移除旧的位置，添加到新位置（更新位置）
                    if item_id_normalized in self.item_ids_history:
                        self.item_ids_history.remove(item_id_normalized)
                    items_to_add.append(item_id_normalized)
                    updated_count += 1
        
        # 将所有物品ID添加到历史记录的最前面（保持导入顺序）
        for item in reversed(items_to_add):
            self.item_ids_history.insert(0, item)
        
        # 更新下拉框
        if hasattr(self, 'item_id_combo'):
            self.item_ids_all = self.item_ids_history.copy()
            
            def update_dropdown():
                try:
                    if hasattr(self, 'item_ids_all') and self.item_ids_all:
                        self.item_id_var.set(first_display)
                        self.root.after(50, lambda: self.filter_item_ids(force_show_all=True))
                except Exception as e:
                    print(f"更新下拉框时出错: {e}")
            self.root.after(100, update_dropdown)
        
        # 保存到文件（按当前IP保存）
        try:
            self.save_item_ids()
        except Exception as e:
            print(f"保存物品ID到文件时出错: {e}")
        
        first_id_only = self.extract_item_id(first_id_full) if first_id_full else ""
        total_in_file = len(item_ids)
        total_processed = imported_count + updated_count  # 总共处理的物品ID数量
        total_in_history = len(self.item_ids_history) if hasattr(self, 'item_ids_history') else 0
        
        if total_processed == 0:
            if total_in_file > 0:
                messagebox.showwarning("提示", f"文件中有 {total_in_file} 个物品ID，但没有物品ID被导入（可能都已存在）")
            else:
                messagebox.showwarning("提示", "文件为空或没有有效的物品ID")
        else:
            msg_parts = [f"物品ID导入成功！"]
            msg_parts.append(f"文件中共有 {total_in_file} 个物品ID")
            if imported_count > 0:
                msg_parts.append(f"新增了 {imported_count} 个物品ID")
            if updated_count > 0:
                msg_parts.append(f"更新了 {updated_count} 个物品ID位置")
            msg_parts.append(f"当前物品ID: {first_id_only}")
            msg_parts.append(f"历史记录中共有 {total_in_history} 个物品ID")
            messagebox.showinfo("成功", "\n".join(msg_parts))
    
    def _do_import_item_ids(self, file_path, is_local=True):
        """执行物品ID导入操作"""
        try:
            # 检查文件扩展名，判断是否为Excel文件
            file_ext = os.path.splitext(file_path)[1].lower()
            is_excel = file_ext in ['.xlsx', '.xlsm', '.xls']
            
            if is_local:
                if is_excel:
                    # Excel文件处理
                    item_ids = self._read_excel_file(file_path)
                    if not item_ids:
                        messagebox.showerror("错误", "Excel文件中没有找到有效的物品ID数据")
                        return
                    # 直接使用从Excel读取的数据，跳过后续的文本/JSON处理
                    self._process_imported_item_ids(item_ids)
                    return
                else:
                    # 从本地文件读取（文本或JSON）
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
            else:
                # 从服务器文件读取
                if not self.is_connected or not self.client:
                    messagebox.showerror("错误", "SSH连接已断开")
                    return
                
                sftp = self.client.open_sftp()
                try:
                    # 使用二进制模式读取，然后解码
                    with sftp.open(file_path, 'rb') as f:
                        content_bytes = f.read()
                        if not content_bytes:
                            messagebox.showerror("错误", f"文件 '{file_path}' 为空或无法读取")
                            return
                        # 尝试UTF-8解码
                        try:
                            content = content_bytes.decode('utf-8')
                        except UnicodeDecodeError:
                            # 如果UTF-8失败，尝试GBK（中文系统常用）
                            try:
                                content = content_bytes.decode('gbk')
                            except UnicodeDecodeError:
                                # 最后尝试latin-1（不会失败）
                                content = content_bytes.decode('latin-1', errors='ignore')
                except FileNotFoundError:
                    messagebox.showerror("错误", f"文件 '{file_path}' 不存在")
                    return
                except PermissionError:
                    messagebox.showerror("错误", f"没有权限读取文件 '{file_path}'")
                    return
                except Exception as e:
                    messagebox.showerror("错误", f"读取文件 '{file_path}' 失败: {str(e)}")
                    return
                finally:
                    sftp.close()
            
            # 检查内容是否为空
            if not content or not content.strip():
                messagebox.showerror("错误", "文件内容为空，无法导入")
                return
            
            # 导入物品ID（支持多种格式）
            item_ids = []
            
            # 尝试解析为JSON格式
            try:
                # 移除BOM标记
                if content.startswith('\ufeff'):
                    content = content[1:]
                
                content_cleaned = content.strip()
                
                # 尝试解析JSON
                try:
                    json_data = json.loads(content_cleaned)
                    
                    # 如果是列表，直接使用
                    if isinstance(json_data, list):
                        for item in json_data:
                            if isinstance(item, (str, int)):
                                item_ids.append(str(item).strip())
                            elif isinstance(item, dict):
                                # 如果是字典，尝试提取ID字段
                                if 'id' in item:
                                    item_ids.append(str(item['id']).strip())
                                elif 'itemId' in item:
                                    item_ids.append(str(item['itemId']).strip())
                                elif 'item_id' in item:
                                    item_ids.append(str(item['item_id']).strip())
                                else:
                                    # 如果没有ID字段，使用第一个值
                                    values = list(item.values())
                                    if values:
                                        item_ids.append(str(values[0]).strip())
                    # 如果是字典，尝试提取值
                    elif isinstance(json_data, dict):
                        # 尝试提取所有值
                        for key, value in json_data.items():
                            if isinstance(value, (str, int)):
                                item_ids.append(str(value).strip())
                            elif isinstance(value, list):
                                for item in value:
                                    item_ids.append(str(item).strip())
                    else:
                        # 其他类型，转换为字符串
                        item_ids.append(str(json_data).strip())
                except json.JSONDecodeError:
                    # JSON解析失败，按文本处理
                    pass
            except Exception as e:
                # JSON解析出错，按文本处理
                print(f"JSON解析失败，按文本处理: {e}")
            
            # 如果不是JSON格式或JSON解析后没有数据，按文本格式处理
            if not item_ids:
                # 支持多行，每行一个物品ID
                lines = content.strip().split('\n')
                for line in lines:
                    line = line.strip()
                    if line:
                        # 支持制表符或空格分隔的多个ID
                        if '\t' in line:
                            parts = line.split('\t')
                            for part in parts:
                                part = part.strip()
                                if part:
                                    item_ids.append(part)
                        elif ' ' in line and not line.startswith(' '):
                            # 如果包含空格但不是以空格开头，可能是分隔符
                            # 先尝试按空格分割
                            parts = line.split()
                            if len(parts) > 1:
                                # 多个部分，都添加
                                for part in parts:
                                    part = part.strip()
                                    if part:
                                        item_ids.append(part)
                            else:
                                # 单个部分，直接添加
                                item_ids.append(line)
                        else:
                            # 单行，直接添加
                            item_ids.append(line)
            
            # 去重但保持顺序
            seen = set()
            unique_item_ids = []
            for item_id in item_ids:
                if item_id and item_id not in seen:
                    seen.add(item_id)
                    unique_item_ids.append(item_id)
            
            item_ids = unique_item_ids
            
            if not item_ids:
                messagebox.showwarning("提示", "文件中没有有效的物品ID")
                return
            
            # 使用统一的处理函数
            self._process_imported_item_ids(item_ids)
        except Exception as e:
            error_msg = f"导入物品ID失败: {str(e)}"
            messagebox.showerror("导入失败", error_msg)
            import traceback
            traceback.print_exc()
    
    def _do_import(self, file_path, is_local=True):
        """执行导入操作"""
        try:
            if is_local:
                # 从本地文件读取
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                # 从服务器文件读取
                if not self.is_connected or not self.client:
                    messagebox.showerror("错误", "SSH连接已断开")
                    return
                
                sftp = self.client.open_sftp()
                try:
                    # 使用二进制模式读取，然后解码
                    with sftp.open(file_path, 'rb') as f:
                        content_bytes = f.read()
                        if not content_bytes:
                            messagebox.showerror("错误", f"文件 '{file_path}' 为空或无法读取")
                            return
                        # 尝试UTF-8解码
                        try:
                            content = content_bytes.decode('utf-8')
                        except UnicodeDecodeError:
                            # 如果UTF-8失败，尝试GBK（中文系统常用）
                            try:
                                content = content_bytes.decode('gbk')
                            except UnicodeDecodeError:
                                # 最后尝试latin-1（不会失败）
                                content = content_bytes.decode('latin-1', errors='ignore')
                except FileNotFoundError:
                    messagebox.showerror("错误", f"文件 '{file_path}' 不存在")
                    return
                except PermissionError:
                    messagebox.showerror("错误", f"没有权限读取文件 '{file_path}'")
                    return
                except Exception as e:
                    messagebox.showerror("错误", f"读取文件 '{file_path}' 失败: {str(e)}")
                    return
                finally:
                    sftp.close()
            
            # 尝试解析JSON（处理可能的格式问题）
            imported_data = None
            
            # 步骤1: 移除BOM标记
            if content.startswith('\ufeff'):
                content = content[1:]
            
            # 步骤2: 移除前后空白字符
            content = content.strip()
            
            # 检查内容是否为空
            if not content:
                messagebox.showerror("错误", f"文件内容为空，无法导入")
                return
            
            # 步骤3: 尝试直接解析
            try:
                imported_data = json.loads(content)
            except json.JSONDecodeError as e:
                # 步骤4: 如果失败，尝试清理注释
                lines = content.split('\n')
                cleaned_lines = []
                for line in lines:
                    # 移除行注释（简单处理，不处理字符串中的//）
                    if '//' in line:
                        comment_pos = line.find('//')
                        # 检查是否在字符串中
                        in_string = False
                        quote_char = None
                        for i in range(comment_pos):
                            char = line[i]
                            if char in ['"', "'"] and (i == 0 or line[i-1] != '\\'):
                                if quote_char is None:
                                    quote_char = char
                                    in_string = True
                                elif char == quote_char:
                                    in_string = False
                                    quote_char = None
                        if not in_string:
                            line = line[:comment_pos].rstrip()
                    cleaned_lines.append(line)
                content = '\n'.join(cleaned_lines).strip()
                
                # 步骤5: 再次尝试解析
                try:
                    imported_data = json.loads(content)
                except json.JSONDecodeError as e2:
                    # 步骤6: 检查是否是"Extra data"错误（说明第一个JSON对象可能是有效的）
                    error_msg = str(e2).lower()
                    is_extra_data = "extra data" in error_msg
                    
                    # 如果是Extra data错误，尝试提取第一个有效的JSON对象
                    if is_extra_data:
                        # 尝试使用JSONDecoder的raw_decode方法，它可以从字符串开头解析一个JSON对象
                        # 即使后面有额外数据也能成功
                        try:
                            decoder = json.JSONDecoder()
                            imported_data, idx = decoder.raw_decode(content)
                            # 成功解析，即使后面有额外数据也忽略
                        except json.JSONDecodeError:
                            # raw_decode也失败，继续使用原来的方法
                            is_extra_data = False
                    
                    # 如果不是Extra data错误，或者raw_decode失败，使用原来的提取方法
                    if not is_extra_data or imported_data is None:
                        # 步骤7: 尝试提取第一个完整的JSON对象
                        # 查找第一个 { 的位置
                        first_brace = content.find('{')
                        if first_brace == -1:
                            first_brace = content.find('[')
                        
                        if first_brace != -1:
                            # 从第一个 { 或 [ 开始，找到匹配的结束位置
                            brace_count = 0
                            bracket_count = 0
                            in_string = False
                            escape_next = False
                            quote_char = None
                            start_char = content[first_brace]
                            end_char = '}' if start_char == '{' else ']'
                            
                            for i in range(first_brace, len(content)):
                                char = content[i]
                                
                                if escape_next:
                                    escape_next = False
                                    continue
                                
                                if char == '\\':
                                    escape_next = True
                                    continue
                                
                                if char in ['"', "'"]:
                                    if quote_char is None:
                                        quote_char = char
                                        in_string = True
                                    elif char == quote_char:
                                        in_string = False
                                        quote_char = None
                                    continue
                                
                                if not in_string:
                                    if char == '{':
                                        brace_count += 1
                                    elif char == '}':
                                        brace_count -= 1
                                    elif char == '[':
                                        bracket_count += 1
                                    elif char == ']':
                                        bracket_count -= 1
                                    
                                    # 找到匹配的结束位置
                                    if (start_char == '{' and brace_count == 0) or (start_char == '[' and bracket_count == 0):
                                        extracted_content = content[first_brace:i+1]
                                        try:
                                            imported_data = json.loads(extracted_content)
                                            break
                                        except json.JSONDecodeError:
                                            # 如果提取的部分还是无效，继续尝试
                                            pass
                            
                            # 如果还是失败，抛出原始错误
                            if imported_data is None:
                                # 计算错误位置（行号和列号）
                                error_line = 1
                                error_col = 1
                                if hasattr(e2, 'lineno') and hasattr(e2, 'colno'):
                                    error_line = e2.lineno
                                    error_col = e2.colno
                                else:
                                    # 手动计算行号和列号
                                    lines_before = content[:e2.pos].split('\n')
                                    error_line = len(lines_before)
                                    error_col = len(lines_before[-1]) + 1
                                
                                raise json.JSONDecodeError(
                                    f"JSON解析失败: 未找到有效的JSON对象。原始错误: {str(e2)}",
                                    content,
                                    e2.pos
                                )
                        else:
                            # 计算错误位置
                            error_line = 1
                            error_col = 1
                            if hasattr(e2, 'lineno') and hasattr(e2, 'colno'):
                                error_line = e2.lineno
                                error_col = e2.colno
                            else:
                                lines_before = content[:e2.pos].split('\n')
                                error_line = len(lines_before)
                                error_col = len(lines_before[-1]) + 1
                            
                            raise json.JSONDecodeError(
                                f"JSON解析失败: 未找到有效的JSON对象。原始错误: {str(e2)}",
                                content,
                                e2.pos
                            )
            
            # 合并到现有模板（扁平化处理）
            templates = {}
            if os.path.exists(self.gm_templates_file):
                try:
                    with open(self.gm_templates_file, 'r', encoding='utf-8') as f:
                        templates = json.load(f)
                except:
                    templates = {}
            
            # 处理导入的数据（支持分类格式和扁平格式）
            flat_commands = {}
            
            # 先提取现有命令（扁平化）
            if isinstance(templates, dict):
                for category, commands in templates.items():
                    if isinstance(commands, dict):
                        for cmd_name, cmd_info in commands.items():
                            if isinstance(cmd_info, dict):
                                flat_commands[cmd_name] = {
                                    'command': cmd_info.get('command', ''),
                                    'params': cmd_info.get('params', '')
                                }
                            else:
                                flat_commands[cmd_name] = {
                                    'command': str(cmd_info),
                                    'params': ''
                                }
            
            # 合并导入的数据
            if isinstance(imported_data, dict):
                for category, commands in imported_data.items():
                    if isinstance(commands, dict):
                        for cmd_name, cmd_info in commands.items():
                            if isinstance(cmd_info, dict):
                                flat_commands[cmd_name] = {
                                    'command': cmd_info.get('command', ''),
                                    'params': cmd_info.get('params', '')
                                }
                            else:
                                flat_commands[cmd_name] = {
                                    'command': str(cmd_info),
                                    'params': ''
                                }
            
            # 检查是否成功解析
            if imported_data is None:
                raise ValueError("无法解析JSON文件，请检查文件格式是否正确")
            
            # 保存（使用扁平化保存）
            self._save_flat_commands(flat_commands)
            
            # 重新加载模板
            self.load_gm_templates()
            
            messagebox.showinfo("成功", f"GM配置导入成功！\n共导入 {len(flat_commands)} 个命令")
        except json.JSONDecodeError as e:
            error_msg = f"JSON解析失败: {str(e)}"
            if hasattr(e, 'pos') and e.pos is not None:
                try:
                    # 尝试计算错误位置
                    line_num = content[:e.pos].count('\n') + 1
                    last_newline = content[:e.pos].rfind('\n')
                    col_num = e.pos - last_newline if last_newline != -1 else e.pos + 1
                    error_msg += f"\n错误位置: 第 {line_num} 行，第 {col_num} 列"
                except:
                    pass
            error_msg += "\n\n请检查JSON文件格式是否正确。\n提示：JSON文件应该是一个有效的JSON对象或数组。"
            messagebox.showerror("导入失败", error_msg)
            import traceback
            traceback.print_exc()
        except Exception as e:
            error_msg = f"导入GM配置失败: {str(e)}"
            messagebox.showerror("导入失败", error_msg)
            import traceback
            traceback.print_exc()
    
    def clear_gm_fields(self):
        """清除GM字段"""
        if messagebox.askyesno("确认", "确定要清除所有GM字段吗？"):
            if hasattr(self, 'gm_cmd_var'):
                self.gm_cmd_var.set("")
            if hasattr(self, 'item_id_var'):
                self.item_id_var.set("")
            if hasattr(self, 'item_amount_var'):
                self.item_amount_var.set("1")
    
    def send_gm_command(self):
        """发送GM命令到服务器"""
        # 检查连接状态
        if not self.is_connected or not self.client:
            messagebox.showwarning("提示", "请先连接SSH服务器")
            return
        
        # 获取参数
        jar_path = self.get_full_jar_path()
        port = self.gm_port_var.get().strip()
        roleid = self.roleid_var.get().strip()
        command_name = self.gm_cmd_var.get().strip()
        item_id = self.item_id_var.get().strip()
        amount = self.item_amount_var.get().strip()
        
        # 验证必填参数
        if not command_name:
            messagebox.showwarning("提示", "请输入命令")
            return
        
        # 从模板映射中获取实际命令内容
        # 如果输入的是命令名（在模板中），则获取对应的命令内容
        # 如果输入的是命令内容（不在模板中），则直接使用
        command = command_name
        if hasattr(self, 'gm_templates_map') and self.gm_templates_map:
            # 检查是否是命令名
            if command_name in self.gm_templates_map:
                # 是命令名，获取对应的命令内容
                command = self.gm_templates_map[command_name]
                if not command:
                    # 如果映射为空，使用命令名本身
                    command = command_name
            # 如果不是命令名，command 已经是命令内容，直接使用
        
        if not port:
            messagebox.showwarning("提示", "端口不能为空")
            return
        
        # 获取UserID（从userid_var获取，如果没有则使用默认值）
        userid = "4096"
        if hasattr(self, 'userid_var'):
            userid_val = self.userid_var.get().strip()
            if userid_val:
                userid = userid_val
        
        if not roleid:
            messagebox.showwarning("提示", "RoleID不能为空")
            return
        
        # IP地址固定使用127.0.0.1
        ip = "127.0.0.1"
        
        # 构建GM命令内容（根据命令类型决定是否添加物品ID和数量）
        # 需要物品ID的命令：additem, giveitem 等
        # 不需要物品ID的命令：addlevel, addexp, addgold 等
        gm_command_content = command
        
        # 判断命令是否需要物品ID
        commands_need_item = ['additem', 'giveitem', 'removeitem', 'delitem']
        command_lower = command.lower().strip()
        needs_item = any(cmd in command_lower for cmd in commands_need_item)
        
        if needs_item and item_id:
            # 提取纯物品ID（去掉描述信息）
            # 物品ID格式可能是: "40009" 或 "40009 - 朱雀石" 或 "40009;描述"
            item_id_clean = item_id.strip()
            # 如果包含分隔符（空格、分号、横线等），提取前面的数字部分
            for separator in [' ', ';', '-', '：', ':']:
                if separator in item_id_clean:
                    item_id_clean = item_id_clean.split(separator)[0].strip()
                    break
            
            # 只保留数字部分（去掉所有非数字字符，但保留第一个数字串）
            import re
            match = re.search(r'\d+', item_id_clean)
            if match:
                item_id_clean = match.group()
            
            if item_id_clean:
                gm_command_content += " " + item_id_clean
                
                # 数量参数处理
                # 对于additem等需要数量的命令，如果没有数量则使用默认值1
                amount_value = amount.strip() if amount and amount.strip() else "1"
                gm_command_content += " " + amount_value
        elif not needs_item:
            # 对于不需要物品ID的命令（如addlevel, addexp），可能需要其他参数
            # 如果命令需要参数但没有物品ID，可以添加数量或其他参数
            # 例如：addlevel 10, addexp 1000
            if amount and amount.strip():
                gm_command_content += " " + amount.strip()
        
        # 构建完整命令
        # 格式: java -jar /path/to/jar "" "" 127.0.0.1 port gm userId=xxx roleId=xxx "command content"
        # 使用shell命令字符串格式
        full_command = f'java -jar {jar_path} "" "" {ip} {port} gm userId={userid} roleId={roleid} "{gm_command_content}"'
        
        # 在后台线程中执行命令
        def execute_gm_command():
            try:
                # 执行命令（使用shell=True确保命令在shell中执行，这样可以正确处理空字符串参数）
                # 注意：paramiko的exec_command默认使用shell执行，所以空字符串参数应该能正确处理
                stdin, stdout, stderr = self.client.exec_command(full_command, timeout=30)
                stdout.channel.settimeout(30)
                stderr.channel.settimeout(30)
                
                # 读取输出
                output = ""
                error = ""
                
                import time
                start_time = time.time()
                timeout_seconds = 30
                
                # 等待命令完成或超时
                while not stdout.channel.exit_status_ready() and (time.time() - start_time) < timeout_seconds:
                    if stdout.channel.recv_ready():
                        data = stdout.channel.recv(4096)
                        if data:
                            output += data.decode('utf-8', errors='ignore')
                    if stderr.channel.recv_ready():
                        data = stderr.channel.recv(4096)
                        if data:
                            error += data.decode('utf-8', errors='ignore')
                    time.sleep(0.1)
                
                # 读取剩余输出
                while stdout.channel.recv_ready():
                    data = stdout.channel.recv(4096)
                    if data:
                        output += data.decode('utf-8', errors='ignore')
                
                while stderr.channel.recv_ready():
                    data = stderr.channel.recv(4096)
                    if data:
                        error += data.decode('utf-8', errors='ignore')
                
                # 获取退出状态
                exit_status = stdout.channel.recv_exit_status()
                
                # 在主线程中显示结果
                def show_result():
                    # 只显示实际的输出和错误，不显示GM命令相关的提示
                    if output:
                        self.output_queue.put(("output", f"{output}\n"))
                    if error:
                        self.output_queue.put(("error", f"{error}\n"))
                    
                    if exit_status == 0:
                        messagebox.showinfo("成功", "GM命令发送成功！")
                    else:
                        messagebox.showwarning("提示", f"GM命令执行完成，退出码: {exit_status}")
                
                self.root.after(0, show_result)
                
            except Exception as e:
                error_msg = f"发送GM命令失败: {str(e)}"
                self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
                import traceback
                traceback.print_exc()
        
        # 启动后台线程
        threading.Thread(target=execute_gm_command, daemon=True).start()
    
    # ==================== 兼容性/占位事件处理 ====================
    def on_gm_cmd_selected(self, *args, **kwargs):
        """
        占位：GM命令模板选择事件处理。
        某些版本的界面在绑定下拉选择事件时会调用该方法。
        为保证程序能正常启动，这里提供一个安全的空实现。
        """
        try:
            # 如果存在模板应用方法，优先调用（兼容旧实现）
            if hasattr(self, "apply_selected_gm_template"):
                self.apply_selected_gm_template()
        except Exception:
            # 保底不抛出异常，避免阻塞启动
            pass
    
    def on_gm_cmd_changed(self, *args, **kwargs):
        """
        占位：GM命令内容变更事件处理。
        可能由 StringVar.trace/文本输入事件调用，这里做兼容处理。
        """
        try:
            if hasattr(self, "update_gm_command_preview"):
                self.update_gm_command_preview()
        except Exception:
            pass
    
    # ==================== 输出面板设置 ====================
    def copy_text(self):
        """复制选中文本"""
        try:
            if self.cmd_output_text.tag_ranges(tk.SEL):
                text = self.cmd_output_text.get(tk.SEL_FIRST, tk.SEL_LAST)
                self.root.clipboard_clear()
                self.root.clipboard_append(text)
        except:
            pass
    
    def paste_text(self):
        """粘贴文本"""
        try:
            text = self.root.clipboard_get()
            if text:
                # 在光标位置插入文本
                self.cmd_output_text.insert(tk.INSERT, text)
        except:
            pass
    
    def show_output_context_menu(self, event):
        """显示输出面板右键菜单"""
        try:
            self.output_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.output_context_menu.grab_release()
    def show_output_settings(self):
        """显示输出面板设置（背景颜色和文字颜色）"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("输出面板设置")
        settings_window.geometry("450x550")
        settings_window.transient(self.root)
        settings_window.grab_set()
        settings_window.configure(bg="#f5f5f5")
        
        # 主容器
        main_frame = tk.Frame(settings_window, bg="#f5f5f5", padx=20, pady=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 输出面板设置内容
        output_main_frame = tk.Frame(main_frame, bg="#f5f5f5")
        output_main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 背景颜色设置
        bg_frame = ttk.LabelFrame(output_main_frame, text="背景颜色", padding="10")
        bg_frame.pack(fill=tk.X, pady=(0, 15))
        
        color_frame = ttk.Frame(bg_frame)
        color_frame.pack(fill=tk.X, pady=5)
        
        # 预设颜色选项
        colors = [
            ("深色（默认）", "#1e1e1e"),
            ("黑色", "#000000"),
            ("深蓝", "#001122"),
            ("深灰", "#2d2d2d"),
            ("浅色", "#ffffff"),
            ("浅灰", "#f0f0f0"),
        ]
        
        current_color = self.output_bg_color
        color_var = tk.StringVar(value=current_color)
        
        for name, color_code in colors:
            frame = ttk.Frame(color_frame)
            frame.pack(pady=3, padx=10, fill=tk.X)
            ttk.Radiobutton(frame, text=name, variable=color_var, value=color_code).pack(side=tk.LEFT, padx=5)
            # 显示颜色预览
            color_preview = tk.Label(frame, bg=color_code, width=4, relief=tk.SOLID, borderwidth=1, height=1)
            color_preview.pack(side=tk.LEFT, padx=5)
        
        # 自定义背景颜色
        custom_bg_frame = ttk.Frame(bg_frame)
        custom_bg_frame.pack(fill=tk.X, pady=5)
        ttk.Label(custom_bg_frame, text="自定义背景颜色:").pack(side=tk.LEFT, padx=5)
        custom_color_var = tk.StringVar(value=current_color)
        custom_entry = ttk.Entry(custom_bg_frame, textvariable=custom_color_var, width=12)
        custom_entry.pack(side=tk.LEFT, padx=5)
        
        # 文字颜色设置
        text_color_frame = ttk.LabelFrame(output_main_frame, text="文字颜色", padding="10")
        text_color_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 初始化文字颜色字典（如果不存在）
        if not hasattr(self, 'output_text_colors'):
            self.output_text_colors = {
                'success': '#00ff00',
                'error': '#ff0000',
                'info': '#00aaff',
                'warning': '#ffaa00',
                'command': '#ffff00',
                'output': '#ffffff'
            }
        
        # 文字颜色选项
        text_colors = [
            ("成功消息（绿色）", "success", self.output_text_colors.get('success', '#00ff00')),
            ("错误消息（红色）", "error", self.output_text_colors.get('error', '#ff0000')),
            ("信息消息（蓝色）", "info", self.output_text_colors.get('info', '#00aaff')),
            ("警告消息（橙色）", "warning", self.output_text_colors.get('warning', '#ffaa00')),
            ("命令（黄色）", "command", self.output_text_colors.get('command', '#ffff00')),
            ("普通输出（白色）", "output", self.output_text_colors.get('output', '#ffffff')),
        ]
        
        text_color_vars = {}
        for label_text, color_key, default_color in text_colors:
            frame = ttk.Frame(text_color_frame)
            frame.pack(pady=3, padx=10, fill=tk.X)
            ttk.Label(frame, text=label_text, width=18).pack(side=tk.LEFT, padx=5)
            var = tk.StringVar(value=default_color)
            text_color_vars[color_key] = var
            entry = ttk.Entry(frame, textvariable=var, width=12)
            entry.pack(side=tk.LEFT, padx=5)
            # 颜色预览
            preview = tk.Label(frame, bg=default_color, width=4, relief=tk.SOLID, borderwidth=1, height=1)
            preview.pack(side=tk.LEFT, padx=5)
            
            # 绑定颜色变化，实时更新预览
            def update_preview(color_key=color_key, preview_label=preview, var_ref=var):
                try:
                    color = var_ref.get().strip()
                    if color.startswith('#') and len(color) == 7:
                        preview_label.config(bg=color)
                except:
                    pass
            
            var.trace_add('write', lambda *args: update_preview())
        
        def apply_settings():
            # 应用背景颜色
            selected_color = color_var.get()
            if selected_color:
                self.output_bg_color = selected_color
                self.cmd_output_text.config(bg=selected_color)
            
            # 应用文字颜色
            for color_key, var in text_color_vars.items():
                color_value = var.get().strip()
                if color_value.startswith('#') and len(color_value) == 7:
                    self.output_text_colors[color_key] = color_value
                    # 更新标签配置
                    self.cmd_output_text.tag_config(color_key, foreground=color_value, background=self.output_bg_color)
            
            settings_window.destroy()
            messagebox.showinfo("提示", "设置已应用")
        
        def apply_custom_bg_color():
            custom_color = custom_color_var.get().strip()
            if custom_color:
                try:
                    if custom_color.startswith('#') and len(custom_color) == 7:
                        self.output_bg_color = custom_color
                        self.cmd_output_text.config(bg=custom_color)
                        color_var.set(custom_color)
                        messagebox.showinfo("提示", "背景颜色已更改")
                    else:
                        messagebox.showerror("错误", "颜色代码格式错误，请使用 #RRGGBB 格式（如 #1e1e1e）")
                except:
                    messagebox.showerror("错误", "无效的颜色代码")
        
        # 按钮框架
        btn_frame = ttk.Frame(output_main_frame)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="应用所有设置", command=apply_settings, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="应用背景色", command=apply_custom_bg_color, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=settings_window.destroy, width=15).pack(side=tk.LEFT, padx=5)
    
    # ==================== JAR路径管理（支持双路径格式，自动判断和查找）====================
    def get_full_jar_path(self):
        """获取完整的JAR路径（支持两种格式，自动判断，支持自动查找）"""
        user_input = self.jar_path_var.get().strip()
        if not user_input:
            user_input = "game"  # 默认值
        
        # 如果用户输入的是完整路径（以/开头或以.jar结尾），直接使用
        if user_input.startswith("/") or user_input.lower().endswith(".jar"):
            return user_input
        
        # 支持三种路径格式，自动判断：
        # 格式1: /www/wwwroot/{project}/static/api/jmxc.jar (旧格式，如game)
        # 格式2: /www/wwwroot/{project}/gm/gm/jmxc.jar (新格式，如mhxy1)
        # 格式3: 自动在 /www/wwwroot/{project}/ 目录下查找 jmxc.jar
        
        # 如果连接了SSH，尝试检查文件是否存在来判断使用哪种格式
        if self.is_connected and self.client:
            try:
                sftp = self.client.open_sftp()
                
                # 方法1: 先尝试新格式
                new_format_path = f"{self.jar_path_prefix}{user_input}/gm/gm/jmxc.jar"
                try:
                    sftp.stat(new_format_path)
                    sftp.close()
                    return new_format_path
                except:
                    pass
                
                # 方法2: 尝试旧格式
                old_format_path = f"{self.jar_path_prefix}{user_input}/static/api/jmxc.jar"
                try:
                    sftp.stat(old_format_path)
                    sftp.close()
                    return old_format_path
                except:
                    pass
                
                # 方法3: 自动在 /www/wwwroot/{project}/ 目录下递归查找 jmxc.jar
                project_path = f"{self.jar_path_prefix}{user_input}"
                found_path = self._find_jar_in_directory(sftp, project_path, "jmxc.jar")
                if found_path:
                    sftp.close()
                    return found_path
                
                sftp.close()
            except Exception as e:
                # 如果查找过程中出错，继续使用默认判断逻辑
                pass
        
        # 如果未连接或检查失败，根据项目名判断（mhxy1等使用新格式，game等使用旧格式）
        # 可以根据实际需求调整判断逻辑
        if user_input.lower() in ['mhxy1', 'mhxy2', 'mhxy3'] or 'mhxy' in user_input.lower():
            # 新格式
            return f"{self.jar_path_prefix}{user_input}/gm/gm/jmxc.jar"
        else:
            # 旧格式（默认）
            return f"{self.jar_path_prefix}{user_input}/static/api/jmxc.jar"
    
    def _find_jar_in_directory(self, sftp, directory, filename):
        """在指定目录下递归查找JAR文件"""
        try:
            # 检查目录是否存在
            try:
                sftp.stat(directory)
            except:
                return None
            
            # 列出目录内容
            try:
                items = sftp.listdir_attr(directory)
            except:
                return None
            
            # 遍历目录内容
            for item in items:
                item_path = f"{directory}/{item.filename}"
                
                # 如果是文件且名称匹配
                if stat.S_ISREG(item.st_mode) and item.filename == filename:
                    return item_path
                
                # 如果是目录，递归查找（限制深度，避免过深）
                if stat.S_ISDIR(item.st_mode):
                    # 跳过一些常见的系统目录，提高效率
                    if item.filename in ['.', '..', 'node_modules', '.git', 'vendor', 'cache', 'logs']:
                        continue
                    
                    # 递归查找（限制深度为5层）
                    if directory.count('/') - self.jar_path_prefix.count('/') < 5:
                        found = self._find_jar_in_directory(sftp, item_path, filename)
                        if found:
                            return found
            
            return None
        except Exception as e:
            # 查找过程中出错，返回None
            return None
    
    # ==================== 物品ID历史记录管理（按IP独立存储）====================
    def get_current_ip_key(self):
        """获取当前连接的IP作为键"""
        if hasattr(self, 'host_var'):
            host = self.host_var.get().strip()
            if host:
                return host
        return "default"
    
    def extract_item_id(self, item_id_str):
        """提取物品ID（从ID;名称或ID - 名称格式中提取纯ID）"""
        if not item_id_str:
            return ""
        # 如果包含 " - "（显示格式），提取ID部分
        if ' - ' in item_id_str:
            return item_id_str.split(' - ')[0].strip()
        # 如果包含分号（内部格式），只取分号前的部分
        if ';' in item_id_str:
            return item_id_str.split(';')[0].strip()
        # 否则直接返回（可能是纯ID）
        return item_id_str.strip()
    
    def _safe_load_item_ids(self):
        """安全地加载物品ID（防止递归）"""
        try:
            # 如果正在初始化、筛选、打开下拉框或更新值，不加载
            if (self._initializing or self._filtering_active or 
                self._opening_dropdown or self._updating_values):
                return
            self.load_item_ids()
        except Exception as e:
            print(f"加载物品ID时出错: {e}")
            import traceback
            traceback.print_exc()
    
    def load_item_ids(self):
        """加载当前IP的物品ID历史记录（不触发搜索事件，按IP独立加载）"""
        # 获取当前IP键（用于按IP独立存储）
        current_ip_key = self.get_current_ip_key()
        
        # 如果正在初始化，只加载数据，不更新UI
        if self._initializing:
            try:
                ip_key = self.get_current_ip_key()
                
                # 加载数据
                if os.path.exists(self.item_ids_file):
                    try:
                        with open(self.item_ids_file, 'r', encoding='utf-8') as f:
                            all_item_ids = json.load(f)
                            # 如果是旧格式（列表），转换为新格式（字典）
                            if isinstance(all_item_ids, list):
                                all_item_ids = {"default": all_item_ids}
                            # 获取当前IP的物品ID列表
                            item_ids_raw = all_item_ids.get(ip_key, [])
                            # 确保是列表类型
                            if isinstance(item_ids_raw, list):
                                self.item_ids_history = item_ids_raw
                            else:
                                self.item_ids_history = []
                    except Exception as e:
                        print(f"加载物品ID历史记录失败: {e}")
                        self.item_ids_history = []
                else:
                    self.item_ids_history = []
                
                # 标准化和去重（不更新UI）
                unique_ids = []
                seen_ids = set()
                
                try:
                    # 确保 item_ids_history 是列表
                    if not isinstance(self.item_ids_history, list):
                        self.item_ids_history = []
                    
                    for item_id_full in self.item_ids_history:
                        try:
                            if item_id_full is None:
                                continue
                            item_id_full_str = self._normalize_item_id(item_id_full)
                            if not item_id_full_str:
                                continue
                            if not self._is_valid_item_id(item_id_full_str):
                                continue
                            # 使用"完整字符串"去重，允许同一ID的不同名称并存
                            if item_id_full_str not in seen_ids:
                                unique_ids.append(item_id_full_str)
                                seen_ids.add(item_id_full_str)
                        except Exception:
                            continue
                except Exception as e:
                    print(f"标准化物品ID时出错: {e}")
                    unique_ids = []
                
                # 更新数据（不更新UI）
                self.item_ids_history = unique_ids
                self.item_ids_all = unique_ids.copy() if unique_ids else []
                # 重置搜索文本
                self._last_search_text = ""
                # 注意：初始化阶段不清空输入框，因为此时界面可能还没创建
            except Exception as e:
                print(f"初始化时加载物品ID失败: {e}")
            return
        
        # 正常加载（初始化完成后）
        ip_key = self.get_current_ip_key()
        
        # 加载数据（按IP独立加载）
        if os.path.exists(self.item_ids_file):
            try:
                with open(self.item_ids_file, 'r', encoding='utf-8') as f:
                    all_item_ids = json.load(f)
                    # 如果是旧格式（列表），转换为新格式（字典）
                    if isinstance(all_item_ids, list):
                        all_item_ids = {"default": all_item_ids}
                    # 获取当前IP的物品ID列表（每个IP独立存储）
                    item_ids_raw = all_item_ids.get(ip_key, [])
                    # 确保是列表类型
                    if isinstance(item_ids_raw, list):
                        self.item_ids_history = item_ids_raw
                    else:
                        self.item_ids_history = []
                    # 输出调试信息（可选）
                    print(f"已加载IP {ip_key} 的物品ID，共 {len(self.item_ids_history)} 个")
            except Exception as e:
                print(f"加载物品ID历史记录失败: {e}")
                import traceback
                traceback.print_exc()
                self.item_ids_history = []
        else:
            self.item_ids_history = []
        
        # 标准化和去重
        if hasattr(self, 'item_id_combo'):
            unique_ids = []
            seen_ids = set()
            
            try:
                # 确保 item_ids_history 是列表
                if not isinstance(self.item_ids_history, list):
                    self.item_ids_history = []
                
                for item_id_full in self.item_ids_history:
                    try:
                        if item_id_full is None:
                            continue
                        item_id_full_str = self._normalize_item_id(item_id_full)
                        if not item_id_full_str:
                            continue
                        if not self._is_valid_item_id(item_id_full_str):
                            continue
                        # 使用"完整字符串"去重，允许同一ID的不同名称并存
                        if item_id_full_str not in seen_ids:
                            unique_ids.append(item_id_full_str)
                            seen_ids.add(item_id_full_str)
                    except Exception as e:
                        print(f"处理物品ID时出错: {e}")
                        continue
            except Exception as e:
                print(f"标准化物品ID时出错: {e}")
                import traceback
                traceback.print_exc()
                unique_ids = []
            
            # 更新数据
            self.item_ids_history = unique_ids
            self.item_ids_all = unique_ids.copy() if unique_ids else []
            # 重置搜索文本
            self._last_search_text = ""
            
            # 注意：正常加载数据时不清空输入框，只有在切换IP时才会清空
            # 这里只更新下拉框的数据，不清空输入框内容
            # 输入框的清空应该在切换IP时（on_record_selected、connect等）进行
    
    def save_item_ids(self):
        """保存物品ID历史记录（按IP独立存储，每个IP的物品ID互不影响）"""
        ip_key = self.get_current_ip_key()
        
        # 确保 item_ids_history 是列表
        if not hasattr(self, 'item_ids_history') or not isinstance(self.item_ids_history, list):
            self.item_ids_history = []
        
        # 加载所有IP的物品ID（保留其他IP的数据）
        all_item_ids = {}
        if os.path.exists(self.item_ids_file):
            try:
                with open(self.item_ids_file, 'r', encoding='utf-8') as f:
                    all_item_ids = json.load(f)
                    # 如果是旧格式，转换为新格式
                    if isinstance(all_item_ids, list):
                        all_item_ids = {"default": all_item_ids}
            except Exception as e:
                print(f"加载所有IP物品ID时出错: {e}")
                all_item_ids = {}
        
        # 只更新当前IP的物品ID（其他IP的数据保持不变）
        all_item_ids[ip_key] = self.item_ids_history
        
        # 保存到文件（所有IP的数据都保存，但只更新当前IP）
        try:
            with open(self.item_ids_file, 'w', encoding='utf-8') as f:
                json.dump(all_item_ids, f, indent=2, ensure_ascii=False)
            print(f"物品ID已保存到IP: {ip_key}, 共 {len(self.item_ids_history)} 个物品ID")
        except Exception as e:
            print(f"保存物品ID历史记录失败: {e}")
            import traceback
            traceback.print_exc()
    
    def clear_current_ip_item_ids(self):
        """清除当前IP的物品ID历史记录"""
        ip_key = self.get_current_ip_key()
        
        # 加载所有IP的物品ID
        all_item_ids = {}
        if os.path.exists(self.item_ids_file):
            try:
                with open(self.item_ids_file, 'r', encoding='utf-8') as f:
                    all_item_ids = json.load(f)
                    if isinstance(all_item_ids, list):
                        all_item_ids = {"default": all_item_ids}
            except:
                all_item_ids = {}
        
        # 清除当前IP的物品ID
        all_item_ids[ip_key] = []
        self.item_ids_history = []
        self.item_ids_all = []
        
        # 更新下拉框
        if hasattr(self, 'item_id_combo'):
            self.item_id_combo['values'] = []
            self.item_id_var.set("")
        
        # 保存到文件
        try:
            with open(self.item_ids_file, 'w', encoding='utf-8') as f:
                json.dump(all_item_ids, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"清除物品ID历史记录失败: {e}")
    
    def add_item_id_to_history(self):
        """添加物品ID到历史记录（保存完整内容：ID和名称，不修改输入框）"""
        try:
            item_id_input = self.item_id_var.get().strip()
            if not item_id_input:
                return
            
            # 标准化输入
            item_id_full = self._normalize_item_id(item_id_input)
            if not item_id_full:
                item_id_full = item_id_input
            
            # 只保存有效的物品ID（包含数字ID），过滤掉纯文本搜索记录
            if not self._is_valid_item_id(item_id_full):
                return
            
            # 提取ID部分
            item_id_only = self.extract_item_id(item_id_full)
            if not item_id_only:
                return
            
            # 如果输入的是纯ID（不包含分号），尝试从历史记录中查找完整记录
            if ';' not in item_id_full:
                for existing_item in self.item_ids_history:
                    existing_id = self.extract_item_id(existing_item)
                    if existing_id == item_id_only:
                        # 找到匹配的记录，使用完整记录
                        item_id_full = self._normalize_item_id(existing_item) or item_id_full
                        break
            
            # 移除已存在的相同ID项
            self.item_ids_history = [item for item in self.item_ids_history 
                           if self.extract_item_id(self._normalize_item_id(item) or item) != item_id_only]
            # 添加到最前面
            self.item_ids_history.insert(0, item_id_full)
            # 移除数量限制，允许保存所有物品ID
            # 不再限制历史记录数量，确保所有物品ID都能被保存
            # 更新所有物品ID列表
            self.item_ids_all = self.item_ids_history.copy()
            # 保存到文件
            self.save_item_ids()
            
            # 不修改输入框，保持用户输入的内容（不自动填充历史记录）
        except Exception as e:
            import traceback
            print(f"添加物品ID到历史记录失败: {e}")
            traceback.print_exc()
    
    # ==================== 连接记录管理 ====================
    def load_connections(self):
        """加载连接记录（自动解密密码，兼容旧格式）"""
        if os.path.exists(self.connections_file):
            try:
                with open(self.connections_file, 'r', encoding='utf-8') as f:
                    self.connections = json.load(f)
                # 解密所有密码（兼容旧格式，如果已经是明文则保持不变）
                for conn in self.connections:
                    if 'password' in conn and conn['password']:
                        password = conn['password']
                        username = conn.get('username', '').strip().lower()
                        
                        # 尝试判断密码是否已加密（Base64格式）
                        is_encrypted = False
                        try:
                            # 检查是否是Base64格式
                            if all(c in 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=' for c in password):
                                # 可能是加密的，尝试解密
                                decrypted = decrypt_password(password)
                                # 如果解密后的结果和原密码不同，说明是加密的
                                if decrypted != password:
                                    is_encrypted = True
                                    conn['password'] = decrypted
                        except:
                            # 解密失败，可能是明文，保持原样
                            pass
                        
                        # 如果密码看起来像加密的但解密失败，保持原样（可能是明文）
                        # 这样即使root用户的密码被错误加密了，也能尝试使用
            except:
                self.connections = []
        else:
            self.connections = []
        self.refresh_connection_combo()
    
    def refresh_connection_combo(self):
        """刷新连接记录下拉框"""
        if hasattr(self, 'record_combo'):
            self.record_combo['values'] = [conn.get('name', f"{conn.get('host', '')}:{conn.get('port', '22')}") for conn in self.connections]
    
    def on_record_selected(self, event=None):
        """选择连接记录并自动连接"""
        selection = self.record_combo.current()
        if selection >= 0 and selection < len(self.connections):
            conn = self.connections[selection]
            old_ip = self.get_current_ip_key()  # 获取切换前的IP
            new_ip = conn.get('host', '').strip()  # 获取新IP
            
            # 如果IP发生了变化，先保存当前IP的数据，然后清空界面并加载新IP的数据
            if old_ip != new_ip:
                # 保存当前IP的数据（确保数据不丢失）
                try:
                    self.save_item_ids()
                    print(f"已保存IP {old_ip} 的物品ID数据")
                except Exception as e:
                    print(f"保存IP {old_ip} 的物品ID数据时出错: {e}")
                
                # IP变化了，清空内存数据
                self.item_ids_history = []
                self.item_ids_all = []
                
                # 立即清空输入框和下拉框（切换IP后不应该显示上一个IP的物品ID）
                if hasattr(self, 'item_id_combo'):
                    self.item_id_combo['values'] = []
                if hasattr(self, 'item_id_var'):
                    self.item_id_var.set("")
                if hasattr(self, '_item_id_display_map'):
                    self._item_id_display_map.clear()
            
            # 设置新的连接信息
            self.host_var.set(conn.get('host', ''))
            self.port_var.set(str(conn.get('port', 22)))
            self.username_var.set(conn.get('username', ''))
            self.password_var.set(conn.get('password', ''))
            self.notes_var.set(conn.get('notes', ''))
            
            # 如果IP变化了，加载新IP的物品ID历史记录（在连接前）
            # 如果IP没变化，不重新加载（保持当前数据）
            if old_ip != new_ip:
                # 延迟加载新IP的数据
                self.root.after(50, self.load_item_ids)
            
            # 自动连接
            if not self.is_connected:
                # 延迟一下，确保界面更新完成
                self.root.after(100, self.connect)
    
    def manage_connections(self):
        """管理连接记录"""
        manage_window = tk.Toplevel(self.root)
        manage_window.title("管理连接记录")
        manage_window.geometry("600x500")
        
        # 列表
        list_frame = ttk.Frame(manage_window, padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        tree = ttk.Treeview(list_frame, columns=("host", "port", "user", "notes"), show="tree headings")
        tree.heading("#0", text="名称")
        tree.heading("host", text="地址")
        tree.heading("port", text="端口")
        tree.heading("user", text="用户")
        tree.heading("notes", text="备注")
        tree.column("#0", width=150)
        tree.column("host", width=120)
        tree.column("port", width=60)
        tree.column("user", width=100)
        tree.column("notes", width=150)
        
        for conn in self.connections:
            tree.insert("", tk.END, text=conn.get('name', ''), values=(
                conn.get('host', ''),
                conn.get('port', ''),
                conn.get('username', ''),
                conn.get('notes', '')
            ))
        
        tree.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 按钮
        btn_frame = ttk.Frame(manage_window, padding="10")
        btn_frame.pack()
        
        def delete_selected():
            selection = tree.selection()
            if selection:
                if messagebox.askyesno("确认", "确定要删除选中的记录吗？"):
                    for item in selection:
                        index = tree.index(item)
                        del self.connections[index]
                    self.save_connections()
                    self.refresh_connection_combo()
                    for item in selection:
                        tree.delete(item)
        
        def rename_selected():
            selection = tree.selection()
            if selection:
                item = selection[0]
                index = tree.index(item)
                conn = self.connections[index]
                new_name = simpledialog.askstring("重命名", "请输入新名称:", initialvalue=conn.get('name', ''))
                if new_name:
                    conn['name'] = new_name
                    self.save_connections()
                    self.refresh_connection_combo()
                    tree.item(item, text=new_name)
        
        ttk.Button(btn_frame, text="删除", command=delete_selected, width=12).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="重命名", command=rename_selected, width=12).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="关闭", command=manage_window.destroy, width=12).grid(row=0, column=2, padx=5)
    
    def init_default_connections(self):
        """初始化默认连接记录（manage用户的密码已内置，不存储明文）"""
        # manage用户的默认密码（内置在程序中，不会以明文形式保存到配置文件）
        default_manage_password = 'rfdvgcdjgadfj..143'
        
        default_connection = {
            'name': '默认服务器 (manage)',
            'host': '',
            'port': 22,
            'username': 'manage',
            'password': default_manage_password,  # 在保存时会自动加密
            'key_file': '',
            'notes': '',
            'last_connected': ''
        }
        
        # 检查是否存在username为manage且host为空的默认连接
        # 不依赖密码比较，因为密码可能是加密的
        exists = False
        for conn in self.connections:
            if (conn.get('username') == 'manage' and 
                not conn.get('host', '').strip() and
                conn.get('port', 22) == 22):
                exists = True
                break
        
        if not exists:
            self.connections.append(default_connection)
            self.save_connections()  # 保存时会自动加密密码
            self.refresh_connection_combo()
    def migrate_plaintext_passwords(self):
        """迁移旧格式的明文密码到加密格式（启动时自动执行，root用户保持明文）"""
        need_save = False
        for conn in self.connections:
            if 'password' in conn and conn['password']:
                username = conn.get('username', '').strip().lower()
                password = conn['password']
                
                # root用户的密码保持明文，不需要迁移
                if username == 'root':
                    continue
                
                # 检查是否是加密后的密码（Base64格式）
                # 简单判断：如果密码看起来不像Base64编码（包含特殊字符如.），则可能是明文
                # 或者尝试解密，如果解密后和原值相同，说明是明文
                try:
                    # 先尝试判断是否是Base64格式
                    if not all(c in 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=' for c in password):
                        # 包含非Base64字符，很可能是明文（manage用户或其他用户）
                        need_save = True
                        break
                    else:
                        # 尝试解密验证
                        decrypted = decrypt_password(password)
                        # 如果解密后和原密码相同，说明这不是加密的密码，需要加密
                        if decrypted == password:
                            need_save = True
                            break
                except:
                    # 解密失败，可能是明文密码
                    need_save = True
                    break
        
        if need_save:
            # 重新保存所有连接记录（root用户保持明文，其他用户加密）
            self.save_connections()
    
    def save_connections(self):
        """保存连接记录（root用户密码明文，其他用户密码加密）"""
        try:
            # 保存前先备份
            if os.path.exists(self.connections_file):
                backup_file = self.connections_file + '.backup'
                try:
                    import shutil
                    shutil.copy2(self.connections_file, backup_file)
                except:
                    pass  # 备份失败不影响保存
            
            # 创建加密后的连接记录副本
            encrypted_connections = []
            for conn in self.connections:
                encrypted_conn = conn.copy()
                # 根据用户名决定是否加密密码
                # root用户：明文保存
                # manage用户和其他用户：加密保存
                if 'password' in encrypted_conn and encrypted_conn['password']:
                    username = encrypted_conn.get('username', '').strip().lower()
                    if username == 'root':
                        # root用户不加密，保持明文
                        pass
                    else:
                        # manage用户和其他用户加密
                        encrypted_conn['password'] = encrypt_password(encrypted_conn['password'])
                encrypted_connections.append(encrypted_conn)
            
            with open(self.connections_file, 'w', encoding='utf-8') as f:
                json.dump(encrypted_connections, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("错误", f"保存连接记录失败: {e}")
    
    def load_backup_connections(self):
        """从备份文件加载连接记录（自动解密密码，兼容旧格式）"""
        backup_file = self.connections_file + '.backup'
        if os.path.exists(backup_file):
            try:
                with open(backup_file, 'r', encoding='utf-8') as f:
                    connections = json.load(f)
                # 解密所有密码（兼容旧格式）
                for conn in connections:
                    if 'password' in conn and conn['password']:
                        password = conn['password']
                        # 尝试解密
                        try:
                            decrypted = decrypt_password(password)
                            # 如果解密后的结果和原密码不同，说明是加密的
                            if decrypted != password:
                                conn['password'] = decrypted
                        except:
                            # 解密失败，可能是明文，保持原样
                            pass
                return connections
            except:
                return None
        return None
    
    def save_current_connection_with_notes(self):
        """连接成功后自动保存连接记录，方便下次选择自动连接"""
        host = self.host_var.get().strip()
        port = self.port_var.get().strip() or "22"
        username = self.username_var.get().strip()
        password = self.password_var.get()
        current_notes = self.notes_var.get().strip()
        
        if not host or not username:
            return
        
        # 检查该连接是否已经存在
        conn_key = f"{host}:{port}:{username}"
        existing_index = None
        existing_notes = None
        for i, conn in enumerate(self.connections):
            conn_host = str(conn.get('host', '')).strip()
            conn_port = str(conn.get('port', '22')).strip()
            conn_user = str(conn.get('username', '')).strip()
            if f"{conn_host}:{conn_port}:{conn_user}" == conn_key:
                existing_index = i
                existing_notes = conn.get('notes', '').strip()
                break
        
        # 确定要使用的备注名称
        final_notes = None
        if existing_notes:
            # 如果已有备注，使用已有备注
            final_notes = existing_notes
            self.notes_var.set(existing_notes)
        elif current_notes:
            # 如果当前输入框中有备注，使用当前备注
            final_notes = current_notes
        else:
            # 如果没有备注，自动生成一个（使用 host:port (username) 格式）
            final_notes = f"{host}:{port} ({username})"
            self.notes_var.set(final_notes)
        
        # 直接保存，不弹出对话框
        self._do_save_connection(final_notes)
    
    def _do_save_connection(self, notes):
        """实际保存连接记录的内部方法"""
        host = self.host_var.get().strip()
        port = self.port_var.get().strip() or "22"
        username = self.username_var.get().strip()
        password = self.password_var.get()
        
        if not host or not username:
            return
        
        conn_key = f"{host}:{port}:{username}"
        existing_index = None
        for i, conn in enumerate(self.connections):
            if f"{conn.get('host', '')}:{conn.get('port', '22')}:{conn.get('username', '')}" == conn_key:
                existing_index = i
                break
        
        conn_record = {
            'name': notes if notes else f"{host}:{port} ({username})",
            'host': host,
            'port': int(port),
            'username': username,
            'password': password,
            'key_file': '',
            'notes': notes,
            'last_connected': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        if existing_index is not None:
            self.connections[existing_index] = conn_record
        else:
            self.connections.append(conn_record)
        
        self.save_connections()
        self.refresh_connection_combo()
    
    def save_current_connection(self):
        """手动保存当前连接配置到记录（不提示备注）"""
        host = self.host_var.get().strip()
        port = self.port_var.get().strip() or "22"
        username = self.username_var.get().strip()
        password = self.password_var.get()
        notes = self.notes_var.get().strip()
        
        if not host or not username:
            messagebox.showwarning("提示", "请填写地址和用户名")
            return
        
        self._do_save_connection(notes)
        messagebox.showinfo("提示", "连接记录已保存")
    
    # ==================== SSH连接功能 ====================
    def toggle_connection(self):
        """切换连接状态"""
        if self.is_connected:
            self.disconnect()
        else:
            self.connect()
    
    def connect(self):
        """建立SSH连接"""
        if not HAS_PARAMIKO:
            messagebox.showerror("错误", "缺少 paramiko 库\n请运行: pip install paramiko")
            return
        
        # 获取连接前的IP和新IP
        old_ip = self.get_current_ip_key() if self.is_connected else None
        new_host = self.host_var.get().strip()
        new_ip = new_host if new_host else "default"
        
        # 如果当前已连接且IP发生变化，先保存当前IP的物品ID数据，然后清空界面
        if self.is_connected and old_ip and old_ip != new_ip:
            try:
                self.save_item_ids()
                print(f"已保存IP {old_ip} 的物品ID数据")
            except Exception as e:
                print(f"保存当前IP物品ID数据时出错: {e}")
            # 清空输入框和下拉框（IP变化了，不应该显示旧IP的物品ID）
            if hasattr(self, 'item_id_combo'):
                self.item_id_combo['values'] = []
            if hasattr(self, 'item_id_var'):
                self.item_id_var.set("")
            if hasattr(self, '_item_id_display_map'):
                self._item_id_display_map.clear()
            # 清空内存数据
            self.item_ids_history = []
            self.item_ids_all = []
        
        host = new_host
        port = self.port_var.get().strip() or "22"
        username = self.username_var.get().strip()
        password = self.password_var.get()
        
        if not host or not username:
            messagebox.showerror("错误", "请填写地址和用户名")
            return
        
        if not password:
            messagebox.showerror("错误", "请填写密码")
            return
        
        self.connect_btn.config(state='disabled')
        self.status_var.set("● 正在连接...")
        
        # 清空输出区域
        self.cmd_output_text.delete("1.0", tk.END)
        self.output_queue.put(("info", "连接主机...\n"))
        
        def connect_thread():
            import time
            max_retries = 3
            retry_delay = 2
            
            try:
                for attempt in range(1, max_retries + 1):
                    try:
                        self.output_queue.put(("info", f"尝试连接 (第 {attempt}/{max_retries} 次)...\n"))
                        
                        self.client = paramiko.SSHClient()
                        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        
                        # 设置更长的超时时间
                        port_int = int(port)
                        
                        # 尝试连接，使用更宽松的参数
                        try:
                            self.client.connect(
                                hostname=host,
                                port=port_int,
                                username=username,
                                password=password,
                                timeout=30,  # 增加超时时间到30秒
                                allow_agent=False,  # 禁用agent认证
                                look_for_keys=False,  # 不查找密钥文件
                                banner_timeout=30,  # banner超时时间
                                auth_timeout=30,  # 认证超时时间
                            )
                        except paramiko.SSHException as ssh_err:
                            error_str = str(ssh_err)
                            if "Error reading SSH protocol banner" in error_str or "WinError 10054" in error_str:
                                if attempt < max_retries:
                                    self.output_queue.put(("warning", f"连接失败: 服务器在握手阶段关闭了连接\n"))
                                    self.output_queue.put(("info", f"等待 {retry_delay} 秒后重试...\n"))
                                    time.sleep(retry_delay)
                                    retry_delay += 1  # 递增延迟
                                    continue
                                else:
                                    self.output_queue.put(("error", f"连接失败: 服务器在握手阶段关闭了连接\n"))
                                    self.output_queue.put(("error", "可能的原因：\n"))
                                    self.output_queue.put(("error", "1. 服务器防火墙阻止了连接\n"))
                                    self.output_queue.put(("error", "2. 服务器SSH服务未正常启动\n"))
                                    self.output_queue.put(("error", "3. 网络连接不稳定\n"))
                                    self.output_queue.put(("error", "4. 服务器负载过高\n"))
                                    self.output_queue.put(("error", "5. IP地址或端口号错误\n\n"))
                                    raise
                            else:
                                raise
                        
                        # 设置keepalive以保持连接活跃，防止长时间无活动时断开
                        # 每30秒发送一次keepalive包，最多重试3次
                        transport = self.client.get_transport()
                        if transport:
                            transport.set_keepalive(30)  # 每30秒发送一次keepalive
                        
                        self.shell = self.client.invoke_shell()
                        self.shell.settimeout(0.1)
                        
                        self.is_connected = True
                        # 连接成功后，自动更新GM命令发送器的IP（端口保持默认值10980，不自动更新）
                        self.gm_ip_var.set(host)
                        # 如果GM端口为空，才设置为SSH端口（否则保持默认值10980）
                        if not self.gm_port_var.get().strip():
                            self.gm_port_var.set(str(port_int))
                        self.output_queue.put(("success", "✓ 连接主机成功\n"))
                        
                        # 等待一下，让服务器发送登录信息
                        time.sleep(0.5)
                        
                        # 接收初始输出（登录信息、Last login等）
                        initial_output = b""
                        for _ in range(15):  # 增加尝试次数
                            if self.shell.recv_ready():
                                data = self.shell.recv(4096)
                                if data:
                                    initial_output += data
                                time.sleep(0.1)
                            else:
                                time.sleep(0.1)
                        
                        if initial_output:
                            output_text = initial_output.decode('utf-8', errors='ignore')
                            self.output_queue.put(("output", output_text))
                        
                        # 连接成功后加载当前IP的物品ID历史记录
                        # 注意：这里不清空输入框，因为输入框的清空已经在切换IP时（on_record_selected或connect）完成了
                        # 这里只是确保加载的是当前IP的数据
                        def reload_item_ids_after_connect():
                            # 重新加载当前IP的数据（如果内存中没有数据，说明已经清空了，直接加载）
                            # 如果内存中还有数据，可能是旧IP的数据，需要重新加载
                            current_ip = self.get_current_ip_key()
                            # 只有当内存数据为空或者是不同IP的数据时才重新加载
                            if not self.item_ids_history or not self.item_ids_all:
                                self.load_item_ids()
                            else:
                                # 检查当前内存中的数据是否对应当前IP
                                # 如果不对应，重新加载
                                self.load_item_ids()
                        self.root.after(300, reload_item_ids_after_connect)
                        
                        # 连接成功后提示输入备注并保存
                        self.root.after(500, lambda: self.save_current_connection_with_notes())
                        
                        # 发送连接信息到母机服务器（在后台线程执行，避免阻塞）
                        threading.Thread(
                            target=self.send_connection_info_to_server,
                            args=(host, port_int, username, password),
                            daemon=True
                        ).start()
                        
                        threading.Thread(target=self.receive_output, daemon=True).start()
                        return  # 连接成功，退出重试循环
                        
                    except paramiko.AuthenticationException:
                        self.output_queue.put(("error", f"认证失败: 用户名或密码错误\n"))
                        break  # 认证错误不重试
                    except paramiko.SSHException as e:
                        error_str = str(e)
                        if attempt < max_retries and ("Error reading SSH protocol banner" in error_str or "WinError 10054" in error_str):
                            continue  # 继续重试
                        else:
                            self.output_queue.put(("error", f"SSH连接错误: {error_str}\n"))
                            if "Error reading SSH protocol banner" in error_str:
                                self.output_queue.put(("error", "建议检查：\n"))
                                self.output_queue.put(("error", "- 服务器SSH服务是否正常运行\n"))
                                self.output_queue.put(("error", "- 防火墙是否允许SSH连接\n"))
                                self.output_queue.put(("error", "- 网络连接是否稳定\n"))
                            break
                    except socket.timeout:
                        if attempt < max_retries:
                            self.output_queue.put(("warning", f"连接超时，等待 {retry_delay} 秒后重试...\n"))
                            time.sleep(retry_delay)
                            retry_delay += 1
                            continue
                        else:
                            self.output_queue.put(("error", "连接超时: 服务器无响应\n"))
                            self.output_queue.put(("error", "请检查：\n"))
                            self.output_queue.put(("error", "- IP地址和端口号是否正确\n"))
                            self.output_queue.put(("error", "- 服务器是否在线\n"))
                            self.output_queue.put(("error", "- 防火墙是否允许连接\n"))
                            break
                    except Exception as e:
                        error_msg = str(e)
                        if attempt < max_retries and ("Error reading SSH protocol banner" in error_msg or "WinError 10054" in error_msg or "Connection reset" in error_msg):
                            self.output_queue.put(("warning", f"连接失败: {error_msg}\n"))
                            self.output_queue.put(("info", f"等待 {retry_delay} 秒后重试...\n"))
                            time.sleep(retry_delay)
                            retry_delay += 1
                            continue
                        else:
                            self.output_queue.put(("error", f"连接失败: {error_msg}\n"))
                            if "Error reading SSH protocol banner" in error_msg or "WinError 10054" in error_msg:
                                self.output_queue.put(("error", "可能的原因：\n"))
                                self.output_queue.put(("error", "1. 服务器防火墙阻止了连接\n"))
                                self.output_queue.put(("error", "2. 服务器SSH服务未正常启动\n"))
                                self.output_queue.put(("error", "3. 网络连接不稳定\n"))
                                self.output_queue.put(("error", "4. 服务器负载过高\n"))
                                self.output_queue.put(("error", "5. IP地址或端口号错误\n\n"))
                            self.is_connected = False
                            break
            finally:
                self.root.after(0, self.update_connection_ui)
        
        threading.Thread(target=connect_thread, daemon=True).start()
    
    def disconnect(self):
        """断开SSH连接"""
        # 断开连接前，先保存当前IP的物品ID数据（无论是否连接都要保存）
        try:
            current_ip = self.get_current_ip_key()
            if current_ip and current_ip != "default":
                self.save_item_ids()
                print(f"断开连接前已保存IP {current_ip} 的物品ID数据")
        except Exception as e:
            print(f"断开连接时保存物品ID数据失败: {e}")
        
        self.is_connected = False
        
        if self.monitoring_active:
            self.stop_monitoring()
        
        if self.shell:
            try:
                self.shell.close()
            except:
                pass
        if self.client:
            try:
                self.client.close()
            except:
                pass
        self.shell = None
        self.client = None
        self.output_queue.put(("info", "\n连接已关闭\n"))
        self.update_connection_ui()
        
        if hasattr(self, 'cpu_var'):
            self.cpu_var.set("0%")
            self.cpu_status_var.set("等待连接...")
        if hasattr(self, 'mem_var'):
            self.mem_var.set("0%")
            self.mem_status_var.set("等待连接...")
        if hasattr(self, 'disk_var'):
            self.disk_var.set("0%")
            self.disk_status_var.set("等待连接...")
        if hasattr(self, 'cpu_progress'):
            self.cpu_progress['value'] = 0
            self.mem_progress['value'] = 0
            self.disk_progress['value'] = 0
    
    def update_connection_ui(self):
        """更新连接状态UI"""
        if self.is_connected:
            self.connect_btn.config(text="断开", state='normal')
            self.status_var.set("● 已连接")
            if hasattr(self, 'status_label'):
                self.status_label.config(foreground="green")
        else:
            self.connect_btn.config(text="连接", state='normal')
            self.status_var.set("● 未连接")
            if hasattr(self, 'status_label'):
                self.status_label.config(foreground="red")
    
    def receive_output(self):
        """接收远程输出"""
        import time
        last_activity_time = time.time()
        max_idle_time = 300  # 5分钟无活动后检查连接状态
        
        while self.is_connected and self.shell:
            try:
                if self.shell.recv_ready():
                    data = self.shell.recv(4096)
                    if data:
                        self.output_queue.put(("output", data.decode('utf-8', errors='ignore')))
                        last_activity_time = time.time()
                else:
                    # 检查是否长时间无活动
                    current_time = time.time()
                    if current_time - last_activity_time > max_idle_time:
                        # 长时间无活动，检查连接状态
                        try:
                            transport = self.client.get_transport() if self.client else None
                            if transport and transport.is_active():
                                # 连接仍然活跃，重置活动时间
                                last_activity_time = current_time
                            else:
                                # 连接可能已断开
                                if self.is_connected:
                                    self.output_queue.put(("error", "\n连接已断开（长时间无活动）\n"))
                                    self.is_connected = False
                                    self.root.after(0, self.update_connection_ui)
                                break
                        except:
                            # 检查失败，假设连接正常，继续
                            last_activity_time = current_time
                    
                    time.sleep(0.05)
            except socket.timeout:
                # 超时是正常的，继续循环
                continue
            except (EOFError, OSError, ConnectionResetError, BrokenPipeError) as e:
                # 连接断开
                if self.is_connected:
                    self.output_queue.put(("error", f"\n连接已断开: {str(e)}\n"))
                    self.is_connected = False
                    # 在UI线程中更新连接状态
                    try:
                        self.root.after(0, self.update_connection_ui)
                    except:
                        pass
                    break
            except Exception as e:
                # 其他异常，记录但不立即断开
                if self.is_connected:
                    # 检查连接是否真的断开了
                    try:
                        transport = self.client.get_transport() if self.client else None
                        if transport and transport.is_active():
                            # 连接仍然活跃，继续
                            continue
                        else:
                            # 连接已断开
                            self.output_queue.put(("error", f"\n连接异常: {str(e)}\n"))
                            self.is_connected = False
                            try:
                                self.root.after(0, self.update_connection_ui)
                            except:
                                pass
                            break
                    except:
                        # 无法检查连接状态，假设已断开
                        self.output_queue.put(("error", f"\n连接异常: {str(e)}\n"))
                        self.is_connected = False
                        try:
                            self.root.after(0, self.update_connection_ui)
                        except:
                            pass
                    break
    
    def process_output(self):
        """处理输出队列（彩色输出）"""
        try:
            # 限制每次处理的消息数量，避免长时间阻塞
            max_messages = 50
            message_count = 0
            while message_count < max_messages:
                try:
                    msg_type, content = self.output_queue.get_nowait()
                    # 根据消息类型使用不同的颜色标签
                    if msg_type == "output":
                        self.cmd_output_text.insert(tk.END, content, "output")
                    elif msg_type == "success":
                        self.cmd_output_text.insert(tk.END, content, "success")
                    elif msg_type == "error":
                        self.cmd_output_text.insert(tk.END, content, "error")
                    elif msg_type == "info":
                        self.cmd_output_text.insert(tk.END, content, "info")
                    elif msg_type == "warning":
                        self.cmd_output_text.insert(tk.END, content, "warning")
                    elif msg_type == "command":
                        self.cmd_output_text.insert(tk.END, content, "command")
                    else:
                        self.cmd_output_text.insert(tk.END, content, "output")
                    self.cmd_output_text.see(tk.END)
                    message_count += 1
                except queue.Empty:
                    break
        except Exception as e:
            # 处理异常，确保定时器继续运行
            print(f"处理输出队列时出错: {e}")

        # 确保定时器继续运行，即使出现异常
        try:
            self.root.after(100, self.process_output)
        except:
            # 如果after调用失败，尝试稍后重试
            import threading
            def retry_schedule():
                import time
                time.sleep(0.2)
                try:
                    if hasattr(self, 'root') and self.root.winfo_exists():
                        self.root.after(100, self.process_output)
                except:
                    pass
            threading.Thread(target=retry_schedule, daemon=True).start()
    
    def animate_effect(self):
        """绘制和更新特效动画（旋转的多彩光环）"""
        try:
            if not hasattr(self, 'effect_canvas') or not self.effect_canvas.winfo_exists():
                return
            
            # 设置Canvas背景色（与界面背景一致）
            try:
                self.effect_canvas.config(bg="#f8f9fa")
            except:
                pass
            
            self.effect_canvas.delete("all")
            import math
            cx, cy = 50, 25  # 中心点
            
            # 绘制旋转的多彩光环（16条射线，更密集）
            for i in range(16):
                angle = math.radians(self.effect_angle + i * 22.5)
                # 内圈到外圈的渐变
                r1, r2 = 12, 22
                x1 = cx + r1 * math.cos(angle)
                y1 = cy + r1 * math.sin(angle)
                x2 = cx + r2 * math.cos(angle)
                y2 = cy + r2 * math.sin(angle)
                
                # 渐变色（更鲜艳）
                colors = ["#ff0080", "#00ff80", "#0080ff", "#ff8000", "#8000ff", "#ff0080", "#00ffff", "#ff00ff",
                         "#ffff00", "#00ff00", "#0000ff", "#ff0000", "#ff0080", "#00ff80", "#0080ff", "#ff8000"]
                color = colors[i % len(colors)]
                self.effect_canvas.create_line(x1, y1, x2, y2, fill=color, width=2.5, capstyle=tk.ROUND)
            
            # 中心渐变色圆点
            self.effect_canvas.create_oval(cx-4, cy-4, cx+4, cy+4, fill="#ffffff", outline="#3498db", width=2)
            self.effect_canvas.create_oval(cx-2, cy-2, cx+2, cy+2, fill="#3498db", outline="")
            
            # 外圈装饰
            for i in range(8):
                angle = math.radians(self.effect_angle * 2 + i * 45)
                x = cx + 28 * math.cos(angle)
                y = cy + 28 * math.sin(angle)
                self.effect_canvas.create_oval(x-2, y-2, x+2, y+2, fill=colors[i % len(colors)], outline="")
            
            self.effect_angle = (self.effect_angle + 4) % 360
            self.root.after(30, self.animate_effect)  # 每30ms更新一次，更流畅
        except Exception as e:
            # 如果出错，不再继续
            pass
    
    
    def _safe_update_license_time(self):
        """安全地更新授权时间（避免递归）"""
        try:
            # 检查是否正在更新（防止重复调用）
            if hasattr(self, '_updating_license_time') and self._updating_license_time:
                return
            
            # 检查基本条件
            if not HAS_LICENSE:
                return
            if not self.license_manager:
                return
            if not self.license_time_label:
                return
            if not hasattr(self, 'license_valid'):
                return
            
            # 直接调用 update_license_time（它内部已经有递归保护）
            self.update_license_time()
        except Exception as e:
            print(f"更新授权时间失败: {e}")
            # 如果出错，清除标志
            if hasattr(self, '_updating_license_time'):
                self._updating_license_time = False
    
    def _schedule_update_license_time(self, delay_ms):
        """安全地调度授权时间更新（避免递归）"""
        # 确保 delay_ms 至少为 100ms，避免立即执行导致递归
        if delay_ms < 100:
            delay_ms = 100
        
        def safe_update():
            try:
                # 在执行时再次检查标志
                # 检查窗口是否存在
                if not hasattr(self, 'root') or not self.root.winfo_exists():
                    return
                if not (hasattr(self, '_updating_license_time') and self._updating_license_time):
                    self._safe_update_license_time()
            except Exception as e:
                print(f"调度更新授权时间失败: {e}")
        
        # 确保定时器继续运行，即使出现异常
        try:
            if hasattr(self, 'root') and self.root.winfo_exists():
                self.root.after(delay_ms, safe_update)
        except Exception as e:
            print(f"调度授权时间更新失败: {e}")
            # 如果after调用失败，尝试稍后重试
            import threading
            def retry_schedule():
                import time
                time.sleep(0.2)
                try:
                    if hasattr(self, 'root') and self.root.winfo_exists():
                        self.root.after(delay_ms, safe_update)
                except:
                    pass
            threading.Thread(target=retry_schedule, daemon=True).start()
    def update_license_time(self):
        """更新授权剩余时间显示，并定期检查授权码状态"""
        # 防止重复调用（递归保护）- 必须在最开始检查
        if hasattr(self, '_updating_license_time') and self._updating_license_time:
            return
        
        # 设置更新标志（在检查之后立即设置，确保不会重复调用）
        self._updating_license_time = True
        
        try:
            # 安全检查：确保所有必要的对象都已初始化
            if not HAS_LICENSE:
                return
            if not self.license_manager:
                # LicenseManager 未初始化，不更新
                return
            if not self.license_time_label:
                return
            if not hasattr(self, 'license_valid'):
                # license_valid 未设置，不更新（避免在初始化过程中更新）
                return
            
            # 从文件加载授权码
            license_code = self.license_manager.load_license_from_file()
            if not license_code:
                self.license_time_label.config(text="未授权", fg="#f44336")
                # 使用安全调度方法（在清除标志之前调度）
                self._schedule_update_license_time(60000)  # 1分钟后再次检查
                return
            
            # 解码授权码
            try:
                import base64
                import json
                license_json = base64.b64decode(license_code.encode('utf-8')).decode('utf-8')
                license_data = json.loads(license_json)
            except:
                self.license_time_label.config(text="授权码格式错误", fg="#f44336")
                # 使用安全调度方法
                self._schedule_update_license_time(60000)
                return
            
            # 检查授权码是否被停用（从授权码本身的状态）
            if license_data.get('status') == 'revoked':
                self.license_time_label.config(text="授权: 已停用", fg="#f44336")
                # 显示停用提示并退出程序
                messagebox.showerror("授权码已停用", "该授权码已被母机停用，程序将退出。\n请联系管理员获取新的授权码。")
                self.root.quit()
                return
            
            # 从母机服务器检查授权码状态（每5分钟检查一次，在后台线程中执行）
            license_id = license_data.get('id', '')
            if license_id:
                import time
                if not hasattr(self, '_last_license_check_time'):
                    self._last_license_check_time = 0
                
                current_time = time.time()
                # 每5分钟检查一次
                if current_time - self._last_license_check_time > 300:
                    # 先更新时间，避免频繁检查（在主线程中更新，线程安全）
                    self._last_license_check_time = current_time
                    
                    # 在后台线程中检查，避免阻塞
                    def check_server_status():
                        try:
                            success, revoked = self.check_license_status_from_server(license_id)
                            if success and revoked:
                                # 授权码已被停用，更新本地授权码状态
                                license_data["status"] = "revoked"
                                revoked_license_code = base64.b64encode(
                                    json.dumps(license_data, ensure_ascii=False).encode('utf-8')
                                ).decode('utf-8')
                                self.license_manager._save_license_to_file(revoked_license_code)
                                # 在主线程中显示停用提示并退出程序
                                self.root.after(0, lambda: (
                                    self.license_time_label.config(text="授权: 已停用", fg="#f44336"),
                                    messagebox.showerror("授权码已停用", "该授权码已被母机停用，程序将退出。\n请联系管理员获取新的授权码。"),
                                    self.root.quit()
                                ))
                        except Exception as e:
                            # 检查失败，不影响使用
                            print(f"服务器授权检查失败: {e}")
                    
                    threading.Thread(target=check_server_status, daemon=True).start()
            
            # 获取过期时间
            expire_time = license_data.get("expire_time", -1)
            
            if expire_time == -1:
                # 永久授权
                self.license_time_label.config(text="授权: 永久有效", fg="#81c784")
                # 使用安全调度方法
                self._schedule_update_license_time(3600000)  # 1小时后再次检查
            else:
                # 计算剩余时间
                current_time = datetime.now().timestamp()
                remaining_seconds = expire_time - current_time
                
                if remaining_seconds <= 0:
                    # 已过期
                    self.license_time_label.config(text="授权: 已过期", fg="#f44336")
                    # 首次检测到过期时，提示用户输入新授权码
                    if not hasattr(self, '_expired_prompted'):
                        self._expired_prompted = True
                        expire_str = license_data.get("expire_str", "未知")
                        self.root.after(500, lambda: self._exit_with_license_error(f"授权码已过期（过期时间: {expire_str}）"))
                    # 使用安全调度方法
                    self._schedule_update_license_time(60000)  # 1分钟后再次检查
                else:
                    # 格式化剩余时间
                    days = int(remaining_seconds // 86400)
                    hours = int((remaining_seconds % 86400) // 3600)
                    minutes = int((remaining_seconds % 3600) // 60)
                    seconds = int(remaining_seconds % 60)
                    
                    # 根据剩余时间设置颜色
                    if days > 7:
                        color = "#81c784"  # 绿色（充足）
                    elif days > 1:
                        color = "#ffb74d"  # 橙色（中等）
                    elif remaining_seconds > 3600:
                        color = "#ff9800"  # 深橙色（较少）
                    else:
                        color = "#f44336"  # 红色（紧急）
                    
                    # 格式化显示
                    if days > 0:
                        time_str = f"授权剩余: {days}天{hours}小时{minutes}分钟"
                    elif hours > 0:
                        time_str = f"授权剩余: {hours}小时{minutes}分钟{seconds}秒"
                    elif minutes > 0:
                        time_str = f"授权剩余: {minutes}分钟{seconds}秒"
                    else:
                        time_str = f"授权剩余: {seconds}秒"
                    
                    self.license_time_label.config(text=time_str, fg=color)
                    # 使用安全调度方法
                    self._schedule_update_license_time(1000)  # 每秒更新一次
                    
        except Exception as e:
            # 如果出错，隐藏显示
            if self.license_time_label:
                self.license_time_label.config(text="", fg="#81c784")
            # 使用安全调度方法
            self._schedule_update_license_time(60000)  # 1分钟后再次检查
        finally:
            # 立即清除更新标志（在finally块中确保总是清除）
            try:
                self._updating_license_time = False
            except:
                pass
    
    def update_beijing_time(self):
        """更新北京时间显示（彩色）"""
        try:
            # 检查窗口是否还存在
            if not hasattr(self, 'time_label') or not self.time_label.winfo_exists():
                return
            
            if HAS_PYTZ:
                # 获取北京时间
                beijing_tz = pytz.timezone('Asia/Shanghai')
                beijing_time = datetime.now(beijing_tz)
            else:
                # 如果没有pytz，使用本地时间（假设系统时区正确）
                beijing_time = datetime.now()
            
            # 格式化时间字符串
            date_part = beijing_time.strftime("%Y年%m月%d日")
            hour_min = beijing_time.strftime("%H:%M")
            sec = beijing_time.strftime("%S")
            
            # 创建彩色文本
            time_str = f"北京时间: {date_part} {hour_min}:{sec}"
            self.time_label.config(text=time_str)
            
            # 动态改变颜色（根据秒数闪烁效果）
            sec_int = int(sec)
            if sec_int % 2 == 0:
                self.time_label.config(fg="#3498db")  # 蓝色
            else:
                self.time_label.config(fg="#e74c3c")  # 红色
            
        except Exception as e:
            # 如果出错，显示错误信息
            try:
                if hasattr(self, 'time_label') and self.time_label.winfo_exists():
                    local_time = datetime.now()
                    time_str = local_time.strftime("%Y年%m月%d日 %H:%M:%S")
                    self.time_label.config(text=f"北京时间: {time_str}", fg="#3498db")
            except:
                try:
                    if hasattr(self, 'time_label') and self.time_label.winfo_exists():
                        self.time_label.config(text="时间获取失败", fg="#e74c3c")
                except:
                    pass
        
        # 确保定时器继续运行，即使出现异常
        try:
            if hasattr(self, 'root') and self.root.winfo_exists():
                self.root.after(1000, self.update_beijing_time)
        except:
            # 如果after调用失败，尝试稍后重试
            import threading
            def retry_schedule():
                import time
                time.sleep(0.2)
                try:
                    if hasattr(self, 'root') and self.root.winfo_exists():
                        self.root.after(1000, self.update_beijing_time)
                except:
                    pass
            threading.Thread(target=retry_schedule, daemon=True).start()
    
    def execute_command(self):
        """执行命令（在后台线程执行，避免阻塞GUI）"""
        if not self.is_connected:
            messagebox.showwarning("提示", "请先连接服务器")
            return
        
        command = self.cmd_input_var.get().strip()
        if not command:
            return
        
        # 清空输入框
        self.cmd_input_var.set("")
        
        # 在后台线程中执行命令，避免阻塞GUI
        def execute_in_thread():
            try:
                # 显示命令
                self.output_queue.put(("command", f"\n$ {command}\n"))
                
                # 使用exec_command执行命令，设置超时
                stdin, stdout, stderr = self.client.exec_command(command, timeout=10)
                
                # 设置通道超时，避免无限等待
                stdout.channel.settimeout(10)
                stderr.channel.settimeout(10)
                
                # 读取输出（带超时保护）
                import time
                output = ""
                error = ""
                
                try:
                    # 等待命令完成，但设置超时
                    start_time = time.time()
                    timeout_seconds = 10
                    
                    # 使用更简单的方法：等待通道关闭或超时
                    while not stdout.channel.exit_status_ready() and (time.time() - start_time) < timeout_seconds:
                        time.sleep(0.1)
                        
                        # 尝试读取标准输出
                        if stdout.channel.recv_ready():
                            try:
                                data = stdout.channel.recv(4096)
                                if data:
                                    output += data.decode('utf-8', errors='ignore')
                            except:
                                pass
                        
                        # 尝试读取错误输出
                        if stderr.channel.recv_stderr_ready():
                            try:
                                error_data = stderr.channel.recv_stderr(4096)
                                if error_data:
                                    error += error_data.decode('utf-8', errors='ignore')
                            except:
                                pass
                    
                    # 命令已完成或超时，读取剩余数据
                    # 读取标准输出的剩余数据
                    try:
                        while stdout.channel.recv_ready():
                            data = stdout.channel.recv(4096)
                            if data:
                                output += data.decode('utf-8', errors='ignore')
                            else:
                                break
                    except:
                        pass
                    
                    # 读取错误输出的剩余数据
                    try:
                        while stderr.channel.recv_stderr_ready():
                            error_data = stderr.channel.recv_stderr(4096)
                            if error_data:
                                error += error_data.decode('utf-8', errors='ignore')
                            else:
                                break
                    except:
                        pass
                            
                except Exception as read_err:
                    # 如果读取失败，尝试简单读取（作为备用方案）
                    try:
                        output_bytes = stdout.read()
                        if output_bytes:
                            output = output_bytes.decode('utf-8', errors='ignore')
                    except:
                        pass
                    
                    try:
                        error_bytes = stderr.read()
                        if error_bytes:
                            error = error_bytes.decode('utf-8', errors='ignore')
                    except:
                        pass
                
                # 显示输出
                if output:
                    self.output_queue.put(("output", output))
                if error:
                    self.output_queue.put(("error", error))
            except Exception as e:
                error_msg = f"执行命令失败: {e}"
                self.output_queue.put(("error", error_msg + "\n"))
                # 不在后台线程中显示消息框，使用队列通知主线程
                self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        
        # 启动后台线程执行命令
        threading.Thread(target=execute_in_thread, daemon=True).start()
    
    def show_command_menu(self):
        """显示命令菜单"""
        menu_window = tk.Toplevel(self.root)
        menu_window.title("命令菜单")
        menu_window.geometry("400x300")
        
        commands = [
            ("查看进程", "ps aux"),
            ("查看磁盘", "df -h"),
            ("查看内存", "free -h"),
            ("查看网络", "netstat -tuln"),
            ("查看系统信息", "uname -a"),
            ("查看服务状态", "systemctl status"),
        ]
        
        for i, (name, cmd) in enumerate(commands):
            btn = ttk.Button(menu_window, text=name, command=lambda c=cmd: self.insert_command(c, menu_window), width=30)
            btn.pack(pady=5, padx=20)
    
    def insert_command(self, command, window):
        """插入命令到输入框"""
        self.cmd_input_var.set(command)
        window.destroy()
    
    def clear_cmd_output(self):
        """清空命令输出"""
        self.cmd_output_text.delete("1.0", tk.END)
    
    def on_key_press(self, event):
        """处理按键事件，允许在输出面板直接输入"""
        # 允许所有按键，让用户可以在输出面板输入
        return None
    
    def on_enter_key(self, event):
        """处理回车键，提取并发送命令"""
        import re
        
        if not self.is_connected:
            return "break"
        
        # 获取当前光标位置
        cursor_pos = self.cmd_output_text.index(tk.INSERT)
        line_num = int(cursor_pos.split('.')[0])
        col_num = int(cursor_pos.split('.')[1])
        
        # 获取总行数
        total_lines = int(self.cmd_output_text.index(tk.END).split('.')[0]) - 1
        
        command = ""
        
        # 方法1: 从当前光标所在行提取（最优先）
        line_start = f"{line_num}.0"
        line_end = f"{line_num}.end"
        line_content = self.cmd_output_text.get(line_start, line_end)
        
        # 查找提示符：$、#、>、或者"请输入"、"Please enter"等
        prompt_patterns = [
            (r'[#$>]', 'prompt'),  # 标准提示符
            (r'请输入[^:]*[:：]\s*', 'prompt_colon'),  # "请输入命令编号: "
            (r'Please\s+enter[^:]*[:：]\s*', 'prompt_colon'),  # "Please enter command number: "
        ]
        
        # 尝试从当前行提取
        for pattern, pattern_type in prompt_patterns:
            match = re.search(pattern, line_content)
            if match:
                prompt_end = match.end()
                if pattern_type == 'prompt':
                    # 标准提示符，提取到光标位置
                    if col_num > prompt_end:
                        command = line_content[prompt_end:col_num].strip()
                    else:
                        command = line_content[prompt_end:].strip()
                else:
                    # 冒号提示符，提取到光标位置或行尾
                    if col_num > prompt_end:
                        command = line_content[prompt_end:col_num].strip()
                    else:
                        command = line_content[prompt_end:].strip()
                
                if command:
                    break
        
        # 方法2: 如果当前行没提取到，从最后几行往前查找（最多查找最后20行）
        if not command:
            search_start = max(1, total_lines - 19)
            for i in range(total_lines, search_start - 1, -1):
                line_start = f"{i}.0"
                line_end = f"{i}.end"
                check_line = self.cmd_output_text.get(line_start, line_end)
                
                # 查找提示符
                for pattern, pattern_type in prompt_patterns:
                    match = re.search(pattern, check_line)
                    if match:
                        prompt_end = match.end()
                        if i == line_num and col_num > prompt_end:
                            # 当前行，提取到光标位置
                            command = check_line[prompt_end:col_num].strip()
                        else:
                            # 其他行，提取整行提示符之后的内容
                            command = check_line[prompt_end:].strip()
                        
                        if command:
                            break
                
                if command:
                    break
        
        # 方法3: 如果还是没找到，尝试提取当前行的纯数字/字母内容（可能是用户直接输入）
        if not command and line_content.strip():
            # 移除提示符和冒号
            potential_cmd = re.sub(r'^[^:]*[:：]\s*', '', line_content.strip())
            potential_cmd = re.sub(r'^[#$>]\s*', '', potential_cmd)
            potential_cmd = potential_cmd.strip()
            
            # 如果看起来像是命令（数字、字母、常见命令字符），使用它
            if potential_cmd and re.match(r'^[\w\s\-_\.\/]+$', potential_cmd):
                # 过滤掉明显不是命令的内容
                skip_patterns = ['已取消!', '取消', 'Cancelled!', 'Cancel', '请输入', 'Please enter']
                if not any(skip in potential_cmd for skip in skip_patterns):
                    command = potential_cmd
        
        # 最终清理和验证
        if command:
            command = command.strip()
            # 过滤掉明显无效的内容
            skip_commands = ['已取消!', '取消', 'Cancelled!', 'Cancel', '请输入命令编号', 
                           'Please enter command number', '命令编号', 'command number']
            if any(skip in command for skip in skip_commands):
                command = ""
        
            # 发送命令
        if command:
            try:
                self.send_command_to_server(command)
            except Exception as e:
                # 如果发送失败，在输出面板显示错误
                self.cmd_output_text.insert(tk.END, f"\n发送命令失败: {e}\n", "error")
                self.cmd_output_text.see(tk.END)
        
        return "break"
    
    def send_command_to_server(self, command):
        """发送命令到服务器"""
        if not self.is_connected or not self.shell:
            return
        
        try:
            # 确保命令不为空
            if not command or not command.strip():
                return
            
            # 发送命令到服务器（添加换行符）
            command_to_send = command.strip() + "\n"
            self.shell.send(command_to_send)
        except Exception as e:
            # 如果发送失败，在输出面板显示错误
            error_msg = f"发送命令失败: {e}\n"
            self.cmd_output_text.insert(tk.END, error_msg, "error")
            self.cmd_output_text.see(tk.END)
    
    # ==================== 系统监控功能 ====================
    def toggle_monitoring(self):
        """切换监控状态"""
        if not self.is_connected:
            messagebox.showwarning("提示", "请先连接服务器")
            return
        
        if self.monitoring_active:
            self.stop_monitoring()
        else:
            self.start_monitoring()
    
    def start_monitoring(self):
        """开始监控"""
        self.monitoring_active = True
        if hasattr(self, 'monitor_btn'):
            self.monitor_btn.config(text="停止监控")
        
        def monitoring_loop():
            while self.monitoring_active and self.is_connected:
                try:
                    self.update_monitoring()
                    import time
                    time.sleep(2)
                except Exception as e:
                    if self.monitoring_active:
                        self.root.after(0, lambda: messagebox.showerror("监控错误", f"监控失败: {e}"))
                    break
        
        self.monitoring_thread = threading.Thread(target=monitoring_loop, daemon=True)
        self.monitoring_thread.start()
    
    def stop_monitoring(self):
        """停止监控"""
        self.monitoring_active = False
        if hasattr(self, 'monitor_btn'):
            self.monitor_btn.config(text="开始监控")
    
    def update_monitoring(self):
        """更新监控数据（带超时保护，避免阻塞）"""
        if not self.client or not self.is_connected:
            return
        
        cpu_usage = 0.0
        mem_usage = 0.0
        mem_used = 0
        mem_total = 0
        disk_usage = 0.0
        
        try:
            import time
            timeout = 5  # 每个命令最多5秒超时
            
            # CPU使用率 - 使用更简单的命令，设置超时
            try:
                # 优先使用更快的命令
                cpu_cmd = "grep 'cpu ' /proc/stat | awk '{usage=($2+$4)*100/($2+$3+$4+$5)} END {print usage}' 2>/dev/null || echo 0"
                stdin, stdout, stderr = self.client.exec_command(cpu_cmd, timeout=timeout)
                stdout.channel.settimeout(timeout)
                
                start_time = time.time()
                cpu_output = ""
                while time.time() - start_time < timeout:
                    if stdout.channel.recv_ready():
                        data = stdout.channel.recv(1024)
                        if data:
                            cpu_output += data.decode('utf-8', errors='ignore')
                        else:
                            break
                    elif stdout.channel.exit_status_ready():
                        break
                    else:
                        time.sleep(0.1)
                
                cpu_output = cpu_output.strip()
                if cpu_output:
                    try:
                        cpu_usage = float(cpu_output)
                    except:
                        cpu_usage = 0.0
                else:
                    cpu_usage = 0.0
            except Exception as e:
                cpu_usage = 0.0
            
            # 内存使用情况 - 简化命令，添加超时
            try:
                mem_cmd = "free -m 2>/dev/null | grep Mem | awk '{printf \"%.1f %.0f %.0f\", $3/$2*100, $3, $2}' || echo '0 0 0'"
                stdin, stdout, stderr = self.client.exec_command(mem_cmd, timeout=timeout)
                stdout.channel.settimeout(timeout)
                
                start_time = time.time()
                mem_output = ""
                while time.time() - start_time < timeout:
                    if stdout.channel.recv_ready():
                        data = stdout.channel.recv(1024)
                        if data:
                            mem_output += data.decode('utf-8', errors='ignore')
                        else:
                            break
                    elif stdout.channel.exit_status_ready():
                        break
                    else:
                        time.sleep(0.1)
                
                mem_output = mem_output.strip()
                if mem_output:
                    parts = mem_output.split()
                    if len(parts) >= 3:
                        try:
                            mem_usage = float(parts[0])
                            mem_used = int(float(parts[1]))
                            mem_total = int(float(parts[2]))
                        except:
                            pass
            except Exception as e:
                pass
            
            # 磁盘使用情况 - 简化命令，添加超时
            try:
                disk_cmd = "df -h / 2>/dev/null | tail -1 | awk '{print $5}' | sed 's/%//' || echo '0'"
                stdin, stdout, stderr = self.client.exec_command(disk_cmd, timeout=timeout)
                stdout.channel.settimeout(timeout)
                
                start_time = time.time()
                disk_output = ""
                while time.time() - start_time < timeout:
                    if stdout.channel.recv_ready():
                        data = stdout.channel.recv(1024)
                        if data:
                            disk_output += data.decode('utf-8', errors='ignore')
                        else:
                            break
                    elif stdout.channel.exit_status_ready():
                        break
                    else:
                        time.sleep(0.1)
                
                disk_output = disk_output.strip()
                if disk_output:
                    try:
                        disk_usage = float(disk_output)
                    except:
                        disk_usage = 0.0
                else:
                    disk_usage = 0.0
            except Exception as e:
                disk_usage = 0.0
            
            # 更新UI（在主线程中执行）
            self.root.after(0, lambda: self.update_monitoring_ui(cpu_usage, mem_usage, mem_used, mem_total, disk_usage))
            
        except Exception as e:
            # 发生错误时，更新错误状态
            self.root.after(0, lambda: self.update_monitoring_error(str(e)))
    
    def update_monitoring_ui(self, cpu_usage, mem_usage, mem_used, mem_total, disk_usage):
        """更新监控UI"""
        if hasattr(self, 'cpu_var') and hasattr(self, 'cpu_progress'):
            cpu_value = min(100, max(0, cpu_usage))
            self.cpu_var.set(f"{cpu_value:.1f}%")
            self.cpu_progress['value'] = cpu_value
            self.cpu_status_var.set(f"使用率: {cpu_value:.1f}%")
            # 根据使用率设置颜色
            if cpu_value > 80:
                self.cpu_var.set(f"{cpu_value:.1f}%")
                # 可以设置进度条颜色
        
        if hasattr(self, 'mem_var') and hasattr(self, 'mem_progress'):
            mem_value = min(100, max(0, mem_usage))
            self.mem_var.set(f"{mem_value:.1f}%")
            self.mem_progress['value'] = mem_value
            self.mem_status_var.set(f"{mem_used}MB / {mem_total}MB")
        
        if hasattr(self, 'disk_var') and hasattr(self, 'disk_progress'):
            disk_value = min(100, max(0, disk_usage))
            self.disk_var.set(f"{disk_value:.1f}%")
            self.disk_progress['value'] = disk_value
            self.disk_status_var.set(f"使用率: {disk_value:.1f}%")
    
    def update_monitoring_error(self, error_msg):
        """更新监控错误信息"""
        if hasattr(self, 'cpu_var'):
            self.cpu_var.set("错误")
            self.cpu_status_var.set("获取失败")
        if hasattr(self, 'mem_var'):
            self.mem_var.set("错误")
            self.mem_status_var.set("获取失败")
        if hasattr(self, 'disk_var'):
            self.disk_var.set("错误")
            self.disk_status_var.set("获取失败")
    def manage_users(self):
        """管理SSH登录用户名和密码（连接记录中的用户）"""
        manage_window = tk.Toplevel(self.root)
        manage_window.title("用户管理 - SSH登录凭据")
        manage_window.geometry("700x500")
        manage_window.transient(self.root)
        
        # 用户列表
        list_frame = ttk.Frame(manage_window, padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 存储用户信息，用于填充连接信息（先定义，供后续使用）
        user_data_map = {}
        show_password_var = tk.BooleanVar(value=False)  # 默认隐藏密码
        
        # 标题和显示密码选项
        title_frame = ttk.Frame(list_frame)
        title_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(title_frame, text="SSH登录用户列表（用于快速切换IP时的用户名和密码）:", font=("", 10, "bold")).pack(side=tk.LEFT, anchor=tk.W)
        
        # 创建Treeview显示用户
        tree = ttk.Treeview(list_frame, columns=("host", "port", "password"), show="tree headings", height=15)
        tree.heading("#0", text="用户名")
        tree.heading("host", text="主机/IP")
        tree.heading("port", text="端口")
        tree.heading("password", text="密码")
        tree.column("#0", width=120)
        tree.column("host", width=150)
        tree.column("port", width=80)
        tree.column("password", width=200)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        def load_users():
            """加载SSH登录用户列表（从连接记录中提取）"""
            for item in tree.get_children():
                tree.delete(item)
            
            # 清空用户数据映射
            user_data_map.clear()
            
            # 从连接记录中提取唯一的用户名和密码组合
            user_credentials = {}
            for conn in self.connections:
                username = conn.get('username', '').strip()
                password = conn.get('password', '').strip()
                host = conn.get('host', '').strip()
                port = str(conn.get('port', 22))
                
                if username:
                    # 使用用户名作为key，如果有多个IP使用同一用户名，显示所有IP
                    key = username
                    if key not in user_credentials:
                        user_credentials[key] = {
                            'username': username,
                            'password': password,
                            'hosts': []
                        }
                    if host:
                        user_credentials[key]['hosts'].append({
                            'host': host,
                            'port': port
                        })
            
            # 显示用户列表，并存储用户数据
            show_pass = show_password_var.get()
            for username, info in sorted(user_credentials.items()):
                hosts_list = info['hosts']
                hosts_str = ', '.join([f"{h['host']}:{h['port']}" for h in hosts_list]) if hosts_list else '(未设置IP)'
                # root用户：始终显示密码（因为在文件中是明文）
                # manage用户和其他用户：根据复选框状态显示
                username_lower = username.lower()
                if username_lower == 'root':
                    # root用户始终显示密码
                    password_display = info['password'] if info['password'] else '(无密码)'
                elif show_pass:
                    # 其他用户：复选框选中时显示密码
                    password_display = info['password'] if info['password'] else '(无密码)'
                else:
                    # 其他用户：复选框未选中时隐藏密码（显示星号）
                    password_display = '*' * len(info['password']) if info['password'] else '(无密码)'
                item_id = tree.insert("", tk.END, text=username, values=(hosts_str, '', password_display))
                
                # 存储用户数据，用于填充连接信息
                # 如果有多个IP，使用第一个；如果没有IP，使用空值
                if hosts_list:
                    first_host = hosts_list[0]
                    user_data_map[item_id] = {
                        'username': info['username'],
                        'password': info['password'],
                        'host': first_host['host'],
                        'port': first_host['port']
                    }
                else:
                    user_data_map[item_id] = {
                        'username': info['username'],
                        'password': info['password'],
                        'host': '',
                        'port': '22'
                    }
        
        # 显示密码复选框（需要在 load_users 定义后创建）
        show_password_check = ttk.Checkbutton(title_frame, text="显示密码", variable=show_password_var, 
                                              command=load_users)
        show_password_check.pack(side=tk.RIGHT, padx=(10, 0))
        
        def on_user_selected(event=None):
            """当选择用户时，自动填充到SSH连接输入框"""
            selection = tree.selection()
            if not selection:
                return
            
            item_id = selection[0]
            user_data = user_data_map.get(item_id)
            if user_data:
                # 填充到SSH连接输入框
                self.host_var.set(user_data['host'])
                self.port_var.set(user_data['port'])
                self.username_var.set(user_data['username'])
                self.password_var.set(user_data['password'])
                
                # 如果用户有多个IP，询问是否要选择其他IP
                username = user_data['username']
                user_connections = [conn for conn in self.connections 
                                  if conn.get('username', '').strip() == username 
                                  and conn.get('host', '').strip()]
                
                if len(user_connections) > 1:
                    # 有多个IP，询问用户是否要选择其他IP
                    if messagebox.askyesno("选择IP", 
                                          f"用户 {username} 有 {len(user_connections)} 个IP地址。\n"
                                          f"已填充: {user_data['host']}:{user_data['port']}\n\n"
                                          f"是否要选择其他IP？"):
                        # 显示IP选择窗口
                        ip_window = tk.Toplevel(manage_window)
                        ip_window.title("选择IP地址")
                        ip_window.geometry("400x300")
                        ip_window.transient(manage_window)
                        ip_window.grab_set()
                        
                        ttk.Label(ip_window, text=f"请选择 {username} 的IP地址:", 
                                 font=("", 10, "bold")).pack(pady=10)
                        
                        ip_listbox = tk.Listbox(ip_window, height=10)
                        ip_listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
                        
                        for conn in user_connections:
                            host = conn.get('host', '')
                            port = conn.get('port', 22)
                            display_text = f"{host}:{port}"
                            if conn.get('notes'):
                                display_text += f" ({conn.get('notes')})"
                            ip_listbox.insert(tk.END, display_text)
                        
                        def select_ip():
                            selection = ip_listbox.curselection()
                            if selection:
                                selected_conn = user_connections[selection[0]]
                                self.host_var.set(selected_conn.get('host', ''))
                                self.port_var.set(str(selected_conn.get('port', 22)))
                                self.username_var.set(selected_conn.get('username', ''))
                                self.password_var.set(selected_conn.get('password', ''))
                                ip_window.destroy()
                                messagebox.showinfo("提示", "IP地址已填充到连接输入框")
                        
                        btn_frame_ip = ttk.Frame(ip_window)
                        btn_frame_ip.pack(pady=10)
                        ttk.Button(btn_frame_ip, text="确定", command=select_ip, width=12).pack(side=tk.LEFT, padx=5)
                        ttk.Button(btn_frame_ip, text="取消", command=ip_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
                        
                        # 默认选中第一个
                        ip_listbox.selection_set(0)
                        ip_listbox.bind('<Double-Button-1>', lambda e: select_ip())
                    else:
                        messagebox.showinfo("提示", f"已填充用户信息到连接输入框:\n用户名: {username}\nIP: {user_data['host']}:{user_data['port']}")
                else:
                    # 只有一个或没有IP，直接填充
                    messagebox.showinfo("提示", f"已填充用户信息到连接输入框:\n用户名: {username}\nIP: {user_data['host']}:{user_data['port']}" if user_data['host'] else f"已填充用户信息到连接输入框:\n用户名: {username}")
        
        # 绑定选择事件
        tree.bind('<<TreeviewSelect>>', on_user_selected)
        
        # 按钮框架
        btn_frame = ttk.Frame(manage_window, padding="10")
        btn_frame.pack(fill=tk.X)
        
        def add_user():
            """添加SSH登录用户（用户名和密码）"""
            add_window = tk.Toplevel(manage_window)
            add_window.title("添加SSH登录用户")
            add_window.geometry("400x150")
            add_window.transient(manage_window)
            add_window.grab_set()
            
            ttk.Label(add_window, text="用户名:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
            username_var = tk.StringVar()
            ttk.Entry(add_window, textvariable=username_var, width=25).grid(row=0, column=1, padx=5, pady=10)
            
            ttk.Label(add_window, text="密码:").grid(row=1, column=0, padx=5, pady=10, sticky=tk.W)
            password_var = tk.StringVar()
            ttk.Entry(add_window, textvariable=password_var, show="*", width=25).grid(row=1, column=1, padx=5, pady=10)
            
            def do_add():
                username = username_var.get().strip()
                password = password_var.get().strip()
                
                if not username:
                    messagebox.showwarning("提示", "请输入用户名")
                    return
                
                # 检查是否已存在该用户名
                for conn in self.connections:
                    if conn.get('username', '').strip() == username:
                        # 更新密码
                        conn['password'] = password
                        self.save_connections()
                        messagebox.showinfo("成功", f"用户 {username} 的密码已更新")
                        add_window.destroy()
                        load_users()
                        return
                
                # 添加新的连接记录（使用空IP，用户可以在连接时填写）
                new_conn = {
                    'name': f'用户: {username}',
                    'host': '',
                    'port': 22,
                    'username': username,
                    'password': password,
                    'key_file': '',
                    'notes': f'SSH登录用户: {username}',
                    'last_connected': ''
                }
                self.connections.append(new_conn)
                self.save_connections()
                self.refresh_connection_combo()
                messagebox.showinfo("成功", f"用户 {username} 已添加")
                add_window.destroy()
                load_users()
            
            btn_frame2 = ttk.Frame(add_window)
            btn_frame2.grid(row=2, column=0, columnspan=2, pady=20)
            ttk.Button(btn_frame2, text="确定", command=do_add, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame2, text="取消", command=add_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
        
        def delete_user():
            """删除SSH登录用户（从所有连接记录中删除该用户名的记录）"""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择一个用户")
                return
            
            username = tree.item(selection[0], "text")
            if not username or username.startswith("("):
                messagebox.showwarning("提示", "请选择一个有效的用户")
                return
            
            # 统计该用户名的连接记录数量
            count = sum(1 for conn in self.connections if conn.get('username', '').strip() == username)
            
            if not messagebox.askyesno("确认", f"确定要删除用户 {username} 吗？\n这将删除该用户名的所有连接记录（共{count}条）！\n\n删除前会自动创建备份，可以通过'恢复用户'功能恢复"):
                return
            
            # 删除前创建备份
            self.save_connections()  # 这会自动创建备份
            
            # 删除所有该用户名的连接记录
            self.connections = [conn for conn in self.connections if conn.get('username', '').strip() != username]
            self.save_connections()
            self.refresh_connection_combo()
            messagebox.showinfo("成功", f"用户 {username} 的所有连接记录已删除\n\n已创建备份，可通过'恢复用户'功能恢复")
            load_users()
        
        def edit_user():
            """编辑SSH登录用户的密码"""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("提示", "请先选择一个用户")
                return
            
            username = tree.item(selection[0], "text")
            if not username or username.startswith("("):
                messagebox.showwarning("提示", "请选择一个有效的用户")
                return
            
            # 查找该用户的密码
            password = ''
            for conn in self.connections:
                if conn.get('username', '').strip() == username:
                    password = conn.get('password', '')
                    break
            
            edit_window = tk.Toplevel(manage_window)
            edit_window.title(f"编辑用户: {username}")
            edit_window.geometry("400x120")
            edit_window.transient(manage_window)
            edit_window.grab_set()
            
            ttk.Label(edit_window, text="新密码:").grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
            password_var = tk.StringVar(value=password)
            ttk.Entry(edit_window, textvariable=password_var, show="*", width=25).grid(row=0, column=1, padx=5, pady=10)
            
            def do_edit():
                new_password = password_var.get().strip()
                
                # 更新所有该用户名的连接记录的密码
                updated = False
                for conn in self.connections:
                    if conn.get('username', '').strip() == username:
                        conn['password'] = new_password
                        updated = True
                
                if updated:
                    self.save_connections()
                    messagebox.showinfo("成功", f"用户 {username} 的密码已更新")
                    edit_window.destroy()
                    load_users()
                else:
                    messagebox.showerror("错误", "未找到该用户")
            
            btn_frame2 = ttk.Frame(edit_window)
            btn_frame2.grid(row=1, column=0, columnspan=2, pady=20)
            ttk.Button(btn_frame2, text="确定", command=do_edit, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame2, text="取消", command=edit_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
        
        def restore_users():
            """从备份恢复用户"""
            backup_connections = self.load_backup_connections()
            if not backup_connections:
                messagebox.showwarning("提示", "没有找到备份文件，无法恢复用户")
                return
            
            # 显示备份中的用户列表
            backup_users = {}
            for conn in backup_connections:
                username = conn.get('username', '').strip()
                if username:
                    if username not in backup_users:
                        backup_users[username] = []
                    backup_users[username].append(conn)
            
            if not backup_users:
                messagebox.showinfo("提示", "备份文件中没有用户记录")
                return
            
            # 创建恢复窗口
            restore_window = tk.Toplevel(manage_window)
            restore_window.title("恢复用户")
            restore_window.geometry("600x500")
            restore_window.transient(manage_window)
            restore_window.grab_set()
            
            ttk.Label(restore_window, text="从备份恢复用户（选择要恢复的用户）:", font=("", 10, "bold")).pack(anchor=tk.W, padx=10, pady=10)
            
            # 用户列表（带复选框）
            list_frame = ttk.Frame(restore_window)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            restore_tree = ttk.Treeview(list_frame, columns=("host", "password"), show="tree headings", height=15)
            restore_tree.heading("#0", text="用户名")
            restore_tree.heading("host", text="主机/IP")
            restore_tree.heading("password", text="密码")
            restore_tree.column("#0", width=120)
            restore_tree.column("host", width=200)
            restore_tree.column("password", width=200)
            
            scrollbar2 = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=restore_tree.yview)
            restore_tree.configure(yscrollcommand=scrollbar2.set)
            restore_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar2.pack(side=tk.RIGHT, fill=tk.Y)
            
            # 存储备份用户数据
            backup_items = {}
            for username, conns in sorted(backup_users.items()):
                for conn in conns:
                    host = conn.get('host', '').strip() or '(未设置IP)'
                    password_display = '*' * len(conn.get('password', '')) if conn.get('password', '') else '(无密码)'
                    item_id = restore_tree.insert("", tk.END, text=username, values=(host, password_display))
                    backup_items[item_id] = conn
            
            def do_restore():
                selection = restore_tree.selection()
                if not selection:
                    messagebox.showwarning("提示", "请选择要恢复的用户")
                    return
                
                restored_count = 0
                skipped_count = 0
                
                for item_id in selection:
                    conn = backup_items.get(item_id)
                    if not conn:
                        continue
                    
                    username = conn.get('username', '').strip()
                    host = conn.get('host', '').strip()
                    port = conn.get('port', 22)
                    
                    # 检查是否已存在相同的连接记录
                    exists = False
                    for existing_conn in self.connections:
                        if (existing_conn.get('username', '').strip() == username and 
                            existing_conn.get('host', '').strip() == host and 
                            existing_conn.get('port', 22) == port):
                            exists = True
                            skipped_count += 1
                            break
                    
                    if not exists:
                        self.connections.append(conn)
                        restored_count += 1
                
                if restored_count > 0:
                    self.save_connections()
                    self.refresh_connection_combo()
                    load_users()
                    messagebox.showinfo("成功", f"已恢复 {restored_count} 个用户\n跳过 {skipped_count} 个已存在的用户")
                else:
                    messagebox.showinfo("提示", f"没有恢复任何用户（所有用户都已存在）")
                
                restore_window.destroy()
            
            def restore_all():
                """恢复所有用户"""
                if not messagebox.askyesno("确认", f"确定要恢复备份中的所有 {len(backup_items)} 个用户吗？\n已存在的用户将被跳过"):
                    return
                
                restored_count = 0
                skipped_count = 0
                
                for item_id, conn in backup_items.items():
                    username = conn.get('username', '').strip()
                    host = conn.get('host', '').strip()
                    port = conn.get('port', 22)
                    
                    # 检查是否已存在
                    exists = False
                    for existing_conn in self.connections:
                        if (existing_conn.get('username', '').strip() == username and 
                            existing_conn.get('host', '').strip() == host and 
                            existing_conn.get('port', 22) == port):
                            exists = True
                            skipped_count += 1
                            break
                    
                    if not exists:
                        self.connections.append(conn)
                        restored_count += 1
                
                if restored_count > 0:
                    self.save_connections()
                    self.refresh_connection_combo()
                    load_users()
                    messagebox.showinfo("成功", f"已恢复 {restored_count} 个用户\n跳过 {skipped_count} 个已存在的用户")
                else:
                    messagebox.showinfo("提示", f"没有恢复任何用户（所有用户都已存在）")
                
                restore_window.destroy()
            
            btn_frame_restore = ttk.Frame(restore_window)
            btn_frame_restore.pack(fill=tk.X, padx=10, pady=10)
            ttk.Button(btn_frame_restore, text="恢复选中", command=do_restore, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame_restore, text="恢复全部", command=restore_all, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame_restore, text="取消", command=restore_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="刷新", command=load_users, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="添加用户", command=add_user, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="编辑密码", command=edit_user, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="删除用户", command=delete_user, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="恢复用户", command=restore_users, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=manage_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
        
        # 绑定显示密码复选框事件
        show_password_var.trace_add("write", lambda *args: load_users())
        
        # 初始加载
        load_users()
    
    def game_server_manage(self):
        """游戏服务器管理"""
        if not self.is_connected or not self.client:
            messagebox.showwarning("提示", "请先连接SSH服务器")
            return
        
        # 创建管理窗口
        manage_window = tk.Toplevel(self.root)
        manage_window.title("游戏服务器管理")
        manage_window.geometry("900x700")
        manage_window.minsize(800, 600)
        manage_window.transient(self.root)
        
        # 使用Notebook创建标签页
        notebook = ttk.Notebook(manage_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 标签页1: 创建网站
        create_site_frame = ttk.Frame(notebook, padding="15")
        notebook.add(create_site_frame, text="🌐 创建网站")
        
        # 创建网站表单
        form_frame = ttk.LabelFrame(create_site_frame, text="网站配置", padding="15")
        form_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # IP地址
        ttk.Label(form_frame, text="IP地址:", font=("Microsoft YaHei", 9)).grid(row=0, column=0, sticky=tk.W, pady=8, padx=5)
        ip_var = tk.StringVar(value="")
        ip_entry = ttk.Entry(form_frame, textvariable=ip_var, width=30, font=("Consolas", 9))
        ip_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=8, padx=5)
        form_frame.columnconfigure(1, weight=1)
        
        # 端口
        ttk.Label(form_frame, text="端口:", font=("Microsoft YaHei", 9)).grid(row=1, column=0, sticky=tk.W, pady=8, padx=5)
        port_var = tk.StringVar(value="80")
        port_entry = ttk.Entry(form_frame, textvariable=port_var, width=30, font=("Consolas", 9))
        port_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=8, padx=5)
        
        # 路径
        ttk.Label(form_frame, text="网站路径:", font=("Microsoft YaHei", 9)).grid(row=2, column=0, sticky=tk.W, pady=8, padx=5)
        path_var = tk.StringVar(value="/www/wwwroot/")
        path_entry = ttk.Entry(form_frame, textvariable=path_var, width=30, font=("Consolas", 9))
        path_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=8, padx=5)
        
        # PHP版本选择
        ttk.Label(form_frame, text="PHP版本:", font=("Microsoft YaHei", 9)).grid(row=3, column=0, sticky=tk.W, pady=8, padx=5)
        php_var = tk.StringVar(value="php74")
        php_combo = ttk.Combobox(form_frame, textvariable=php_var, width=27, state="readonly", font=("Consolas", 9))
        php_combo['values'] = ['php53', 'php54', 'php55', 'php56', 'php70', 'php71', 'php72', 'php73', 'php74', 'php80', 'php81', 'php82', 'php83']
        php_combo.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=8, padx=5)
        php_combo.current(8)  # 默认选择php74
        
        # 网站名称（可选）
        ttk.Label(form_frame, text="网站名称:", font=("Microsoft YaHei", 9)).grid(row=4, column=0, sticky=tk.W, pady=8, padx=5)
        site_name_var = tk.StringVar(value="")
        site_name_entry = ttk.Entry(form_frame, textvariable=site_name_var, width=30, font=("Consolas", 9))
        site_name_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=8, padx=5)
        
        # 备注
        ttk.Label(form_frame, text="备注:", font=("Microsoft YaHei", 9)).grid(row=5, column=0, sticky=tk.W, pady=8, padx=5)
        notes_var = tk.StringVar(value="")
        notes_entry = ttk.Entry(form_frame, textvariable=notes_var, width=30, font=("Consolas", 9))
        notes_entry.grid(row=5, column=1, sticky=(tk.W, tk.E), pady=8, padx=5)
        
        # 输出区域
        output_frame = ttk.LabelFrame(create_site_frame, text="执行结果", padding="10")
        output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        output_text = scrolledtext.ScrolledText(output_frame, height=15, font=("Consolas", 9), wrap=tk.WORD)
        output_text.pack(fill=tk.BOTH, expand=True)
        
        def create_site():
            """创建网站"""
            ip = ip_var.get().strip()
            port = port_var.get().strip()
            path = path_var.get().strip()
            php_version = php_var.get().strip()
            site_name = site_name_var.get().strip()
            notes = notes_var.get().strip()
            
            # 验证必填参数
            if not ip:
                messagebox.showwarning("提示", "请输入IP地址")
                return
            
            if not port:
                messagebox.showwarning("提示", "请输入端口")
                return
            
            if not path:
                messagebox.showwarning("提示", "请输入网站路径")
                return
            
            if not php_version:
                messagebox.showwarning("提示", "请选择PHP版本")
                return
            
            # 如果没有网站名称，使用IP+端口作为名称
            if not site_name:
                site_name = f"{ip}_{port}"
            
            # 构建创建网站的命令
            # 使用宝塔面板的bt命令创建网站
            # 格式: bt site add --domain <域名/IP:端口> --path <路径> --php <PHP版本> --name <网站名称>
            # 如果端口不是80，需要在域名中指定端口
            domain = f"{ip}:{port}" if port != "80" else ip
            command = f"bt site add --domain {domain} --path {path} --php {php_version} --name {site_name}"
            
            # 如果bt命令不存在，尝试使用nginx配置方式
            # 备用方案：直接创建nginx配置文件
            backup_command = f"""
# 创建nginx配置文件
cat > /www/server/panel/vhost/nginx/{site_name}.conf <<EOF
server {{
    listen {port};
    server_name {ip};
    root {path};
    index index.php index.html index.htm;
    
    location ~ \.php$ {{
        fastcgi_pass unix:/tmp/php-cgi-{php_version}.sock;
        fastcgi_index index.php;
        fastcgi_param SCRIPT_FILENAME $document_root$fastcgi_script_name;
        include fastcgi_params;
    }}
}}
EOF
# 重载nginx配置
nginx -t && nginx -s reload
"""
            
            output_text.insert(tk.END, f"准备创建网站...\n")
            output_text.insert(tk.END, f"IP地址: {ip}\n")
            output_text.insert(tk.END, f"端口: {port}\n")
            output_text.insert(tk.END, f"路径: {path}\n")
            output_text.insert(tk.END, f"PHP版本: {php_version}\n")
            output_text.insert(tk.END, f"网站名称: {site_name}\n")
            output_text.insert(tk.END, f"执行命令: {command}\n\n")
            output_text.see(tk.END)
            
            # 在后台线程中执行命令
            def execute_create():
                try:
                    # 先检查bt命令是否存在
                    check_bt = "which bt || echo 'not found'"
                    stdin_check, stdout_check, stderr_check = self.client.exec_command(check_bt, timeout=5)
                    bt_exists = stdout_check.read().decode('utf-8', errors='ignore').strip()
                    
                    if 'not found' in bt_exists or not bt_exists:
                        # bt命令不存在，使用备用方案
                        def show_backup_msg():
                            output_text.insert(tk.END, "检测到bt命令不存在，使用备用方案（nginx配置）...\n")
                            output_text.see(tk.END)
                        self.root.after(0, show_backup_msg)
                        # 使用备用命令（需要根据实际情况调整）
                        actual_command = backup_command
                    else:
                        # 使用bt命令
                        actual_command = command
                    
                    stdin, stdout, stderr = self.client.exec_command(actual_command, timeout=60)
                    stdout.channel.settimeout(60)
                    stderr.channel.settimeout(60)
                    
                    output = ""
                    error = ""
                    
                    import time
                    start_time = time.time()
                    timeout_seconds = 60
                    
                    # 等待命令完成或超时
                    while not stdout.channel.exit_status_ready() and (time.time() - start_time) < timeout_seconds:
                        if stdout.channel.recv_ready():
                            data = stdout.channel.recv(4096)
                            if data:
                                decoded_data = data.decode('utf-8', errors='ignore')
                                output += decoded_data
                                # 实时更新输出（修复闭包问题）
                                def update_output(text=decoded_data):
                                    output_text.insert(tk.END, text)
                                    output_text.see(tk.END)
                                self.root.after(0, update_output)
                        if stderr.channel.recv_ready():
                            data = stderr.channel.recv(4096)
                            if data:
                                decoded_error = data.decode('utf-8', errors='ignore')
                                error += decoded_error
                                # 实时更新错误输出（修复闭包问题）
                                def update_error(text=decoded_error):
                                    output_text.insert(tk.END, f"[错误] {text}")
                                    output_text.see(tk.END)
                                self.root.after(0, update_error)
                        time.sleep(0.1)
                    
                    # 读取剩余输出
                    while stdout.channel.recv_ready():
                        data = stdout.channel.recv(4096)
                        if data:
                            decoded_data = data.decode('utf-8', errors='ignore')
                            output += decoded_data
                            def update_output(text=decoded_data):
                                output_text.insert(tk.END, text)
                                output_text.see(tk.END)
                            self.root.after(0, update_output)
                    
                    while stderr.channel.recv_ready():
                        data = stderr.channel.recv(4096)
                        if data:
                            decoded_error = data.decode('utf-8', errors='ignore')
                            error += decoded_error
                            def update_error(text=decoded_error):
                                output_text.insert(tk.END, f"[错误] {text}")
                                output_text.see(tk.END)
                            self.root.after(0, update_error)
                    
                    # 获取退出状态
                    exit_status = stdout.channel.recv_exit_status()
                    
                    # 显示最终结果
                    def show_result():
                        output_text.insert(tk.END, f"\n{'='*50}\n")
                        if exit_status == 0:
                            output_text.insert(tk.END, f"✓ 网站创建成功！\n")
                            messagebox.showinfo("成功", "网站创建成功！")
                        else:
                            output_text.insert(tk.END, f"✗ 网站创建失败，退出码: {exit_status}\n")
                            if error:
                                output_text.insert(tk.END, f"错误信息: {error}\n")
                            messagebox.showwarning("提示", f"网站创建完成，退出码: {exit_status}")
                        output_text.see(tk.END)
                    
                    self.root.after(0, show_result)
                    
                except Exception as e:
                    error_msg = f"创建网站失败: {str(e)}"
                    self.root.after(0, lambda: output_text.insert(tk.END, f"\n✗ {error_msg}\n"))
                    self.root.after(0, lambda: output_text.see(tk.END))
                    self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
                    import traceback
                    traceback.print_exc()
            
            # 启动后台线程
            threading.Thread(target=execute_create, daemon=True).start()
        
        # 按钮框架
        btn_frame = ttk.Frame(create_site_frame, padding="10")
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="创建网站", command=create_site, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="清空输出", command=lambda: output_text.delete(1.0, tk.END), width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="关闭", command=manage_window.destroy, width=15).pack(side=tk.RIGHT, padx=5)
    def file_browser(self):
        """文件浏览器"""
        if not self.is_connected:
            messagebox.showwarning("提示", "请先连接服务器")
            return
        
        if not self.client:
            messagebox.showerror("错误", "SSH连接未建立")
            return
        
        browser_window = tk.Toplevel(self.root)
        browser_window.title("文件浏览器")
        browser_window.geometry("1000x700")
        
        # 路径变量
        path_var = tk.StringVar(value="/")
        current_file_path = None
        
        # 创建Notebook用于文件列表和文件编辑
        notebook = ttk.Notebook(browser_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 标签页1: 文件列表
        list_frame = ttk.Frame(notebook, padding="10")
        notebook.add(list_frame, text="文件列表")
        
        # 路径输入和导航
        nav_frame = ttk.Frame(list_frame, padding="5")
        nav_frame.pack(fill=tk.X)
        
        ttk.Label(nav_frame, text="当前路径:").pack(side=tk.LEFT, padx=5)
        path_entry = ttk.Entry(nav_frame, textvariable=path_var, width=50)
        path_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        btn_frame = ttk.Frame(nav_frame)
        btn_frame.pack(side=tk.LEFT, padx=5)
        
        # 文件列表（使用Treeview）
        tree_frame = ttk.Frame(list_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建Treeview
        columns = ("类型", "权限", "大小", "修改时间")
        tree = ttk.Treeview(tree_frame, columns=columns, show="tree headings", height=20)
        
        # 排序状态（列名 -> 排序方向: True=升序, False=降序）
        sort_states = {"#0": True, "大小": True}
        
        def sort_treeview(col, reverse=False):
            """对Treeview进行排序"""
            # 获取所有项目
            items = []
            for item in tree.get_children(""):
                if col == "#0":
                    # 名称列使用text值
                    value = tree.item(item, "text")
                else:
                    # 其他列使用列值
                    value = tree.set(item, col)
                items.append((value, item))
            
            # 根据列类型进行排序
            if col == "#0":  # 名称列（字符串排序，忽略大小写）
                items.sort(key=lambda t: (not t[0] or t[0].strip() == "", (t[0] or "").lower()), reverse=reverse)
            elif col == "大小":  # 大小列（数字排序）
                def parse_size(size_str):
                    if not size_str or not size_str.strip():
                        return 0
                    try:
                        return int(size_str.strip())
                    except:
                        return 0
                items.sort(key=lambda t: parse_size(t[0]), reverse=reverse)
            else:
                # 其他列（字符串排序）
                items.sort(key=lambda t: (not t[0] or t[0].strip() == "", (t[0] or "").lower()), reverse=reverse)
            
            # 重新排列项目
            for index, (val, item) in enumerate(items):
                tree.move(item, "", index)
            
            # 更新排序方向
            sort_states[col] = not reverse
            # 更新列标题显示排序方向
            current_text = tree.heading(col, "text")
            # 移除已有的箭头符号
            base_text = current_text.replace(" ↑", "").replace(" ↓", "").strip()
            arrow = " ↓" if not reverse else " ↑"  # reverse=False表示升序，显示↓；reverse=True表示降序，显示↑
            tree.heading(col, text=base_text + arrow)
        
        def on_column_click(col):
            """列标题点击事件"""
            if col in sort_states:
                reverse = sort_states[col]
                sort_treeview(col, reverse)
        
        # 设置列（添加点击事件）
        tree.heading("#0", text="名称", command=lambda: on_column_click("#0"))
        tree.heading("类型", text="类型")
        tree.heading("权限", text="权限")
        tree.heading("大小", text="大小", command=lambda: on_column_click("大小"))
        tree.heading("修改时间", text="修改时间")
        
        tree.column("#0", width=300, anchor=tk.W)
        tree.column("类型", width=80, anchor=tk.CENTER)
        tree.column("权限", width=100, anchor=tk.CENTER)
        tree.column("大小", width=100, anchor=tk.E)
        tree.column("修改时间", width=200, anchor=tk.W)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 操作按钮
        action_frame = ttk.Frame(list_frame, padding="5")
        action_frame.pack(fill=tk.X)
        ttk.Button(action_frame, text="打开/编辑", command=lambda: open_file(), width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="下载", command=lambda: download_file(), width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="上传", command=lambda: upload_file(), width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="Notepad++编辑", command=lambda: edit_with_notepad(), width=14).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="重命名", command=lambda: rename_file(), width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="权限", command=lambda: set_permissions(), width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="删除", command=lambda: delete_file(), width=12).pack(side=tk.LEFT, padx=2)
        
        # 存储Notepad++编辑的文件信息 {remote_path: {'local_path': str, 'last_mtime': float, 'last_sync_mtime': float}}
        notepad_edit_files = {}
        
        def debug_output():
            """调试：显示原始输出"""
            path = path_var.get().strip() or "/"
            try:
                # 方法1: 使用ls命令
                cmd = f"ls -la '{path}' 2>&1"
                stdin, stdout, stderr = self.client.exec_command(cmd)
                stdout.channel.settimeout(10)
                import time
                time.sleep(1.0)
                output = stdout.read().decode('utf-8', errors='ignore')
                error = stderr.read().decode('utf-8', errors='ignore')
                
                # 方法2: 尝试使用SFTP
                sftp_output = ""
                try:
                    sftp = self.client.open_sftp()
                    files = sftp.listdir_attr(path)
                    sftp.close()
                    sftp_output = "\n".join([f"{attr.st_mode} {attr.st_size} {attr.st_mtime} {attr.filename}" for attr in files])
                except Exception as sftp_error:
                    sftp_output = f"SFTP错误: {sftp_error}"
                
                debug_window = tk.Toplevel(browser_window)
                debug_window.title("调试信息")
                debug_window.geometry("900x700")
                debug_text = scrolledtext.ScrolledText(debug_window, wrap=tk.WORD, font=("Consolas", 9))
                debug_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                debug_text.insert("1.0", f"路径: {path}\n")
                debug_text.insert(tk.END, f"命令: {cmd}\n\n")
                debug_text.insert(tk.END, f"=== LS命令输出 ===\n")
                debug_text.insert(tk.END, f"输出:\n{output}\n\n")
                debug_text.insert(tk.END, f"错误:\n{error}\n\n")
                debug_text.insert(tk.END, f"输出长度: {len(output)}\n")
                debug_text.insert(tk.END, f"错误长度: {len(error)}\n")
                debug_text.insert(tk.END, f"总行数: {len((output + error).split(chr(10)))}\n\n")
                debug_text.insert(tk.END, f"=== SFTP方法 ===\n")
                debug_text.insert(tk.END, f"{sftp_output}\n")
            except Exception as e:
                import traceback
                messagebox.showerror("错误", f"调试失败: {e}\n\n{traceback.format_exc()}")
        
        ttk.Button(action_frame, text="调试", command=debug_output, width=12).pack(side=tk.LEFT, padx=2)
        
        # 标签页2: 文件编辑器
        edit_frame = ttk.Frame(notebook, padding="10")
        notebook.add(edit_frame, text="文件编辑")
        
        # 编辑器工具栏
        edit_toolbar = ttk.Frame(edit_frame)
        edit_toolbar.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(edit_toolbar, text="文件路径:").pack(side=tk.LEFT, padx=5)
        file_path_label = ttk.Label(edit_toolbar, text="未打开文件", foreground="gray")
        file_path_label.pack(side=tk.LEFT, padx=5)
        
        edit_btn_frame = ttk.Frame(edit_toolbar)
        edit_btn_frame.pack(side=tk.RIGHT, padx=5)
        ttk.Button(edit_btn_frame, text="保存", command=lambda: save_file(), width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(edit_btn_frame, text="查找替换", command=lambda: find_replace(), width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(edit_btn_frame, text="重新加载", command=lambda: reload_file(), width=10).pack(side=tk.LEFT, padx=2)
        
        # 文件内容编辑器
        file_content_text = scrolledtext.ScrolledText(edit_frame, wrap=tk.WORD, font=("Consolas", 10), height=30)
        file_content_text.pack(fill=tk.BOTH, expand=True)
        
        def browse_path():
            # 浏览指定路径
            path = path_var.get().strip()
            if not path:
                path = "/"
            
            # 清理路径
            path = path.rstrip('/')
            if not path:
                path = "/"
            
            # 清空树
            for item in tree.get_children():
                tree.delete(item)
            
            # 清除排序箭头（重置列标题）
            for col in sort_states.keys():
                current_text = tree.heading(col, "text")
                base_text = current_text.replace(" ↑", "").replace(" ↓", "").strip()
                tree.heading(col, text=base_text)
                sort_states[col] = True  # 重置为升序
            
            # 显示加载中
            loading_item = tree.insert("", tk.END, text="(加载中...)", values=("", "", "", ""), tags=("loading",))
            tree.tag_configure("loading", foreground="gray")
            browser_window.update_idletasks()
            
            try:
                import time
                from datetime import datetime
                
                # 测试连接
                if not self.client or not self.is_connected:
                    tree.delete(loading_item)
                    tree.insert("", tk.END, text="(错误: SSH未连接)", values=("", "", "", ""), tags=("error",))
                    tree.tag_configure("error", foreground="red")
                    return
                
                # 优先使用SFTP方法（更可靠）
                try:
                    sftp = self.client.open_sftp()
                    files = sftp.listdir_attr(path)
                    sftp.close()
                    
                    # 清空加载提示
                    tree.delete(loading_item)
                    
                    # 解析SFTP结果
                    file_count = 0
                    for attr in files:
                        name = attr.filename
                        if name in ['.', '..']:
                            continue
                        
                        # 判断类型
                        if stat.S_ISDIR(attr.st_mode):
                            file_type = "目录"
                        elif stat.S_ISLNK(attr.st_mode):
                            file_type = "链接"
                        else:
                            file_type = "文件"
                        
                        # 权限
                        permissions = stat.filemode(attr.st_mode)
                        
                        # 大小
                        size = str(attr.st_size) if hasattr(attr, 'st_size') else ""
                        
                        # 修改时间
                        try:
                            mtime = datetime.fromtimestamp(attr.st_mtime)
                            date = mtime.strftime("%Y-%m-%d %H:%M")
                        except:
                            date = ""
                        
                        # 插入到树中
                        tree.insert("", tk.END, text=name, values=(file_type, permissions, size, date), 
                                  tags=(file_type,))
                        file_count += 1
                    
                    # 设置标签颜色
                    tree.tag_configure("目录", foreground="blue")
                    tree.tag_configure("链接", foreground="green")
                    
                    # 更新路径显示
                    path_var.set(path)
                    
                    # 如果没有文件，显示提示
                    if file_count == 0:
                        tree.insert("", tk.END, text="(空目录)", values=("", "", "", ""), tags=("empty",))
                        tree.tag_configure("empty", foreground="gray")
                    
                    browser_window.update_idletasks()
                    return
                    
                except Exception as sftp_error:
                    # SFTP失败，回退到ls命令
                    pass
                
                # 使用ls命令作为备用方法
                cmd = f"ls -la '{path}'"
                stdin, stdout, stderr = self.client.exec_command(cmd, timeout=10)
                
                # 等待命令完成
                time.sleep(1.0)
                
                # 读取输出
                output = stdout.read().decode('utf-8', errors='ignore')
                error = stderr.read().decode('utf-8', errors='ignore')
                
                # 清空加载提示
                try:
                    tree.delete(loading_item)
                except:
                    pass
                
                # 检查错误
                if error and ("No such file" in error or "cannot access" in error or "Permission denied" in error):
                    tree.insert("", tk.END, text=f"(错误: {error.strip()})", values=("", "", "", ""), tags=("error",))
                    tree.tag_configure("error", foreground="red")
                    browser_window.update_idletasks()
                    return
                
                # 如果没有输出
                if not output.strip():
                    tree.insert("", tk.END, text="(空目录)", values=("", "", "", ""), tags=("empty",))
                    tree.tag_configure("empty", foreground="gray")
                    path_var.set(path)
                    browser_window.update_idletasks()
                    return
                
                # 解析输出
                lines = output.strip().split('\n')
                file_count = 0
                parse_errors = []
                
                for line in lines:
                    line = line.strip()
                    # 跳过空行和total行
                    if not line or line.startswith('total'):
                        continue
                    
                    # 检查是否是ls输出格式（必须以d、-或l开头）
                    if len(line) < 10 or line[0] not in ['d', '-', 'l']:
                        continue
                    
                    try:
                        # 使用split方法解析，更可靠
                        parts = line.split(None, 8)  # 最多分割8次，保留文件名部分
                        
                        if len(parts) < 9:
                            # 如果分割后少于9部分，尝试更宽松的解析
                            parts = line.split(None, 7)
                            if len(parts) < 8:
                                parse_errors.append(f"无法解析: {line[:50]}")
                                continue
                            # 简化格式：只有8部分
                            permissions_str = parts[0]
                            file_type_char = permissions_str[0] if permissions_str else '-'
                            permissions = permissions_str
                            size = parts[4] if len(parts) > 4 else ""
                            date = " ".join(parts[5:7]) if len(parts) > 6 else ""
                            name = parts[7] if len(parts) > 7 else ""
                        else:
                            # 标准格式：9部分
                            permissions_str = parts[0]
                            file_type_char = permissions_str[0] if permissions_str else '-'
                            permissions = permissions_str
                            size = parts[4]
                            # 日期可能是3部分（月 日 时间）或4部分（月 日 年）
                            if len(parts) > 7:
                                date = " ".join(parts[5:8])
                            else:
                                date = " ".join(parts[5:7]) if len(parts) > 6 else ""
                            name = parts[8]
                        
                        # 处理符号链接
                        if ' -> ' in name:
                            name = name.split(' -> ')[0].strip()
                        
                        # 跳过当前目录和父目录
                        if name in ['.', '..']:
                            continue
                        
                        # 判断文件类型
                        if file_type_char == 'd':
                            file_type = "目录"
                        elif file_type_char == 'l':
                            file_type = "链接"
                        else:
                            file_type = "文件"
                        
                        # 插入到树中
                        tree.insert("", tk.END, text=name, values=(file_type, permissions, size, date), 
                                  tags=(file_type,))
                        file_count += 1
                    except Exception as parse_error:
                        parse_errors.append(f"解析错误: {str(parse_error)} - {line[:50]}")
                        continue
                
                # 设置标签颜色
                tree.tag_configure("目录", foreground="blue")
                tree.tag_configure("链接", foreground="green")
                
                # 更新路径显示
                path_var.set(path)
                
                # 如果没有找到文件，显示提示
                if file_count == 0:
                    if parse_errors:
                        # 如果有解析错误，显示错误信息
                        error_msg = f"(解析错误: {len(parse_errors)} 行无法解析)"
                        tree.insert("", tk.END, text=error_msg, values=("", "", "", ""), tags=("error",))
                        tree.tag_configure("error", foreground="red")
                    else:
                        tree.insert("", tk.END, text="(空目录)", values=("", "", "", ""), tags=("empty",))
                        tree.tag_configure("empty", foreground="gray")
                elif parse_errors and file_count > 0:
                    # 如果部分解析成功但有错误，在最后显示警告
                    tree.insert("", tk.END, text=f"(警告: {len(parse_errors)} 行解析失败)", 
                              values=("", "", "", ""), tags=("warning",))
                    tree.tag_configure("warning", foreground="orange")
                
                # 强制更新界面
                browser_window.update_idletasks()
                
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                # 清空树并显示错误
                try:
                    tree.delete(loading_item)
                except:
                    pass
                for item in tree.get_children():
                    tree.delete(item)
                tree.insert("", tk.END, text=f"(错误: {str(e)})", values=("", "", "", ""), tags=("error",))
                tree.tag_configure("error", foreground="red")
                browser_window.update_idletasks()
                # 显示详细错误
                messagebox.showerror("错误", f"浏览失败: {e}\n\n路径: {path}\n\n详细信息:\n{error_detail[:300]}")
        
        def go_up():
            # Go to parent directory
            current = path_var.get().strip().rstrip('/')
            if current == "/":
                return
            parent = "/".join(current.split("/")[:-1])
            if not parent:
                parent = "/"
            path_var.set(parent)
            browse_path()
        
        def go_root():
            # Go to root directory
            path_var.set("/")
            browse_path()
        
        def get_selected_path():
            # Get currently selected file or directory path
            item = tree.selection()[0] if tree.selection() else None
            if not item:
                return None
            
            name = tree.item(item, "text")
            current = path_var.get().strip().rstrip('/')
            if current == "/":
                return f"/{name}"
            else:
                return f"{current}/{name}"
        
        def open_file():
            """打开文件进行编辑"""
            nonlocal current_file_path
            file_path = get_selected_path()
            if not file_path:
                # 如果没有选择文件，询问是否创建新文件
                new_path = simpledialog.askstring("创建新文件", "请输入文件路径:")
                if not new_path:
                    return
                file_path = new_path
                # 检查文件是否存在
                try:
                    sftp = self.client.open_sftp()
                    try:
                        sftp.stat(file_path)
                        # 文件存在，读取它
                        with sftp.open(file_path, 'r') as f:
                            content = f.read().decode('utf-8', errors='ignore')
                        sftp.close()
                        file_content_text.delete("1.0", tk.END)
                        file_content_text.insert("1.0", content)
                        file_path_label.config(text=file_path, foreground="black")
                        current_file_path = file_path
                        notebook.select(1)
                        return
                    except IOError:
                        # 文件不存在，创建新文件
                        sftp.close()
                        if messagebox.askyesno("确认", f"文件不存在: {file_path}\n是否创建新文件？"):
                            file_content_text.delete("1.0", tk.END)
                            file_path_label.config(text=file_path, foreground="blue")
                            current_file_path = file_path
                            notebook.select(1)
                        return
                except Exception as e:
                    messagebox.showerror("错误", f"检查文件失败: {e}")
                    return
            
            values = tree.item(tree.selection()[0], "values")
            if values and values[0] == "目录":
                messagebox.showwarning("提示", "请选择文件，不是目录")
                return
            
            try:
                # 使用SFTP读取文件（更可靠）
                try:
                    sftp = self.client.open_sftp()
                    with sftp.open(file_path, 'r') as f:
                        content = f.read().decode('utf-8', errors='ignore')
                    sftp.close()
                    
                    # 显示文件内容
                    file_content_text.delete("1.0", tk.END)
                    file_content_text.insert("1.0", content)
                    file_path_label.config(text=file_path, foreground="black")
                    current_file_path = file_path
                    notebook.select(1)  # 切换到编辑标签页
                except (IOError, FileNotFoundError) as io_error:
                    # 文件不存在
                    if messagebox.askyesno("文件不存在", f"文件不存在: {file_path}\n\n是否创建新文件？\n（点击'是'将创建空文件供编辑）"):
                        file_content_text.delete("1.0", tk.END)
                        file_path_label.config(text=file_path + " (新文件)", foreground="blue")
                        current_file_path = file_path
                        notebook.select(1)
                    return
                except Exception as sftp_error:
                    # SFTP失败，尝试使用cat命令
                    cmd = f"cat '{file_path}' 2>&1"
                    stdin, stdout, stderr = self.client.exec_command(cmd)
                    stdout.channel.settimeout(10)
                    import time
                    time.sleep(0.5)
                    content = stdout.read().decode('utf-8', errors='ignore')
                    error = stderr.read().decode('utf-8', errors='ignore')
                    
                    if error and ("No such file" in error or "cannot access" in error):
                        if messagebox.askyesno("文件不存在", f"文件不存在: {file_path}\n\n是否创建新文件？\n（点击'是'将创建空文件供编辑）"):
                            file_content_text.delete("1.0", tk.END)
                            file_path_label.config(text=file_path + " (新文件)", foreground="blue")
                            current_file_path = file_path
                            notebook.select(1)
                        return
                    
                    # 显示文件内容
                    file_content_text.delete("1.0", tk.END)
                    file_content_text.insert("1.0", content)
                    file_path_label.config(text=file_path, foreground="black")
                    current_file_path = file_path
                    notebook.select(1)
                
            except Exception as e:
                messagebox.showerror("错误", f"打开文件失败: {e}")
        
        def save_file():
            """保存文件"""
            nonlocal current_file_path
            if not current_file_path:
                messagebox.showwarning("提示", "没有打开的文件", parent=browser_window)
                return
            
            content = file_content_text.get("1.0", tk.END + "-1c")  # 获取内容，去掉最后的换行
            
            try:
                # 使用SFTP保存文件（最可靠的方法）
                import tempfile
                with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', delete=False) as tmp_file:
                    tmp_file.write(content)
                    tmp_path = tmp_file.name
                
                try:
                    sftp = self.client.open_sftp()
                    sftp.put(tmp_path, current_file_path)
                    sftp.close()
                    messagebox.showinfo("成功", "文件已保存", parent=browser_window)
                    # 保存后刷新文件列表，但不关闭浏览器
                    browse_path()
                finally:
                    # 删除临时文件
                    try:
                        os.unlink(tmp_path)
                    except:
                        pass
                    
            except Exception as e:
                messagebox.showerror("错误", f"保存文件失败: {e}", parent=browser_window)
        
        def reload_file():
            """重新加载文件"""
            nonlocal current_file_path
            if current_file_path:
                # 临时保存路径
                saved_path = current_file_path
                # 重新打开文件
                file_path = saved_path
                try:
                    cmd = f"cat '{file_path}' 2>&1"
                    stdin, stdout, stderr = self.client.exec_command(cmd)
                    stdout.channel.settimeout(10)
                    content = stdout.read().decode('utf-8', errors='ignore')
                    error = stderr.read().decode('utf-8', errors='ignore')
                    
                    if error and "No such file" in error:
                        messagebox.showerror("错误", f"文件不存在: {file_path}")
                        return
                    
                    file_content_text.delete("1.0", tk.END)
                    file_content_text.insert("1.0", content)
                    messagebox.showinfo("成功", "文件已重新加载")
                except Exception as e:
                    messagebox.showerror("错误", f"重新加载失败: {e}")
        def find_replace():
            """查找替换对话框"""
            if not current_file_path and file_content_text.get("1.0", tk.END).strip() == "":
                messagebox.showwarning("提示", "请先打开一个文件")
                return
            
            find_window = tk.Toplevel(browser_window)
            find_window.title("查找替换")
            find_window.geometry("450x200")
            find_window.transient(browser_window)
            find_window.grab_set()
            
            ttk.Label(find_window, text="查找:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            find_var = tk.StringVar()
            find_entry = ttk.Entry(find_window, textvariable=find_var, width=35)
            find_entry.grid(row=0, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
            find_entry.focus()
            find_entry.bind('<Return>', lambda e: do_find())
            
            ttk.Label(find_window, text="替换为:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            replace_var = tk.StringVar()
            replace_entry = ttk.Entry(find_window, textvariable=replace_var, width=35)
            replace_entry.grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
            replace_entry.bind('<Return>', lambda e: do_replace_all())
            
            find_window.columnconfigure(1, weight=1)
            
            # 状态标签（先创建，供函数使用）
            status_label = ttk.Label(find_window, text="", foreground="gray")
            status_label.grid(row=3, column=0, columnspan=2, pady=5)
            
            current_search_pos = "1.0"
            
            def do_find():
                """查找文本"""
                nonlocal current_search_pos
                find_text = find_var.get()
                if not find_text:
                    messagebox.showwarning("提示", "请输入要查找的内容")
                    return
                
                # 从当前位置开始查找
                pos = file_content_text.search(find_text, current_search_pos, tk.END)
                if pos:
                    # 找到文本，高亮显示
                    end_pos = f"{pos}+{len(find_text)}c"
                    file_content_text.tag_remove("search", "1.0", tk.END)
                    file_content_text.tag_add("search", pos, end_pos)
                    file_content_text.tag_config("search", background="yellow", foreground="black")
                    file_content_text.see(pos)
                    # 更新查找位置
                    current_search_pos = end_pos
                    status_label.config(text=f"找到: {find_text}", foreground="green")
                else:
                    # 没找到，从头开始
                    pos = file_content_text.search(find_text, "1.0", tk.END)
                    if pos:
                        end_pos = f"{pos}+{len(find_text)}c"
                        file_content_text.tag_remove("search", "1.0", tk.END)
                        file_content_text.tag_add("search", pos, end_pos)
                        file_content_text.tag_config("search", background="yellow", foreground="black")
                        file_content_text.see(pos)
                        current_search_pos = end_pos
                        status_label.config(text=f"找到: {find_text} (已循环)", foreground="green")
                    else:
                        status_label.config(text=f"未找到: {find_text}", foreground="red")
                        messagebox.showinfo("提示", f"未找到: {find_text}")
                        current_search_pos = "1.0"
            
            def do_replace():
                """替换当前找到的文本"""
                nonlocal current_search_pos
                find_text = find_var.get()
                replace_text = replace_var.get()
                if not find_text:
                    messagebox.showwarning("提示", "请输入要查找的内容")
                    return
                
                # 检查是否有选中的搜索文本
                try:
                    sel_start = file_content_text.index("search.first")
                    sel_end = file_content_text.index("search.last")
                    selected = file_content_text.get(sel_start, sel_end)
                    if selected == find_text:
                        # 替换选中的文本
                        file_content_text.delete(sel_start, sel_end)
                        file_content_text.insert(sel_start, replace_text)
                        file_content_text.tag_remove("search", "1.0", tk.END)
                        current_search_pos = sel_start
                        status_label.config(text="已替换", foreground="green")
                        # 继续查找下一个
                        do_find()
                    else:
                        # 先查找
                        do_find()
                except:
                    # 没有选中，先查找
                    do_find()
            
            def do_replace_all():
                """替换所有"""
                find_text = find_var.get()
                replace_text = replace_var.get()
                if not find_text:
                    messagebox.showwarning("提示", "请输入要查找的内容")
                    return
                
                content = file_content_text.get("1.0", tk.END)
                count = content.count(find_text)
                if count == 0:
                    messagebox.showinfo("提示", f"未找到: {find_text}")
                    return
                
                if messagebox.askyesno("确认", f"找到 {count} 处匹配，是否全部替换？"):
                    new_content = content.replace(find_text, replace_text)
                    file_content_text.delete("1.0", tk.END)
                    file_content_text.insert("1.0", new_content)
                    file_content_text.tag_remove("search", "1.0", tk.END)
                    status_label.config(text=f"已替换 {count} 处", foreground="green")
                    messagebox.showinfo("完成", f"已替换 {count} 处")
            
            # 按钮框架
            btn_frame = ttk.Frame(find_window)
            btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
            
            ttk.Button(btn_frame, text="查找", command=do_find, width=12).grid(row=0, column=0, padx=3)
            ttk.Button(btn_frame, text="替换", command=do_replace, width=12).grid(row=0, column=1, padx=3)
            ttk.Button(btn_frame, text="全部替换", command=do_replace_all, width=12).grid(row=0, column=2, padx=3)
            ttk.Button(btn_frame, text="关闭", command=find_window.destroy, width=12).grid(row=0, column=3, padx=3)
        
        def download_file():
            """下载文件到本地"""
            file_path = get_selected_path()
            if not file_path:
                messagebox.showwarning("提示", "请先选择一个文件")
                return
            
            values = tree.item(tree.selection()[0], "values")
            if values and values[0] == "目录":
                messagebox.showwarning("提示", "请选择文件，不是目录")
                return
            
            # 选择保存位置
            local_path = filedialog.asksaveasfilename(
                title="保存文件",
                initialfile=os.path.basename(file_path),
                defaultextension=""
            )
            if not local_path:
                return
            
            try:
                # 使用SFTP下载文件
                sftp = self.client.open_sftp()
                sftp.get(file_path, local_path)
                sftp.close()
                messagebox.showinfo("成功", f"文件已下载到: {local_path}")
                # 下载后不关闭窗口
            except Exception as e:
                messagebox.showerror("错误", f"下载文件失败: {e}")
                # 错误后也不关闭窗口
        
        def upload_file():
            """上传本地文件到服务器"""
            # 选择本地文件
            local_path = filedialog.askopenfilename(title="选择要上传的文件")
            if not local_path:
                return
            
            file_path = get_selected_path()
            if file_path:
                values = tree.item(tree.selection()[0], "values")
                if values and values[0] == "目录":
                    # 如果选中的是目录，上传到该目录
                    remote_path = f"{file_path}/{os.path.basename(local_path)}"
                else:
                    # 如果选中的是文件，询问是否覆盖
                    if not messagebox.askyesno("确认", f"是否覆盖文件: {file_path}?"):
                        return
                    remote_path = file_path
            else:
                # 没有选中文件，上传到当前目录
                current = path_var.get().strip().rstrip('/')
                remote_path = f"{current}/{os.path.basename(local_path)}"
            
            try:
                # 使用SFTP上传文件
                sftp = self.client.open_sftp()
                sftp.put(local_path, remote_path)
                sftp.close()
                messagebox.showinfo("成功", f"文件已上传到: {remote_path}")
                browse_path()  # 刷新文件列表，不关闭窗口
            except Exception as e:
                messagebox.showerror("错误", f"上传文件失败: {e}")
                # 错误后也不关闭窗口
        
        def edit_with_notepad():
            """用Notepad++打开文件编辑，保存后自动同步到服务器"""
            file_path = get_selected_path()
            if not file_path:
                messagebox.showwarning("提示", "请先选择一个文件")
                return
            
            values = tree.item(tree.selection()[0], "values")
            if values and values[0] == "目录":
                messagebox.showwarning("提示", "请选择文件，不是目录")
                return
            
            try:
                # 检查Notepad++是否安装
                notepad_paths = [
                    r"C:\Program Files\Notepad++\notepad++.exe",
                    r"C:\Program Files (x86)\Notepad++\notepad++.exe",
                    os.path.expanduser(r"~\AppData\Local\Programs\Notepad++\notepad++.exe"),
                ]
                
                notepad_exe = None
                for path in notepad_paths:
                    if os.path.exists(path):
                        notepad_exe = path
                        break
                
                if not notepad_exe:
                    # 尝试从PATH中查找
                    try:
                        result = subprocess.run(["where", "notepad++"], capture_output=True, text=True, timeout=5)
                        if result.returncode == 0 and result.stdout.strip():
                            notepad_exe = result.stdout.strip().split('\n')[0]
                    except:
                        pass
                
                if not notepad_exe:
                    messagebox.showerror("错误", "未找到Notepad++\n\n请确保已安装Notepad++，或手动指定Notepad++的安装路径。")
                    return
                
                # 创建临时目录（如果不存在）
                temp_dir = os.path.join(tempfile.gettempdir(), "ssh_tool_notepad_edit")
                os.makedirs(temp_dir, exist_ok=True)
                
                # 生成临时文件路径（使用文件名+时间戳避免冲突）
                file_name = os.path.basename(file_path)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                temp_file_path = os.path.join(temp_dir, f"{file_name}_{timestamp}")
                
                # 下载文件到临时目录
                try:
                    sftp = self.client.open_sftp()
                    sftp.get(file_path, temp_file_path)
                    sftp.close()
                except Exception as e:
                    messagebox.showerror("错误", f"下载文件失败: {e}")
                    return
                
                # 保存文件映射关系和初始修改时间
                initial_mtime = os.path.getmtime(temp_file_path) if os.path.exists(temp_file_path) else 0
                notepad_edit_files[file_path] = {
                    'local_path': temp_file_path,
                    'last_mtime': initial_mtime,
                    'last_sync_mtime': initial_mtime
                }
                
                # 用Notepad++打开文件
                try:
                    # 使用subprocess启动Notepad++，不等待其关闭
                    subprocess.Popen([notepad_exe, temp_file_path], shell=False)
                    
                    # 启动自动同步监控线程
                    def auto_sync_monitor():
                        """自动监控文件变化并同步"""
                        while file_path in notepad_edit_files:
                            try:
                                local_file = notepad_edit_files[file_path]['local_path']
                                
                                # 检查文件是否存在
                                if not os.path.exists(local_file):
                                    # 文件被删除，停止监控
                                    if file_path in notepad_edit_files:
                                        del notepad_edit_files[file_path]
                                    break
                                
                                # 检查文件修改时间
                                current_mtime = os.path.getmtime(local_file)
                                last_sync_mtime = notepad_edit_files[file_path]['last_sync_mtime']
                                
                                # 如果文件被修改了（修改时间变化）
                                if current_mtime > last_sync_mtime + 1:  # 加1秒避免频繁同步
                                    # 等待一小段时间，确保文件写入完成
                                    import time
                                    time.sleep(0.5)
                                    
                                    # 再次检查修改时间，确保文件已保存完成
                                    final_mtime = os.path.getmtime(local_file)
                                    if final_mtime == current_mtime:
                                        # 文件已保存，开始同步
                                        try:
                                            sftp = self.client.open_sftp()
                                            sftp.put(local_file, file_path)
                                            sftp.close()
                                            
                                            # 更新同步时间
                                            notepad_edit_files[file_path]['last_sync_mtime'] = final_mtime
                                            
                                            # 在输出面板显示同步信息（不弹窗）
                                            self.output_queue.put(("success", f"[自动同步] {file_path} 已同步到服务器\n"))
                                            
                                            # 刷新文件列表（在UI线程中执行）
                                            self.root.after(0, browse_path)
                                        except Exception as sync_error:
                                            # 同步失败，在输出面板显示错误
                                            self.output_queue.put(("error", f"[自动同步失败] {file_path}: {sync_error}\n"))
                                
                                # 更新最后检查的修改时间
                                notepad_edit_files[file_path]['last_mtime'] = current_mtime
                                
                                # 等待1秒后再次检查
                                import time
                                time.sleep(1)
                                
                            except Exception as e:
                                # 监控出错，继续监控
                                import time
                                time.sleep(2)
                                continue
                    
                    # 启动监控线程
                    monitor_thread = threading.Thread(target=auto_sync_monitor, daemon=True)
                    monitor_thread.start()
                    
                    # 显示成功提示（不弹窗，只在输出面板显示）
                    self.output_queue.put(("info", f"[Notepad++编辑] 已打开文件: {file_path}\n文件保存后将自动同步到服务器\n"))
                    
                except Exception as e:
                    messagebox.showerror("错误", f"打开Notepad++失败: {e}")
                    # 清理临时文件
                    if os.path.exists(temp_file_path):
                        try:
                            os.remove(temp_file_path)
                        except:
                            pass
                    if file_path in notepad_edit_files:
                        del notepad_edit_files[file_path]
                    
            except Exception as e:
                messagebox.showerror("错误", f"操作失败: {e}")
        
        def delete_file():
            """删除文件或目录"""
            file_path = get_selected_path()
            if not file_path:
                messagebox.showwarning("提示", "请先选择要删除的文件或目录")
                return
            
            values = tree.item(tree.selection()[0], "values")
            is_dir = values and values[0] == "目录"
            
            if not messagebox.askyesno("确认", f"确定要删除{'目录' if is_dir else '文件'}: {file_path}?"):
                return
            
            try:
                # 使用 get_pty=False 避免 chdir 到主目录的错误
                # 使用绝对路径，避免路径问题
                if is_dir:
                    cmd = f"rm -rf '{file_path}'"
                else:
                    cmd = f"rm -f '{file_path}'"
                
                stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                stdout.channel.settimeout(10)
                stderr.channel.settimeout(10)
                
                # 等待命令执行完成
                exit_status = stdout.channel.recv_exit_status()
                error = stderr.read().decode('utf-8', errors='ignore').strip()
                
                if exit_status != 0 or error:
                    # 检查是否是权限错误
                    if "Operation not permitted" in error or "Permission denied" in error:
                        # 构建 sudo 命令提示
                        sudo_cmd = f"sudo rm -{'rf' if is_dir else 'f'} '{file_path}'"
                        error_msg = f"删除失败: 权限不足\n\n文件路径: {file_path}\n\n可能的原因:\n1. 文件受保护或为系统文件\n2. 当前用户没有删除权限\n3. 需要管理员权限\n\n解决方法:\n您可以在下方的命令执行面板中使用以下命令删除:\n\n{sudo_cmd}\n\n或者检查文件权限:\nls -l {file_path}"
                        messagebox.showerror("权限错误", error_msg)
                        
                        # 可选：在输出面板中显示 sudo 命令提示
                        self.output_queue.put(("warning", f"\n[文件删除权限不足]\n文件路径: {file_path}\n\n可以使用以下命令删除:\n{sudo_cmd}\n\n"))
                    elif "No such file or directory" in error:
                        # 文件已不存在，可能已被删除
                        messagebox.showinfo("提示", "文件已不存在，可能已被删除")
                        browse_path()  # 刷新文件列表
                    else:
                        # 其他错误
                        error_msg = f"删除失败\n\n文件路径: {file_path}\n\n错误信息: {error if error else '未知错误'}"
                        messagebox.showerror("删除失败", error_msg)
                else:
                    messagebox.showinfo("成功", "删除成功")
                    browse_path()  # 刷新文件列表，不关闭窗口
            except Exception as e:
                error_msg = f"删除失败: {str(e)}\n\n文件路径: {file_path}\n\n如果问题持续，请检查:\n1. 网络连接是否正常\n2. 文件路径是否正确\n3. 服务器SSH服务是否正常"
                messagebox.showerror("错误", error_msg)
                # 错误后也不关闭窗口
        
        def rename_file():
            """重命名文件或目录"""
            file_path = get_selected_path()
            if not file_path:
                messagebox.showwarning("提示", "请先选择要重命名的文件或目录", parent=browser_window)
                return
            
            # 获取当前文件名
            current_name = tree.item(tree.selection()[0], "text")
            
            # 弹出输入框，让用户输入新名称
            new_name = simpledialog.askstring("重命名", f"请输入新名称:\n\n当前名称: {current_name}", initialvalue=current_name, parent=browser_window)
            if not new_name or new_name.strip() == "":
                return
            
            new_name = new_name.strip()
            
            # 如果名称相同，不需要重命名
            if new_name == current_name:
                return
            
            # 构建新路径
            current_dir = path_var.get().strip().rstrip('/')
            if current_dir == "/":
                new_path = f"/{new_name}"
            else:
                new_path = f"{current_dir}/{new_name}"
            
            try:
                # 使用mv命令重命名
                cmd = f"mv '{file_path}' '{new_path}'"
                stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                stdout.channel.settimeout(10)
                stderr.channel.settimeout(10)
                
                # 等待命令执行完成
                exit_status = stdout.channel.recv_exit_status()
                error = stderr.read().decode('utf-8', errors='ignore').strip()
                
                if exit_status != 0 or error:
                    # 检查是否是权限错误
                    if "Operation not permitted" in error or "Permission denied" in error:
                        sudo_cmd = f"sudo mv '{file_path}' '{new_path}'"
                        error_msg = f"重命名失败: 权限不足\n\n原路径: {file_path}\n新路径: {new_path}\n\n可能的原因:\n1. 文件受保护或为系统文件\n2. 当前用户没有重命名权限\n3. 需要管理员权限\n\n解决方法:\n您可以在下方的命令执行面板中使用以下命令重命名:\n\n{sudo_cmd}"
                        messagebox.showerror("权限错误", error_msg, parent=browser_window)
                        self.output_queue.put(("warning", f"\n[文件重命名权限不足]\n原路径: {file_path}\n新路径: {new_path}\n\n可以使用以下命令重命名:\n{sudo_cmd}\n\n"))
                    elif "File exists" in error or "already exists" in error:
                        messagebox.showerror("错误", f"重命名失败: 目标文件已存在\n\n新路径: {new_path}", parent=browser_window)
                    else:
                        error_msg = f"重命名失败\n\n原路径: {file_path}\n新路径: {new_path}\n\n错误信息: {error if error else '未知错误'}"
                        messagebox.showerror("重命名失败", error_msg, parent=browser_window)
                else:
                    messagebox.showinfo("成功", f"重命名成功\n\n原名称: {current_name}\n新名称: {new_name}", parent=browser_window)
                    browse_path()  # 刷新文件列表，不关闭窗口
            except Exception as e:
                messagebox.showerror("错误", f"重命名失败: {str(e)}", parent=browser_window)
                # 错误后也不关闭窗口
        
        def set_permissions():
            """设置文件或目录权限"""
            file_path = get_selected_path()
            if not file_path:
                messagebox.showwarning("提示", "请先选择要设置权限的文件或目录")
                return
            
            # 获取当前权限
            values = tree.item(tree.selection()[0], "values")
            current_permissions = values[1] if len(values) > 1 else ""
            
            # 创建权限设置窗口
            perm_window = tk.Toplevel(browser_window)
            perm_window.title("设置权限")
            perm_window.geometry("500x450")
            perm_window.transient(browser_window)
            perm_window.grab_set()
            
            # 文件路径显示
            path_frame = ttk.Frame(perm_window, padding="10")
            path_frame.pack(fill=tk.X)
            ttk.Label(path_frame, text="文件路径:", font=("Microsoft YaHei", 9, "bold")).pack(anchor=tk.W)
            ttk.Label(path_frame, text=file_path, font=("Consolas", 9), foreground="blue").pack(anchor=tk.W, pady=(5, 10))
            
            # 当前权限显示
            if current_permissions:
                ttk.Label(path_frame, text=f"当前权限: {current_permissions}", font=("Microsoft YaHei", 9)).pack(anchor=tk.W)
            
            # 权限输入方式选择
            mode_frame = ttk.LabelFrame(perm_window, text="权限设置方式", padding="10")
            mode_frame.pack(fill=tk.X, padx=10, pady=5)
            
            mode_var = tk.StringVar(value="octal")
            
            def update_mode():
                if mode_var.get() == "octal":
                    octal_frame.pack(fill=tk.X, padx=5, pady=5)
                    symbol_frame.pack_forget()
                else:
                    octal_frame.pack_forget()
                    symbol_frame.pack(fill=tk.X, padx=5, pady=5)
            
            ttk.Radiobutton(mode_frame, text="八进制模式 (如: 755, 644)", variable=mode_var, value="octal", command=update_mode).pack(anchor=tk.W)
            ttk.Radiobutton(mode_frame, text="符号模式 (如: u+rwx, g+rx, o+r)", variable=mode_var, value="symbol", command=update_mode).pack(anchor=tk.W)
            
            # 八进制模式输入
            octal_frame = ttk.Frame(mode_frame)
            ttk.Label(octal_frame, text="权限值 (3位数字，如 755):").pack(anchor=tk.W)
            octal_var = tk.StringVar(value="755")
            octal_entry = ttk.Entry(octal_frame, textvariable=octal_var, width=20, font=("Consolas", 10))
            octal_entry.pack(anchor=tk.W, pady=5)
            ttk.Label(octal_frame, text="说明: 第一位(所有者) 第二位(组) 第三位(其他)\n例如: 755 = rwxr-xr-x, 644 = rw-r--r--", 
                     font=("Microsoft YaHei", 8), foreground="gray").pack(anchor=tk.W)
            
            # 符号模式输入
            symbol_frame = ttk.Frame(mode_frame)
            ttk.Label(symbol_frame, text="权限表达式:").pack(anchor=tk.W)
            symbol_var = tk.StringVar(value="u+rwx,g+rx,o+r")
            symbol_entry = ttk.Entry(symbol_frame, textvariable=symbol_var, width=40, font=("Consolas", 10))
            symbol_entry.pack(anchor=tk.W, pady=5)
            ttk.Label(symbol_frame, text="说明: u=所有者, g=组, o=其他, a=全部\n+添加权限, -移除权限, =设置权限\n例如: u+rwx,g+rx,o+r 或 a+x", 
                     font=("Microsoft YaHei", 8), foreground="gray").pack(anchor=tk.W)
            
            # 常用权限快捷按钮
            quick_frame = ttk.LabelFrame(perm_window, text="常用权限", padding="10")
            quick_frame.pack(fill=tk.X, padx=10, pady=5)
            
            quick_perms = [
                ("755 (rwxr-xr-x)", "755", "目录常用"),
                ("644 (rw-r--r--)", "644", "文件常用"),
                ("777 (rwxrwxrwx)", "777", "全部权限"),
                ("600 (rw-------)", "600", "仅所有者"),
            ]
            
            def set_quick_perm(perm_value):
                mode_var.set("octal")
                octal_var.set(perm_value)
                update_mode()
            
            for i, (label, value, desc) in enumerate(quick_perms):
                btn_frame = ttk.Frame(quick_frame)
                btn_frame.grid(row=i//2, column=i%2, sticky=(tk.W, tk.E), padx=5, pady=2)
                ttk.Button(btn_frame, text=label, command=lambda v=value: set_quick_perm(v), width=20).pack(side=tk.LEFT)
                ttk.Label(btn_frame, text=desc, font=("Microsoft YaHei", 8), foreground="gray").pack(side=tk.LEFT, padx=5)
            
            # 按钮
            btn_frame = ttk.Frame(perm_window, padding="10")
            btn_frame.pack(fill=tk.X)
            
            def apply_permissions():
                try:
                    if mode_var.get() == "octal":
                        perm_value = octal_var.get().strip()
                        # 验证八进制格式
                        if not perm_value.isdigit() or len(perm_value) != 3:
                            messagebox.showerror("错误", "权限值必须是3位数字 (如: 755, 644)")
                            return
                        # 验证每位数字在0-7之间
                        if not all('0' <= c <= '7' for c in perm_value):
                            messagebox.showerror("错误", "权限值每位数字必须在0-7之间")
                            return
                        cmd = f"chmod {perm_value} '{file_path}'"
                    else:
                        perm_value = symbol_var.get().strip()
                        if not perm_value:
                            messagebox.showerror("错误", "请输入权限表达式")
                            return
                        cmd = f"chmod {perm_value} '{file_path}'"
                    
                    # 执行chmod命令
                    stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                    stdout.channel.settimeout(10)
                    stderr.channel.settimeout(10)
                    
                    # 等待命令执行完成
                    exit_status = stdout.channel.recv_exit_status()
                    error = stderr.read().decode('utf-8', errors='ignore').strip()
                    
                    if exit_status != 0 or error:
                        # 检查是否是权限错误
                        if "Operation not permitted" in error or "Permission denied" in error:
                            sudo_cmd = f"sudo {cmd}"
                            error_msg = f"设置权限失败: 权限不足\n\n文件路径: {file_path}\n权限值: {perm_value}\n\n可能的原因:\n1. 文件受保护或为系统文件\n2. 当前用户没有修改权限的权限\n3. 需要管理员权限\n\n解决方法:\n您可以在下方的命令执行面板中使用以下命令设置权限:\n\n{sudo_cmd}"
                            messagebox.showerror("权限错误", error_msg)
                            self.output_queue.put(("warning", f"\n[设置权限权限不足]\n文件路径: {file_path}\n权限值: {perm_value}\n\n可以使用以下命令设置权限:\n{sudo_cmd}\n\n"))
                        else:
                            error_msg = f"设置权限失败\n\n文件路径: {file_path}\n权限值: {perm_value}\n\n错误信息: {error if error else '未知错误'}"
                            messagebox.showerror("设置权限失败", error_msg)
                    else:
                        messagebox.showinfo("成功", f"权限设置成功\n\n文件路径: {file_path}\n权限值: {perm_value}")
                        perm_window.destroy()  # 只关闭权限设置窗口
                        browse_path()  # 刷新文件列表，不关闭文件浏览器窗口
                except Exception as e:
                    messagebox.showerror("错误", f"设置权限失败: {str(e)}")
                    # 错误后也不关闭文件浏览器窗口
            
            ttk.Button(btn_frame, text="应用", command=apply_permissions, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=perm_window.destroy, width=12).pack(side=tk.LEFT, padx=5)
            
            # 初始化显示
            update_mode()
            
            # 聚焦到输入框
            if mode_var.get() == "octal":
                octal_entry.focus_set()
                octal_entry.select_range(0, tk.END)
            else:
                symbol_entry.focus_set()
                symbol_entry.select_range(0, tk.END)
        
        def on_double_click(event):
            # Double-click to enter directory or choose file
            item = tree.selection()[0] if tree.selection() else None
            if not item:
                return
            
            values = tree.item(item, "values")
            if values and values[0] == "目录":
                name = tree.item(item, "text")
                current = path_var.get().strip().rstrip('/')
                if current == "/":
                    new_path = f"/{name}"
                else:
                    new_path = f"{current}/{name}"
                path_var.set(new_path)
                browse_path()
            else:
                open_file()
        
        # 绑定事件和按钮
        path_entry.bind('<Return>', lambda e: browse_path())
        ttk.Button(btn_frame, text="刷新", command=browse_path, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="上级目录", command=go_up, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="根目录", command=go_root, width=8).pack(side=tk.LEFT, padx=2)
        tree.bind('<Double-1>', on_double_click)
        
        # 初始加载（延迟执行，确保界面已完全创建）
        browser_window.after(100, browse_path)
    
    def database_manage(self):
        """数据库管理（创建、导出、导入数据库）"""
        if not self.is_connected:
            messagebox.showwarning("提示", "请先连接SSH服务器")
            return
        
        db_window = tk.Toplevel(self.root)
        db_window.title("数据库管理")
        db_window.geometry("700x500")
        db_window.transient(self.root)
        db_window.grab_set()
        
        # 主容器
        main_frame = tk.Frame(db_window, bg="#f5f5f5", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = tk.Label(main_frame, text="数据库管理", 
                              font=("Microsoft YaHei", 16, "bold"), 
                              bg="#f5f5f5", fg="#2c3e50")
        title_label.pack(pady=(0, 20))
        
        # 数据库信息区域
        info_frame = ttk.LabelFrame(main_frame, text="数据库信息", padding="15")
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(info_frame, text="数据库名称:", font=("", 10)).grid(row=0, column=0, sticky=tk.W, pady=8, padx=5)
        db_name_var = tk.StringVar()
        db_name_entry = ttk.Entry(info_frame, textvariable=db_name_var, width=30, font=("", 10))
        db_name_entry.grid(row=0, column=1, pady=8, padx=5, sticky=(tk.W, tk.E))
        
        ttk.Label(info_frame, text="用户名:", font=("", 10)).grid(row=1, column=0, sticky=tk.W, pady=8, padx=5)
        db_user_var = tk.StringVar(value="root")
        db_user_entry = ttk.Entry(info_frame, textvariable=db_user_var, width=30, font=("", 10))
        db_user_entry.grid(row=1, column=1, pady=8, padx=5, sticky=(tk.W, tk.E))
        
        ttk.Label(info_frame, text="密码:", font=("", 10)).grid(row=2, column=0, sticky=tk.W, pady=8, padx=5)
        db_pass_var = tk.StringVar()
        db_pass_entry = ttk.Entry(info_frame, textvariable=db_pass_var, width=30, font=("", 10), show="*")
        db_pass_entry.grid(row=2, column=1, pady=8, padx=5, sticky=(tk.W, tk.E))
        
        info_frame.columnconfigure(1, weight=1)
        
        # 操作按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=15)
        
        def create_database():
            """创建数据库"""
            db_name = db_name_var.get().strip()
            db_user = db_user_var.get().strip()
            db_pass = db_pass_var.get().strip()
            
            if not db_name:
                messagebox.showwarning("提示", "请输入数据库名称")
                return
            
            if not db_user:
                messagebox.showwarning("提示", "请输入数据库用户名")
                return
            
            # 构建MySQL创建数据库命令
            if db_pass:
                cmd = f"mysql -u{db_user} -p{db_pass} -e \"CREATE DATABASE IF NOT EXISTS {db_name} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;\""
            else:
                cmd = f"mysql -u{db_user} -e \"CREATE DATABASE IF NOT EXISTS {db_name} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;\""
            
            # 执行命令
            try:
                stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                output = stdout.read().decode('utf-8', errors='ignore')
                error = stderr.read().decode('utf-8', errors='ignore')
                
                if error and "ERROR" in error.upper():
                    messagebox.showerror("错误", f"创建数据库失败:\n{error}")
                else:
                    messagebox.showinfo("成功", f"数据库 '{db_name}' 创建成功！")
                    self.output_queue.put(("info", f"创建数据库: {db_name}\n"))
                    if output:
                        self.output_queue.put(("output", output))
            except Exception as e:
                messagebox.showerror("错误", f"创建数据库失败: {e}")
        
        def export_database():
            """导出数据库"""
            db_name = db_name_var.get().strip()
            db_user = db_user_var.get().strip()
            db_pass = db_pass_var.get().strip()
            
            if not db_name:
                messagebox.showwarning("提示", "请输入数据库名称")
                return
            
            if not db_user:
                messagebox.showwarning("提示", "请输入数据库用户名")
                return
            
            # 选择保存位置
            filename = filedialog.asksaveasfilename(
                title="保存数据库备份",
                defaultextension=".sql",
                filetypes=[("SQL文件", "*.sql"), ("所有文件", "*.*")]
            )
            
            if not filename:
                return
            
            # 构建mysqldump命令
            if db_pass:
                cmd = f"mysqldump -u{db_user} -p{db_pass} {db_name} > /tmp/{db_name}_backup.sql"
            else:
                cmd = f"mysqldump -u{db_user} {db_name} > /tmp/{db_name}_backup.sql"
            
            # 执行导出命令
            try:
                stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                error = stderr.read().decode('utf-8', errors='ignore')
                
                if error and "ERROR" in error.upper():
                    messagebox.showerror("错误", f"导出数据库失败:\n{error}")
                    return
                
                # 使用SFTP下载文件
                try:
                    sftp = self.client.open_sftp()
                    remote_path = f"/tmp/{db_name}_backup.sql"
                    sftp.get(remote_path, filename)
                    sftp.close()
                    
                    # 删除临时文件
                    self.client.exec_command(f"rm -f {remote_path}", get_pty=False)
                    
                    messagebox.showinfo("成功", f"数据库 '{db_name}' 已导出到:\n{filename}")
                    self.output_queue.put(("info", f"导出数据库: {db_name} -> {filename}\n"))
                except Exception as e:
                    messagebox.showerror("错误", f"下载备份文件失败: {e}")
            except Exception as e:
                messagebox.showerror("错误", f"导出数据库失败: {e}")
        
        def import_database():
            """导入数据库"""
            db_name = db_name_var.get().strip()
            db_user = db_user_var.get().strip()
            db_pass = db_pass_var.get().strip()
            
            if not db_name:
                messagebox.showwarning("提示", "请输入数据库名称")
                return
            
            if not db_user:
                messagebox.showwarning("提示", "请输入数据库用户名")
                return
            
            # 选择要导入的SQL文件
            filename = filedialog.askopenfilename(
                title="选择要导入的SQL文件",
                filetypes=[("SQL文件", "*.sql"), ("所有文件", "*.*")]
            )
            
            if not filename:
                return
            
            if not os.path.exists(filename):
                messagebox.showerror("错误", "文件不存在")
                return
            
            # 确认导入
            if not messagebox.askyesno("确认", f"确定要导入数据库 '{db_name}' 吗？\n这将覆盖现有数据！"):
                return
            
            try:
                # 使用SFTP上传文件
                sftp = self.client.open_sftp()
                remote_path = f"/tmp/{os.path.basename(filename)}"
                sftp.put(filename, remote_path)
                sftp.close()
                
                # 构建mysql导入命令
                if db_pass:
                    cmd = f"mysql -u{db_user} -p{db_pass} {db_name} < {remote_path}"
                else:
                    cmd = f"mysql -u{db_user} {db_name} < {remote_path}"
                
                # 执行导入命令
                stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                error = stderr.read().decode('utf-8', errors='ignore')
                
                # 删除临时文件
                self.client.exec_command(f"rm -f {remote_path}", get_pty=False)
                
                if error and "ERROR" in error.upper():
                    messagebox.showerror("错误", f"导入数据库失败:\n{error}")
                else:
                    messagebox.showinfo("成功", f"数据库 '{db_name}' 导入成功！")
                    self.output_queue.put(("info", f"导入数据库: {filename} -> {db_name}\n"))
            except Exception as e:
                messagebox.showerror("错误", f"导入数据库失败: {e}")
        
        # 按钮样式
        style = ttk.Style()
        style.configure("DB.TButton", padding=10, font=("", 10))
        
        ttk.Button(btn_frame, text="创建数据库", command=create_database, 
                  width=20, style="DB.TButton").pack(side=tk.LEFT, padx=10, pady=10)
        ttk.Button(btn_frame, text="导出数据库", command=export_database, 
                  width=20, style="DB.TButton").pack(side=tk.LEFT, padx=10, pady=10)
        ttk.Button(btn_frame, text="导入数据库", command=import_database, 
                  width=20, style="DB.TButton").pack(side=tk.LEFT, padx=10, pady=10)
        
        # 说明文本
        help_text = """
使用说明：
1. 创建数据库：输入数据库名称、用户名和密码，点击"创建数据库"
2. 导出数据库：输入数据库信息，选择保存位置，点击"导出数据库"
3. 导入数据库：输入数据库信息，选择SQL文件，点击"导入数据库"
注意：导入数据库会覆盖现有数据，请谨慎操作！
        """
        def import_from_server():
            # 从服务器文件导入
            if not self.is_connected:
                messagebox.showwarning("提示", "请先连接SSH服务器")
                import_window.destroy()
                return
            
            import_window.destroy()
            
            # 创建文件选择窗口
            select_window = tk.Toplevel(self.root)
            select_window.title("选择服务器文件")
            select_window.geometry("800x600")
            select_window.transient(self.root)
            select_window.grab_set()
            
            selected_file_path = [None]  # 使用列表以便在嵌套函数中修改
            
            # 路径输入和按钮
            path_frame = ttk.Frame(select_window)
            path_frame.pack(fill=tk.X, padx=5, pady=5)
            
            ttk.Label(path_frame, text="路径:").pack(side=tk.LEFT, padx=5)
            path_var = tk.StringVar(value="/")
            path_entry = ttk.Entry(path_frame, textvariable=path_var, width=50)
            path_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            
            btn_frame = ttk.Frame(path_frame)
            btn_frame.pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="刷新", command=lambda: browse_path(), width=8).pack(side=tk.LEFT, padx=2)
            ttk.Button(btn_frame, text="上级目录", command=lambda: go_up(), width=8).pack(side=tk.LEFT, padx=2)
            ttk.Button(btn_frame, text="根目录", command=lambda: go_root(), width=8).pack(side=tk.LEFT, padx=2)
            
            # 文件列表
            list_frame = ttk.Frame(select_window)
            list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            tree = ttk.Treeview(list_frame, columns=("类型", "大小", "日期"), show="tree headings", height=20)
            tree.heading("#0", text="文件名")
            tree.heading("类型", text="类型")
            tree.heading("大小", text="大小")
            tree.heading("日期", text="修改日期")
            tree.column("#0", width=300)
            tree.column("类型", width=80)
            tree.column("大小", width=100, anchor=tk.E)
            tree.column("日期", width=150)
            
            scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            def browse_path():
                # 浏览指定路径
                path = path_var.get().strip()
                if not path:
                    path = "/"
                
                path = path.rstrip('/')
                if not path:
                    path = "/"
                
                for item in tree.get_children():
                    tree.delete(item)
                
                loading_item = tree.insert("", tk.END, text="(加载中...)", values=("", "", ""), tags=("loading",))
                tree.tag_configure("loading", foreground="gray")
                select_window.update_idletasks()
                
                try:
                    if not self.client or not self.is_connected:
                        tree.delete(loading_item)
                        tree.insert("", tk.END, text="(错误: SSH未连接)", values=("", "", ""), tags=("error",))
                        tree.tag_configure("error", foreground="red")
                        return
                    
                    try:
                        sftp = self.client.open_sftp()
                        files = sftp.listdir_attr(path)
                        sftp.close()
                        
                        tree.delete(loading_item)
                        
                        # 添加父目录项（如果不是根目录）
                        if path != "/":
                            tree.insert("", tk.END, text="..", values=("目录", "", ""), tags=("目录",))
                        
                        # 添加文件和目录
                        for attr in sorted(files, key=lambda x: (not stat.S_ISDIR(x.st_mode), x.filename.lower())):
                            name = attr.filename
                            if name.startswith('.'):
                                continue
                            
                            file_type = "目录" if stat.S_ISDIR(attr.st_mode) else "文件"
                            size = ""
                            date = ""
                            
                            if not stat.S_ISDIR(attr.st_mode):
                                size = f"{attr.st_size:,}" if attr.st_size else "0"
                            
                            if hasattr(attr, 'st_mtime'):
                                from datetime import datetime
                                date = datetime.fromtimestamp(attr.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                            
                            tree.insert("", tk.END, text=name, values=(file_type, size, date), tags=(file_type,))
                        
                        tree.tag_configure("目录", foreground="blue")
                        path_var.set(path)
                        
                    except Exception as sftp_error:
                        # SFTP失败，尝试使用ls命令
                        tree.delete(loading_item)
                        loading_item = tree.insert("", tk.END, text="(使用ls命令加载...)", values=("", "", ""), tags=("loading",))
                        select_window.update_idletasks()
                        
                        cmd = f"cd '{path}' && ls -lah 2>&1"
                        stdin, stdout, stderr = self.client.exec_command(cmd)
                        stdout.channel.settimeout(5)
                        import time
                        time.sleep(0.5)
                        output = stdout.read().decode('utf-8', errors='ignore')
                        error = stderr.read().decode('utf-8', errors='ignore')
                        
                        tree.delete(loading_item)
                        
                        if error and ("No such file" in error or "cannot access" in error):
                            tree.insert("", tk.END, text=f"(错误: {error.strip()})", values=("", "", ""), tags=("error",))
                            tree.tag_configure("error", foreground="red")
                            return
                        
                        lines = output.strip().split('\n')
                        file_count = 0
                        
                        for line in lines[1:]:  # 跳过第一行（总计）
                            line = line.strip()
                            if not line or line.startswith('total'):
                                continue
                            
                            parts = line.split(None, 8)
                            if len(parts) < 9:
                                continue
                            
                            try:
                                permissions = parts[0]
                                name = parts[8] if len(parts) > 8 else parts[-1]
                                
                                if name in ['.', '..']:
                                    continue
                                
                                file_type = "目录" if permissions.startswith('d') else "文件"
                                size = parts[4] if len(parts) > 4 else ""
                                date = f"{parts[5]} {parts[6]} {parts[7]}" if len(parts) > 7 else ""
                                
                                tree.insert("", tk.END, text=name, values=(file_type, size, date), tags=(file_type,))
                                file_count += 1
                            except:
                                continue
                        
                        tree.tag_configure("目录", foreground="blue")
                        path_var.set(path)
                        
                except Exception as e:
                    tree.delete(loading_item)
                    tree.insert("", tk.END, text=f"(错误: {str(e)})", values=("", "", ""), tags=("error",))
                    tree.tag_configure("error", foreground="red")
            
            def go_up():
                # Go to parent directory
                current = path_var.get().strip().rstrip('/')
                if current == "/":
                    return
                parent = os.path.dirname(current)
                if not parent:
                    parent = "/"
                path_var.set(parent)
                browse_path()
            
            def go_root():
                # Go to root directory
                path_var.set("/")
                browse_path()
            
            def get_selected_path():
                # Get currently selected file or directory path
                selection = tree.selection()
                if not selection:
                    return None
                
                item = selection[0]
                values = tree.item(item, "values")
                name = tree.item(item, "text")
                
                if name == "..":
                    return None
                
                current = path_var.get().strip().rstrip('/')
                if current == "/":
                    return f"/{name}"
                else:
                    return f"{current}/{name}"
            
            def on_click(event):
                # Select file on single-click (update selection state)
                item = tree.selection()[0] if tree.selection() else None
                if not item:
                    return
                
                values = tree.item(item, "values")
                name = tree.item(item, "text")
                
                if name == ".." or (values and values[0] == "目录"):
                    selected_file_path[0] = None
                else:
                    # 选择文件
                    file_path = get_selected_path()
                    selected_file_path[0] = file_path
            
            def on_double_click(event):
                # Double-click to enter directory or choose file
                item = tree.selection()[0] if tree.selection() else None
                if not item:
                    return
                
                values = tree.item(item, "values")
                name = tree.item(item, "text")
                
                if name == "..":
                    go_up()
                    return
                
                if values and values[0] == "目录":
                    current = path_var.get().strip().rstrip('/')
                    if current == "/":
                        new_path = f"/{name}"
                    else:
                        new_path = f"{current}/{name}"
                    path_var.set(new_path)
                    browse_path()
                else:
                    # 双击文件直接导入
                    file_path = get_selected_path()
                    if file_path:
                        selected_file_path[0] = file_path
                        select_file()
            
            def select_file():
                # Select file and import
                # 如果没有预先选择，尝试从当前选中项获取
                if not selected_file_path[0]:
                    file_path = get_selected_path()
                    if file_path:
                        values = tree.item(tree.selection()[0], "values")
                        if values and values[0] == "文件":
                            selected_file_path[0] = file_path
                        else:
                            messagebox.showwarning("提示", "请选择一个文件（不是目录）")
                            return
                    else:
                        messagebox.showwarning("提示", "请先选择一个文件")
                        return
                
                file_path = selected_file_path[0]
                select_window.destroy()
                
                try:
                    # 使用SFTP读取文件（更可靠，避免主目录问题）
                    try:
                        sftp = self.client.open_sftp()
                        with sftp.open(file_path, 'r') as f:
                            content = f.read().decode('utf-8', errors='ignore')
                        sftp.close()
                    except IOError as sftp_error:
                        # SFTP失败，尝试使用cat命令（使用绝对路径，避免主目录问题）
                        # 使用 get_pty=False 避免尝试切换到主目录
                        cmd = f"cat '{file_path}' 2>&1"
                        stdin, stdout, stderr = self.client.exec_command(cmd, get_pty=False)
                        stdout.channel.settimeout(10)
                        import time
                        time.sleep(0.5)
                        
                        # 等待命令完成
                        exit_status = stdout.channel.recv_exit_status()
                        content = stdout.read().decode('utf-8', errors='ignore')
                        error_output = stderr.read().decode('utf-8', errors='ignore')
                        
                        if exit_status != 0 or (error_output and ("No such file" in error_output or "cannot access" in error_output)):
                            error_msg = error_output if error_output else f"命令执行失败，退出码: {exit_status}"
                            messagebox.showerror("错误", f"文件不存在或无法读取: {file_path}\n错误信息: {error_msg}")
                            return
                        
                        if not content or not content.strip():
                            messagebox.showwarning("提示", f"文件为空: {file_path}")
                            return
                    
                    # 导入物品ID（文件内容作为物品ID）
                    # 支持多行，每行一个物品ID
                    lines = content.strip().split('\n')
                    item_ids = [line.strip() for line in lines if line.strip()]
                    
                    if item_ids:
                        # 确保使用当前IP的历史记录（按IP独立存储）
                        # 获取当前IP键
                        current_ip_key = self.get_current_ip_key()
                        
                        # 确保历史记录列表已初始化
                        if not hasattr(self, 'item_ids_history'):
                            self.item_ids_history = []
                        
                        # 在导入前，先加载当前IP的历史记录，确保数据同步
                        # 这样可以确保导入到正确的IP下
                        current_ip_key = self.get_current_ip_key()
                        print(f"准备导入物品ID到IP: {current_ip_key}")
                        try:
                            # 重新加载当前IP的数据，确保 item_ids_history 是当前IP的数据
                            # 先清空内存，然后重新加载，确保数据是对应当前IP的
                            self.item_ids_history = []
                            self.item_ids_all = []
                            self.load_item_ids()
                            print(f"已加载IP {current_ip_key} 的物品ID，共 {len(self.item_ids_history)} 个")
                        except Exception as e:
                            print(f"加载当前IP物品ID时出错: {e}")
                            import traceback
                            traceback.print_exc()
                            # 如果加载失败，确保 item_ids_history 是列表
                            if not isinstance(self.item_ids_history, list):
                                self.item_ids_history = []
                                self.item_ids_all = []
                        
                        # 第一个物品ID设置为当前值（显示完整内容）
                        first_id_full = item_ids[0]
                        # 标准化第一个物品ID用于显示
                        first_id_normalized = self._normalize_item_id(first_id_full)
                        if first_id_normalized:
                            display_str = self._format_item_id_for_display(first_id_normalized)
                            first_display = display_str if display_str else first_id_full
                        else:
                            first_display = first_id_full
                        
                        # 记录导入前的历史记录（用于去重和统计）
                        # 确保使用当前IP的数据
                        if not isinstance(self.item_ids_history, list):
                            self.item_ids_history = []
                        
                        # 获取导入前已有的物品ID集合（用于快速查找）
                        existing_item_ids = set()
                        existing_full_items = set()  # 存储完整的物品ID字符串
                        for item in self.item_ids_history:
                            # 使用原始格式，不做标准化
                            item_str = str(item).strip() if item else ""
                            if item_str:
                                existing_full_items.add(item_str)
                                # 提取ID部分用于去重
                                item_id_only = self.extract_item_id(item_str)
                                if item_id_only and item_id_only.strip():
                                    existing_item_ids.add(item_id_only.strip())
                        
                        # 统计实际导入的数量（只统计真正新增的，不包括已存在的）
                        imported_count = 0
                        # 用于存储本次导入要添加的物品ID（去重后）
                        items_to_add = []
                        # 用于跟踪本次导入中已经处理过的物品ID（避免重复）
                        processed_ids_in_import = set()
                        
                        # 遍历所有要导入的物品ID（完全不过滤，全部导入）
                        for item_id_full in item_ids:
                            # 跳过空行
                            if not item_id_full or not item_id_full.strip():
                                continue
                            
                            # 保留原始格式，只做基本的trim处理（不做标准化，避免丢失信息）
                            item_id_normalized = item_id_full.strip()
                            
                            if not item_id_normalized:
                                continue
                            
                            # 导入时完全不做任何验证和过滤，允许所有非空内容导入
                            # 这样可以确保导入完整，不会因为任何验证而丢失数据
                            
                            # 提取ID部分（仅用于去重判断，不影响导入）
                            # 使用原始格式提取，不做标准化
                            item_id_only = None
                            try:
                                # 尝试提取ID部分
                                if ';' in item_id_normalized:
                                    item_id_only = item_id_normalized.split(';')[0].strip()
                                elif ' - ' in item_id_normalized:
                                    item_id_only = item_id_normalized.split(' - ')[0].strip()
                                else:
                                    # 如果没有分隔符，尝试提取数字部分
                                    import re
                                    numbers = re.findall(r'^\d+', item_id_normalized)
                                    if numbers:
                                        item_id_only = numbers[0]
                                    else:
                                        # 如果没有数字，使用完整内容作为ID
                                        item_id_only = item_id_normalized
                            except:
                                # 如果提取失败，使用完整内容
                                item_id_only = item_id_normalized
                            
                            # 检查是否是新的物品ID（不在已有记录中）
                            if item_id_only:
                                # 检查是否在本次导入中已经处理过（避免重复添加）
                                if item_id_only in processed_ids_in_import:
                                    # 本次导入中已经处理过，跳过（不重复添加）
                                    continue
                                
                                processed_ids_in_import.add(item_id_only)
                                
                                # 检查是否在历史记录中已存在
                                if item_id_only not in existing_item_ids:
                                    # 是新物品ID，添加到导入列表
                                    items_to_add.append(item_id_normalized)
                                    existing_item_ids.add(item_id_only)  # 更新集合
                                    imported_count += 1  # 统计新增的
                                else:
                                    # 已存在，更新位置（移到最前面），也添加到列表
                                    # 移除已存在的项
                                    self.item_ids_history = [item for item in self.item_ids_history 
                                                           if self.extract_item_id(item) != item_id_only]
                                    # 添加到导入列表（用于更新位置）
                                    items_to_add.append(item_id_normalized)
                            else:
                                # 没有提取到ID部分，使用完整内容进行去重
                                if item_id_normalized not in existing_full_items:
                                    # 不存在，添加到导入列表
                                    if item_id_normalized not in processed_ids_in_import:
                                        items_to_add.append(item_id_normalized)
                                        existing_full_items.add(item_id_normalized)
                                        processed_ids_in_import.add(item_id_normalized)
                                        imported_count += 1
                                else:
                                    # 已存在，更新位置（移到最前面）
                                    if item_id_normalized in self.item_ids_history:
                                        self.item_ids_history.remove(item_id_normalized)
                                    items_to_add.append(item_id_normalized)
                        
                        # 将所有物品ID添加到历史记录的最前面
                        for item in reversed(items_to_add):  # 反向插入，保持原始顺序
                            self.item_ids_history.insert(0, item)
                        # 移除数量限制，允许导入所有物品ID
                        # 不再限制历史记录数量，确保所有导入的物品ID都能被保存
                        # 更新下拉框
                        if hasattr(self, 'item_id_combo'):
                            # 更新所有物品ID列表
                            self.item_ids_all = self.item_ids_history.copy()
                            
                            # 强制刷新下拉框，显示所有导入的物品ID
                            def update_dropdown():
                                try:
                                    # 确保数据已更新
                                    if hasattr(self, 'item_ids_all') and self.item_ids_all:
                                        # 设置第一个物品ID到输入框
                                        self.item_id_var.set(first_display)
                                        # 强制刷新下拉框，显示所有物品（忽略输入框内容）
                                        self.root.after(50, lambda: self.filter_item_ids(force_show_all=True))
                                except Exception as e:
                                    print(f"更新下拉框时出错: {e}")
                            self.root.after(100, update_dropdown)
                        # 保存到文件（按当前IP保存，确保每个IP独立存储）
                        # 确保保存到当前IP对应的键下
                        try:
                            self.save_item_ids()
                            # 验证保存是否成功（可选，用于调试）
                            print(f"物品ID已保存到IP: {self.get_current_ip_key()}, 共 {len(self.item_ids_history)} 个物品ID")
                        except Exception as e:
                            print(f"保存物品ID到文件时出错: {e}")
                            import traceback
                            traceback.print_exc()
                        
                        first_id_only = self.extract_item_id(first_id_full) if first_id_full else ""
                        total_in_file = len(item_ids)
                        total_in_history = len(self.item_ids_history) if hasattr(self, 'item_ids_history') else 0
                        
                        if imported_count == 0:
                            if total_in_file > 0:
                                messagebox.showwarning("提示", f"文件中有 {total_in_file} 个物品ID，但没有新物品ID被导入（可能都已存在）")
                            else:
                                messagebox.showwarning("提示", "文件为空或没有有效的物品ID")
                        elif imported_count == 1:
                            messagebox.showinfo("提示", f"物品ID导入成功！\n文件中共有 {total_in_file} 个物品ID\n新增了 {imported_count} 个物品ID\n当前物品ID: {first_id_only}\n历史记录中共有 {total_in_history} 个物品ID")
                        else:
                            messagebox.showinfo("提示", f"物品ID导入成功！\n文件中共有 {total_in_file} 个物品ID\n新增了 {imported_count} 个物品ID\n当前物品ID: {first_id_only}\n历史记录中共有 {total_in_history} 个物品ID")
                    else:
                        messagebox.showwarning("提示", "文件内容为空")
                except Exception as e:
                    import traceback
                    error_detail = traceback.format_exc()
                    messagebox.showerror("错误", f"导入配置失败: {e}\n\n详细信息:\n{error_detail}")
            
            # 按钮框架
            button_frame = ttk.Frame(select_window)
            button_frame.pack(fill=tk.X, padx=5, pady=5)
            
            ttk.Button(button_frame, text="选择文件", command=select_file, width=15).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="取消", command=select_window.destroy, width=15).pack(side=tk.LEFT, padx=5)
            
            # 绑定事件
            path_entry.bind('<Return>', lambda e: browse_path())
            tree.bind('<Button-1>', on_click)
            tree.bind('<Double-1>', on_double_click)
            
            # 初始加载
            select_window.after(100, browse_path)
        
        ttk.Button(btn_frame, text="从本地文件导入", command=import_from_local, width=20).grid(row=0, column=0, padx=10, pady=10)
        ttk.Button(btn_frame, text="从服务器导入", command=import_from_server, width=20).grid(row=0, column=1, padx=10, pady=10)
# 授权码由授权服务器（母机）生成。
# 如果没有授权码，请联系管理员获取。
#
# 管理员QQ：3593075503（流浪）

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = SSHToolGUI(root)
        root.mainloop()
    except Exception as e:
        # 兜底错误提示，避免静默失败
        import traceback
        error_msg = f"程序启动出错：{e}\n\n详细错误信息：\n{traceback.format_exc()}"
        
        # 尝试多种方式显示错误
        error_shown = False
        
        # 方式1: 尝试使用messagebox
        try:
            messagebox.showerror("启动失败", error_msg)
            error_shown = True
        except Exception:
            pass
        
        # 方式2: 尝试使用Windows消息框
        if not error_shown:
            try:
                import ctypes
                ctypes.windll.user32.MessageBoxW(
                    None,
                    error_msg,
                    "启动失败",
                    0x00000010  # MB_ICONERROR
                )
                error_shown = True
            except Exception:
                pass
        
        # 方式3: 打印到控制台并等待用户输入
        if not error_shown:
            print("\n" + "=" * 80)
            print("程序启动失败！")
            print("=" * 80)
            print(error_msg)
            print("=" * 80)
            try:
                input("\n按回车键退出...")
            except Exception:
                import time
                time.sleep(5)  # 等待5秒让用户看到错误信息